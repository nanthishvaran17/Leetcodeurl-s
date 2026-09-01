import io
import datetime
from typing import Dict, Any, List, Optional
from zoneinfo import ZoneInfo
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from sqlalchemy.orm import Session
from backend.models import Student, Department, Contest, ContestParticipationRecord, LeetCodeProfileStats, LeetCodeProfile
from backend.time_utils import IST, now_ist, format_ist
from backend.logger import logger

class CollegeReportGenerator:
    """
    Production-Grade College LeetCode Performance Report Generator.
    Produces the exact institutional format for NANDHA ENGINEERING COLLEGE.
    
    Format 1: Management Summary (College Admin View)
    Format 2: Detailed Student List (Excel/CSV)
    """

    def __init__(self, db_session_factory):
        self.db_session_factory = db_session_factory
        self.ist = IST

    def _map_year_to_batch(self, year_val: Any) -> str:
        """Helper to map academic year or year level to graduation batch string."""
        s = str(year_val or "").strip().upper()
        if "2023" in s or s in ("III", "3", "3RD", "THIRD"):
            return "2023 - 2027"
        elif "2024" in s or s in ("II", "2", "2ND", "SECOND"):
            return "2024 - 2028"
        elif "2025" in s or s in ("I", "1", "1ST", "FIRST"):
            return "2025 - 2029"
        elif "2022" in s or s in ("IV", "4", "4TH", "FINAL"):
            return "2022 - 2026"
        elif s.isdigit() and len(s) == 4:
            y = int(s)
            return f"{y} - {y + 4}"
        return "2023 - 2027"

    async def get_participants(self, db: Session, contest_id: int) -> List[Dict[str, Any]]:
        """Extracts unified participant records with profiles, contest results, and evidence."""
        students = db.query(Student).filter((Student.is_active == True) | (Student.is_active.is_(None))).all()
        departments = {d.id: d for d in db.query(Department).all()}
        part_records = {p.student_id: p for p in db.query(ContestParticipationRecord).filter_by(contest_id=contest_id).all()}
        stats_map = {s.student_id: s for s in db.query(LeetCodeProfileStats).all()}
        profiles_map = {p.student_id: p for p in db.query(LeetCodeProfile).all()}

        results = []
        for s in students:
            dept = departments.get(s.department_id)
            dept_name = dept.name if dept else "Computer Science and Engineering (IoT)"
            dept_code = dept.code if dept else "CSE"
            part = part_records.get(s.id)
            stats = stats_map.get(s.id)
            prof = profiles_map.get(s.id)

            username = s.username or (part.leetcode_username if part else None) or (prof.canonical_username if prof else None) or ""
            profile_url = f"https://leetcode.com/u/{username}/" if username else ""
            profile_total_solved = (stats.total_solved if stats and stats.total_solved is not None else 0)

            participation_status = part.participation_status if part else "NOT_VERIFIED"
            contest_solved = part.solved_count if part and part.solved_count is not None else 0
            score = part.score if part and part.score is not None else 0
            rank = part.rank if part and part.rank is not None else None
            rating = (part.rating if part and part.rating is not None else (stats.contest_rating if stats and stats.contest_rating else 0)) or 0
            evidence_source = part.evidence_source if part and part.evidence_source else "none"

            # Parse question status breakdown
            q_data = getattr(part, "question_data", None) if part else None
            q1, q2, q3, q4 = "-", "-", "-", "-"
            if isinstance(q_data, list) and len(q_data) >= 4:
                q1 = q_data[0].get("status", "-") if isinstance(q_data[0], dict) else str(q_data[0])
                q2 = q_data[1].get("status", "-") if isinstance(q_data[1], dict) else str(q_data[1])
                q3 = q_data[2].get("status", "-") if isinstance(q_data[2], dict) else str(q_data[2])
                q4 = q_data[3].get("status", "-") if isinstance(q_data[3], dict) else str(q_data[3])
            elif participation_status == "ACTUAL" and contest_solved > 0:
                q1 = "AC" if contest_solved >= 1 else "WA"
                q2 = "AC" if contest_solved >= 2 else "WA"
                q3 = "AC" if contest_solved >= 3 else "WA"
                q4 = "AC" if contest_solved >= 4 else "WA"
            elif participation_status == "VIRTUAL" and contest_solved > 0:
                q1 = "AC" if contest_solved >= 1 else "WA"
                q2 = "AC" if contest_solved >= 2 else "WA"
                q3 = "AC" if contest_solved >= 3 else "WA"
                q4 = "AC" if contest_solved >= 4 else "WA"

            results.append({
                "student_id": s.id,
                "name": s.name,
                "reg_no": s.reg_no or "",
                "department": dept_name,
                "dept_code": dept_code,
                "year_level": s.year_level or "III",
                "batch": self._map_year_to_batch(s.year_level),
                "leetcode_username": username,
                "profile_url": profile_url,
                "profile_total_solved": profile_total_solved,
                "participation_status": participation_status,
                "contest_solved": contest_solved,
                "score": score,
                "rank": rank if rank is not None else "-",
                "rating": round(rating, 1) if rating else 0,
                "evidence_source": evidence_source,
                "q1": q1,
                "q2": q2,
                "q3": q3,
                "q4": q4,
            })

        # Sort by Batch desc, then Rank asc
        results.sort(key=lambda x: (x["batch"], x["rank"] if isinstance(x["rank"], int) else 999999))
        return results

    def generate_summary_data(self, participants: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Builds aggregated batch-wise Management Summary data."""
        batch_order = ["2023 - 2027", "2024 - 2028", "2025 - 2029", "2022 - 2026"]
        batch_data: Dict[str, Dict[str, Any]] = {}

        for b in batch_order:
            batch_data[b] = {
                "above_500": 0,
                "250_500": 0,
                "less_250": 0,
                "less_100": 0,
                "not_started": 0,
                "4Q": 0,
                "3Q": 0,
                "2Q": 0,
                "1Q": 0,
                "rating_gt_1500": 0,
                "rank_lt_20000": 0,
            }

        for p in participants:
            b = p["batch"]
            if b not in batch_data:
                batch_data[b] = {
                    "above_500": 0, "250_500": 0, "less_250": 0, "less_100": 0, "not_started": 0,
                    "4Q": 0, "3Q": 0, "2Q": 0, "1Q": 0, "rating_gt_1500": 0, "rank_lt_20000": 0,
                }

            # Problems Solved breakdown
            solved = p["profile_total_solved"]
            if solved > 500:
                batch_data[b]["above_500"] += 1
            elif solved >= 250:
                batch_data[b]["250_500"] += 1
            elif solved >= 100:
                batch_data[b]["less_250"] += 1
            elif solved > 0:
                batch_data[b]["less_100"] += 1
            else:
                batch_data[b]["not_started"] += 1

            # Contest Performance
            c_solved = p["contest_solved"]
            if c_solved >= 4:
                batch_data[b]["4Q"] += 1
            elif c_solved == 3:
                batch_data[b]["3Q"] += 1
            elif c_solved == 2:
                batch_data[b]["2Q"] += 1
            elif c_solved == 1:
                batch_data[b]["1Q"] += 1

            # Contest Rating & Ranking
            if (p.get("rating") or 0) > 1500:
                batch_data[b]["rating_gt_1500"] += 1

            rank_val = p.get("rank")
            if isinstance(rank_val, (int, float)) and rank_val < 20000:
                batch_data[b]["rank_lt_20000"] += 1

        rows = []
        for b in batch_order:
            if b in batch_data:
                d = batch_data[b]
                rows.append({
                    "Batch": b,
                    "Above 500": d["above_500"],
                    "250-500": d["250_500"],
                    "Less than 250": d["less_250"],
                    "Less than 100": d["less_100"],
                    "Not yet started": d["not_started"],
                    "4Q": d["4Q"],
                    "3Q": d["3Q"],
                    "2Q": d["2Q"],
                    "1Q": d["1Q"],
                    "Rating: >1500": d["rating_gt_1500"],
                    "Ranking: <20000": d["rank_lt_20000"],
                })

        return rows

    def generate_details_data(self, participants: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Builds individual student detailed list conforming to Format 2."""
        rows = []
        for idx, p in enumerate(participants, 1):
            status_raw = str(p.get("participation_status", "")).upper()
            if status_raw in ("ACTUAL", "PUBLIC_ATTENDED", "ATTENDED"):
                status_display = "ACTUAL"
            elif status_raw in ("VIRTUAL", "VIRTUAL_ATTENDED"):
                status_display = "VIRTUAL"
            else:
                status_display = "NOT_ATTENDED"

            rows.append({
                "S.No": idx,
                "Register No": p.get("reg_no", ""),
                "Student Name": p.get("name", ""),
                "Dept": p.get("dept_code", "CSE"),
                "Year": p.get("year_level", ""),
                "Status": status_display,
                "Q1": p.get("q1", "-"),
                "Q2": p.get("q2", "-"),
                "Q3": p.get("q3", "-"),
                "Q4": p.get("q4", "-"),
                "Contest Solved": p.get("contest_solved", 0),
                "Score": p.get("score", 0),
                "Rank": p.get("rank", "-"),
                "Rating": p.get("rating", 0),
                "Profile Total Solved": p.get("profile_total_solved", 0),
                "Username": p.get("leetcode_username", ""),
                "Profile URL": p.get("profile_url", ""),
                "Source": p.get("evidence_source", "none"),
            })

        return rows

    def build_excel_workbook(
        self,
        summary_rows: List[Dict[str, Any]],
        details_rows: List[Dict[str, Any]],
        contest_title: str,
        department_name: str = "Department of Computer Science and Engineering (IoT)",
        report_date_str: Optional[str] = None
    ) -> Workbook:
        """Constructs and styles the official multi-sheet Excel report with openpyxl."""
        wb = Workbook()
        date_display = report_date_str or now_ist().strftime("%d.%m.%Y")

        # ── Sheet 1: Management Summary ─────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Management Summary"

        # Theme Colors
        navy_header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
        blue_sub_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        light_gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        thin_side = Side(border_style="thin", color="D3D3D3")
        grid_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        # Header Titles
        ws1.merge_cells("A1:L1")
        ws1["A1"] = "NANDHA ENGINEERING COLLEGE, ERODE - 638 052."
        ws1["A1"].font = Font(name="Arial", size=14, bold=True, color="002060")
        ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")

        ws1.merge_cells("A2:L2")
        ws1["A2"] = department_name
        ws1["A2"].font = Font(name="Arial", size=12, bold=True, color="333333")
        ws1["A2"].alignment = Alignment(horizontal="center", vertical="center")

        ws1.merge_cells("A3:L3")
        ws1["A3"] = f"Date: {date_display}"
        ws1["A3"].font = Font(name="Arial", size=10, bold=False, color="555555")
        ws1["A3"].alignment = Alignment(horizontal="center", vertical="center")

        ws1.merge_cells("A5:L5")
        ws1["A5"] = f"Leetcode Performance - Weekly Report ({contest_title})"
        ws1["A5"].font = Font(name="Arial", size=13, bold=True, color="002060")
        ws1["A5"].alignment = Alignment(horizontal="center", vertical="center")

        ws1.merge_cells("A6:L6")
        ws1["A6"] = "Name & Designation of the Academic Coordinator: Dr. S. Sivakumar"
        ws1["A6"].font = Font(name="Arial", size=10, italic=True, color="444444")
        ws1["A6"].alignment = Alignment(horizontal="center", vertical="center")

        # Two-tier column headers (Row 8 & Row 9)
        ws1.merge_cells("A8:A9")
        ws1["A8"] = "Batch"
        ws1["A8"].font = Font(bold=True, color="FFFFFF")
        ws1["A8"].fill = navy_header_fill
        ws1["A8"].alignment = Alignment(horizontal="center", vertical="center")

        ws1.merge_cells("B8:F8")
        ws1["B8"] = "Number of Problems Solved"
        ws1["B8"].font = Font(bold=True, color="FFFFFF")
        ws1["B8"].fill = navy_header_fill
        ws1["B8"].alignment = Alignment(horizontal="center", vertical="center")

        ws1.merge_cells("G8:J8")
        ws1["G8"] = "Weekly Contest Attended"
        ws1["G8"].font = Font(bold=True, color="FFFFFF")
        ws1["G8"].fill = navy_header_fill
        ws1["G8"].alignment = Alignment(horizontal="center", vertical="center")

        ws1.merge_cells("K8:L8")
        ws1["K8"] = "Leetcode Contest Rating"
        ws1["K8"].font = Font(bold=True, color="FFFFFF")
        ws1["K8"].fill = navy_header_fill
        ws1["K8"].alignment = Alignment(horizontal="center", vertical="center")

        sub_headers = [
            "Above 500", "250-500", "Less than 250", "Less than 100", "Not yet started",
            "4Q", "3Q", "2Q", "1Q",
            "Rating: >1500", "Ranking: <20000"
        ]
        for col_idx, sub in enumerate(sub_headers, start=2):
            cell = ws1.cell(row=9, column=col_idx, value=sub)
            cell.font = Font(bold=True, color="FFFFFF", size=9)
            cell.fill = blue_sub_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Write Data rows starting from Row 10
        cur_row = 10
        for r in summary_rows:
            row_vals = [
                r["Batch"], r["Above 500"], r["250-500"], r["Less than 250"], r["Less than 100"], r["Not yet started"],
                r["4Q"], r["3Q"], r["2Q"], r["1Q"], r["Rating: >1500"], r["Ranking: <20000"]
            ]
            for col_idx, val in enumerate(row_vals, start=1):
                cell = ws1.cell(row=cur_row, column=col_idx, value=val)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = grid_border
                if col_idx == 1:
                    cell.font = Font(bold=True)
            cur_row += 1

        # Summary Total Row
        ws1.cell(row=cur_row, column=1, value="Total").font = Font(bold=True)
        ws1.cell(row=cur_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws1.cell(row=cur_row, column=1).fill = light_gray_fill
        ws1.cell(row=cur_row, column=1).border = grid_border

        for col_idx in range(2, 13):
            col_letter = get_column_letter(col_idx)
            cell = ws1.cell(row=cur_row, column=col_idx)
            cell.value = f"=SUM({col_letter}10:{col_letter}{cur_row-1})"
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = light_gray_fill
            cell.border = grid_border

        # ── Sheet 2: Student Details ─────────────────────────────────────────────
        ws2 = wb.create_sheet(title="Student Details")

        # Write header
        headers = [
            "S.No", "Register No", "Student Name", "Dept", "Year", "Status",
            "Q1", "Q2", "Q3", "Q4", "Contest Solved", "Score", "Rank", "Rating",
            "Profile Total Solved", "Username", "Profile URL", "Source"
        ]
        for col_idx, h in enumerate(headers, start=1):
            cell = ws2.cell(row=1, column=col_idx, value=h)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = navy_header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write student rows
        for r_idx, r in enumerate(details_rows, start=2):
            for col_idx, h in enumerate(headers, start=1):
                val = r.get(h, "")
                cell = ws2.cell(row=r_idx, column=col_idx, value=val)
                cell.border = grid_border
                if h in ("Status", "Q1", "Q2", "Q3", "Q4", "Contest Solved", "Score", "Rank", "Rating", "Profile Total Solved", "Dept", "Year"):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        # Auto-adjust column widths for both sheets
        for ws in [ws1, ws2]:
            for col in ws.columns:
                max_len = 0
                first_cell = col[0]
                col_letter = get_column_letter(first_cell.column)
                for cell in col:
                    if hasattr(cell, "coordinate") and cell.coordinate in ws.merged_cells:
                        continue
                    if hasattr(cell, "value") and cell.value is not None:
                        val_str = str(cell.value)
                        if len(val_str) > max_len and len(val_str) < 60:
                            max_len = len(val_str)
                ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 40)

        return wb

    async def generate_complete_report(
        self,
        contest_id: int,
        department_name: str = "Department of Computer Science and Engineering (IoT)"
    ) -> Dict[str, Any]:
        """Orchestrates extraction, aggregation, Excel synthesis, and in-memory export."""
        db = self.db_session_factory()
        try:
            contest = db.query(Contest).filter_by(id=contest_id).first()
            contest_title = contest.contest_title if contest else f"Contest {contest_id}"
            contest_slug = contest.contest_slug if contest else f"contest-{contest_id}"

            participants = await self.get_participants(db, contest_id)
            summary_rows = self.generate_summary_data(participants)
            details_rows = self.generate_details_data(participants)

            wb = self.build_excel_workbook(
                summary_rows=summary_rows,
                details_rows=details_rows,
                contest_title=contest_title,
                department_name=department_name
            )

            # Export Excel to bytes
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_bytes = excel_buffer.getvalue()

            # Generate HTML email body
            actual_cnt = sum(1 for p in participants if p["participation_status"] in ("ACTUAL", "PUBLIC_ATTENDED", "ATTENDED"))
            virtual_cnt = sum(1 for p in participants if p["participation_status"] in ("VIRTUAL", "VIRTUAL_ATTENDED"))
            not_attended_cnt = len(participants) - actual_cnt - virtual_cnt

            # Extract top rankers
            ranked_students = [p for p in participants if isinstance(p.get("rank"), int)]
            ranked_students.sort(key=lambda x: x["rank"])
            top_10 = ranked_students[:10]

            email_html = self._create_college_email_html(
                contest_title=contest_title,
                department_name=department_name,
                total_students=len(participants),
                actual_count=actual_cnt,
                virtual_count=virtual_cnt,
                not_attended_count=not_attended_cnt,
                summary_rows=summary_rows,
                top_rankers=top_10,
                report_date=now_ist().strftime("%d.%m.%Y")
            )

            return {
                "contest_id": contest_id,
                "contest_title": contest_title,
                "contest_slug": contest_slug,
                "excel_bytes": excel_bytes,
                "filename": f"Nandha_LeetCode_Weekly_Report_{contest_slug}_{now_ist().strftime('%Y%m%d')}.xlsx",
                "email_html": email_html,
                "total_students": len(participants),
                "actual_count": actual_cnt,
                "virtual_count": virtual_cnt,
                "not_attended_count": not_attended_cnt,
                "top_rankers": top_10
            }
        finally:
            db.close()

    def _create_college_email_html(
        self,
        contest_title: str,
        department_name: str,
        total_students: int,
        actual_count: int,
        virtual_count: int,
        not_attended_count: int,
        summary_rows: List[Dict[str, Any]],
        top_rankers: List[Dict[str, Any]],
        report_date: str
    ) -> str:
        """Constructs responsive HTML email matching Nandha Engineering College specification."""
        summary_table_rows = ""
        for r in summary_rows:
            summary_table_rows += f"""
            <tr>
                <td style="padding: 8px 12px; font-weight: bold; border: 1px solid #ddd; text-align: center;">{r['Batch']}</td>
                <td style="padding: 8px; border: 1px solid #ddd; text-align: center;">{r['Above 500']}</td>
                <td style="padding: 8px; border: 1px solid #ddd; text-align: center;">{r['250-500']}</td>
                <td style="padding: 8px; border: 1px solid #ddd; text-align: center;">{r['Less than 250']}</td>
                <td style="padding: 8px; border: 1px solid #ddd; text-align: center;">{r['Less than 100']}</td>
                <td style="padding: 8px; border: 1px solid #ddd; text-align: center;">{r['Not yet started']}</td>
                <td style="padding: 8px; border: 1px solid #ddd; text-align: center; font-weight: bold; color: #0066CC;">{r['4Q']}</td>
                <td style="padding: 8px; border: 1px solid #ddd; text-align: center;">{r['3Q']}</td>
                <td style="padding: 8px; border: 1px solid #ddd; text-align: center;">{r['2Q']}</td>
                <td style="padding: 8px; border: 1px solid #ddd; text-align: center;">{r['1Q']}</td>
                <td style="padding: 8px; border: 1px solid #ddd; text-align: center; font-weight: bold; color: #28a745;">{r['Rating: >1500']}</td>
                <td style="padding: 8px; border: 1px solid #ddd; text-align: center;">{r['Ranking: <20000']}</td>
            </tr>
            """

        top_rankers_rows = ""
        for idx, s in enumerate(top_rankers, 1):
            top_rankers_rows += f"""
            <tr>
                <td style="padding: 8px; border: 1px solid #ddd; text-align: center;">{idx}</td>
                <td style="padding: 8px; border: 1px solid #ddd;">{s.get('name', '')}</td>
                <td style="padding: 8px; border: 1px solid #ddd; text-align: center;">{s.get('reg_no', '')}</td>
                <td style="padding: 8px; border: 1px solid #ddd; text-align: center; font-weight: bold; color: #002060;">#{s.get('rank', '-')}</td>
                <td style="padding: 8px; border: 1px solid #ddd; text-align: center;">{s.get('contest_solved', 0)} / 4</td>
                <td style="padding: 8px; border: 1px solid #ddd; text-align: center; font-weight: bold;">{s.get('score', 0)}</td>
                <td style="padding: 8px; border: 1px solid #ddd; text-align: center; color: #28a745;">{s.get('rating', 0)}</td>
            </tr>
            """

        if not top_rankers_rows:
            top_rankers_rows = '<tr><td colspan="7" style="padding: 12px; text-align: center; color: #777;">No participants attended live contest yet.</td></tr>'

        return f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <style>
                body {{ font-family: Arial, Helvetica, sans-serif; background: #f4f6f9; margin: 0; padding: 20px; color: #333; }}
                .container {{ max-width: 900px; margin: 0 auto; background: #ffffff; border-radius: 12px; overflow: hidden; box-shadow: 0 4px 16px rgba(0,0,0,0.08); border: 1px solid #e1e4e8; }}
                .header {{ background: #002060; color: #ffffff; padding: 28px 20px; text-align: center; }}
                .header h1 {{ margin: 0; font-size: 20px; letter-spacing: 0.5px; font-weight: bold; }}
                .header h2 {{ margin: 6px 0 0 0; font-size: 15px; font-weight: normal; color: #d0e0ff; }}
                .header p {{ margin: 6px 0 0 0; font-size: 13px; color: #a0c0ff; }}
                .content {{ padding: 24px; }}
                .kpi-grid {{ display: flex; gap: 12px; margin: 16px 0 24px 0; }}
                .kpi-card {{ flex: 1; background: #f8fafc; border-radius: 8px; padding: 14px; text-align: center; border: 1px solid #e2e8f0; }}
                .kpi-num {{ font-size: 24px; font-weight: bold; margin-bottom: 4px; }}
                .kpi-label {{ font-size: 11px; text-transform: uppercase; color: #64748b; font-weight: bold; letter-spacing: 0.5px; }}
                table {{ width: 100%; border-collapse: collapse; margin-top: 10px; font-size: 12px; }}
                th {{ background: #0066CC; color: white; padding: 10px 8px; font-weight: 600; border: 1px solid #0052a3; }}
                .section-title {{ font-size: 15px; font-weight: bold; color: #002060; margin: 24px 0 10px 0; border-bottom: 2px solid #002060; padding-bottom: 4px; display: flex; justify-content: space-between; }}
                .footer {{ background: #f8fafc; padding: 16px 24px; font-size: 12px; color: #64748b; text-align: center; border-top: 1px solid #e2e8f0; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>NANDHA ENGINEERING COLLEGE, ERODE - 638 052.</h1>
                    <h2>{department_name}</h2>
                    <p>Leetcode Performance - Weekly Report ({contest_title}) • Date: {report_date}</p>
                    <p style="font-size: 11px; margin-top: 4px; color: #ffffff;">Name & Designation of Academic Coordinator: <strong>Dr. S. Sivakumar</strong></p>
                </div>

                <div class="content">
                    <div class="kpi-grid">
                        <div class="kpi-card">
                            <div class="kpi-num" style="color: #002060;">{total_students}</div>
                            <div class="kpi-label">Total Roster</div>
                        </div>
                        <div class="kpi-card">
                            <div class="kpi-num" style="color: #16a34a;">{actual_count}</div>
                            <div class="kpi-label">ACTUAL Attended</div>
                        </div>
                        <div class="kpi-card">
                            <div class="kpi-num" style="color: #d97706;">{virtual_count}</div>
                            <div class="kpi-label">VIRTUAL</div>
                        </div>
                        <div class="kpi-card">
                            <div class="kpi-num" style="color: #dc2626;">{not_attended_count}</div>
                            <div class="kpi-label">Not Attended</div>
                        </div>
                    </div>

                    <div class="section-title">
                        <span>📊 Batch-wise Performance Summary</span>
                    </div>
                    <table>
                        <thead>
                            <tr>
                                <th rowspan="2">Batch</th>
                                <th colspan="5">Problems Solved</th>
                                <th colspan="4">Contest Attended</th>
                                <th colspan="2">Rating / Ranking</th>
                            </tr>
                            <tr style="background: #004c99;">
                                <th>&gt;500</th>
                                <th>250-500</th>
                                <th>&lt;250</th>
                                <th>&lt;100</th>
                                <th>Not Started</th>
                                <th>4Q</th>
                                <th>3Q</th>
                                <th>2Q</th>
                                <th>1Q</th>
                                <th>Rating &gt;1500</th>
                                <th>Rank &lt;20000</th>
                            </tr>
                        </thead>
                        <tbody>
                            {summary_table_rows}
                        </tbody>
                    </table>

                    <div class="section-title">
                        <span>🏆 Top 10 Live Contest Performers</span>
                    </div>
                    <table>
                        <thead>
                            <tr>
                                <th>#</th>
                                <th>Student Name</th>
                                <th>Reg No</th>
                                <th>Global Rank</th>
                                <th>Solved</th>
                                <th>Score</th>
                                <th>Contest Rating</th>
                            </tr>
                        </thead>
                        <tbody>
                            {top_rankers_rows}
                        </tbody>
                    </table>

                    <div style="margin-top: 24px; padding: 14px; background: #e0f2fe; border-left: 4px solid #0284c7; border-radius: 6px; font-size: 12px; color: #0369a1;">
                        <strong>📎 Excel Attachment Included:</strong> Contains complete sheet with Management Summary and full 17-column student-wise verification breakdown.
                    </div>
                </div>

                <div class="footer">
                    Automated LeetCode Weekly Contest Tracker • Nandha Engineering College Autonomous System
                </div>
            </div>
        </body>
        </html>
        """
