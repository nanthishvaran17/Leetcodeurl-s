import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, Any, List

def build_weekly_performance_excel(data: Dict[str, Any], filepath: str) -> str:
    """
    MASTER INSTITUTIONAL 18-SHEET EXCEL WORKBOOK GENERATOR
    Generates LeetCode_Weekly_Report_{date}.xlsx containing complete student master roster,
    batch separation, cohort summaries, category rosters, error reports, and snapshot audit logs.
    """
    wb = openpyxl.Workbook()
    default_sheet = wb.active

    report_date = data.get("report_date", "13-08-2026")
    total_students = data.get("total_students", 0)

    # Styling Tokens
    navy_header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    brand_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

    font_title = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    font_sub = Font(name="Arial", size=10, italic=True, color="E5E7EB")
    font_header = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    font_bold = Font(name="Arial", size=10, bold=True)
    font_regular = Font(name="Arial", size=10)

    thin_border_side = Side(style='thin', color='D1D5DB')
    grid_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')

    def add_header_banner(ws, title_text, subtitle_text):
        ws.merge_cells('A1:S1')
        ws.merge_cells('A2:S2')
        cell_t = ws['A1']
        cell_t.value = f"NANDHA ENGINEERING COLLEGE (AUTONOMOUS) — {title_text}"
        cell_t.font = font_title
        cell_t.fill = navy_header_fill
        cell_t.alignment = align_center

        cell_s = ws['A2']
        cell_s.value = f"{subtitle_text} | Report Date: {report_date} | Total Students: {total_students}"
        cell_s.font = font_sub
        cell_s.fill = brand_fill
        cell_s.alignment = align_center

        ws.row_dimensions[1].height = 28
        ws.row_dimensions[2].height = 20

    def style_table_headers(ws, row_idx, headers):
        ws.row_dimensions[row_idx].height = 24
        for col_idx, h in enumerate(headers, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=h)
            c.font = font_header
            c.fill = navy_header_fill
            c.alignment = align_center
            c.border = grid_border

    def auto_fit_columns(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or '')
                if cell.row in (1, 2):
                    continue
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Master Standardized Student Detail Columns
    std_headers = [
        "S.No", "Register No", "Student Name", "Department", "Year", "Batch",
        "LeetCode URL", "Username", "Easy Solved", "Medium Solved", "Hard Solved",
        "Total Solved", "Category", "Profile Rank", "Latest Contest Name",
        "Questions Solved", "Contest Rating", "Contest Rank",
        "Fetch Status", "Last Successful Fetch", "Last Fetch Attempt", "Fetch Error"
    ]

    def extract_std_row(r, idx):
        return [
            idx,
            r.get("reg_no"),
            r.get("name"),
            r.get("dept"),
            r.get("year"),
            r.get("batch"),
            r.get("leetcode_url"),
            r.get("username"),
            r.get("easy") if r.get("easy") is not None else "N/A",
            r.get("medium") if r.get("medium") is not None else "N/A",
            r.get("hard") if r.get("hard") is not None else "N/A",
            r.get("total_solved") if r.get("total_solved") is not None else "N/A",
            r.get("category"),
            f"#{r.get('profile_ranking'):,}" if r.get("profile_ranking") else "N/A",
            r.get("contest_name") or "N/A",
            r.get("contest_q_solved") or "N/A",
            r.get("contest_rating") if r.get("contest_rating") is not None else "N/A",
            f"#{r.get('contest_ranking'):,}" if r.get("contest_ranking") else "N/A",
            r.get("fetch_status"),
            r.get("last_successful_fetch"),
            r.get("last_fetch_attempt"),
            r.get("fetch_error")
        ]

    def write_student_table(ws, title, subtitle, headers, rows):
        add_header_banner(ws, title, subtitle)
        style_table_headers(ws, 4, headers)
        ws.freeze_panes = 'A5'
        for r_i, r_data in enumerate(rows, start=5):
            ws.row_dimensions[r_i].height = 20
            for c_i, val in enumerate(r_data, start=1):
                c = ws.cell(row=r_i, column=c_i, value=val)
                c.font = font_regular
                c.border = grid_border
                c.alignment = align_left if c_i in (3, 7, 8, 19) else align_center
                if r_i % 2 == 0:
                    c.fill = zebra_fill
        auto_fit_columns(ws)

    cur_all = data.get("all_students_current", [])

    # 1. 00_All_Students (Master Sheet — All 273 Students)
    ws00 = wb.create_sheet(title="00_All_Students")
    write_student_table(ws00, "MASTER ROSTER — ALL STUDENTS", f"Complete Master Roster ({total_students} Students)", std_headers, [extract_std_row(r, i) for i, r in enumerate(cur_all, start=1)])

    # 2. Year-Wise Batch Sheets (01_2023-2027, 02_2024-2028, 03_2025-2029)
    batch_map = data.get("batch_map", {})
    sorted_batches = sorted(list(batch_map.keys()))
    sheet_num = 1
    for b_code in sorted_batches:
        b_rows = batch_map[b_code]
        title_b = f"{sheet_num:02d}_{b_code}"
        ws_b = wb.create_sheet(title=title_b)
        write_student_table(ws_b, f"BATCH ROSTER — {b_code}", f"Academic Batch {b_code} ({len(b_rows)} Students)", std_headers, [extract_std_row(r, i) for i, r in enumerate(b_rows, start=1)])
        sheet_num += 1

    # 3. 04_Year_Summary Sheet
    ws04 = wb.create_sheet(title="04_Year_Summary")
    add_header_banner(ws04, "YEAR-WISE BATCH PERFORMANCE SUMMARY", "Summary by Academic Batch Cohort")
    headers04 = ["Batch", "Total Students", "Verified", "Failed", "Above 500", "250-500", "101-250", "Less than 100", "Not Yet Started", "Average Solved", "Total Solved", "Rating >1500", "Ranking <20000"]
    style_table_headers(ws04, 4, headers04)
    r_idx = 5
    for bs in data.get("batch_summaries", []):
        c = bs["current_week"]
        row_vals = [bs["batch"], bs["num_students"], c["verified"], c["failed"], c["above_500"], c["250_500"], c["101_250"], c["less_100"], c["not_started"], c["avg_solved"], c["total_solved"], c["rating_1500"], c["ranking_20000"]]
        for c_idx, val in enumerate(row_vals, start=1):
            cell = ws04.cell(row=r_idx, column=c_idx, value=val)
            cell.font = font_regular
            cell.alignment = align_center
            cell.border = grid_border
        r_idx += 1
    auto_fit_columns(ws04)

    # 4. 05_Department_Summary Sheet
    ws05 = wb.create_sheet(title="05_Department_Summary")
    add_header_banner(ws05, "DEPARTMENT-WISE METRIC SUMMARY", "Summary by Academic Department")
    headers05 = ["Department", "Total Students", "Verified", "Failed", "Above 500", "250-500", "101-250", "Less than 100", "Not Yet Started", "Average Solved", "Total Solved", "Rating >1500", "Ranking <20000"]
    style_table_headers(ws05, 4, headers05)
    r_idx = 5
    for ds in data.get("dept_summaries", []):
        c = ds["current_week"]
        row_vals = [ds["department"], ds["num_students"], c["verified"], c["failed"], c["above_500"], c["250_500"], c["101_250"], c["less_100"], c["not_started"], c["avg_solved"], c["total_solved"], c["rating_1500"], c["ranking_20000"]]
        for c_idx, val in enumerate(row_vals, start=1):
            cell = ws05.cell(row=r_idx, column=c_idx, value=val)
            cell.font = font_regular
            cell.alignment = align_center
            cell.border = grid_border
        r_idx += 1
    auto_fit_columns(ws05)

    # 5. 06_Year_Department_Summary Sheet
    ws06_yd = wb.create_sheet(title="06_Year_Department_Summary")
    add_header_banner(ws06_yd, "BATCH & DEPARTMENT COHORT SUMMARY", "Summary by Batch & Department Breakdown")
    headers06 = ["Batch", "Department", "Total Students", "Verified", "Failed", "Above 500", "250-500", "101-250", "Less than 100", "Not Yet Started", "Average Solved", "Total Solved", "Rating >1500", "Ranking <20000"]
    style_table_headers(ws06_yd, 4, headers06)
    r_idx = 5
    for yd in data.get("year_dept_summaries", []):
        row_vals = [yd["batch"], yd["department"], yd["total_students"], yd["verified"], yd["failed"], yd["above_500"], yd["250_500"], yd["101_250"], yd["less_100"], yd["not_started"], yd["avg_solved"], yd["total_solved"], yd["rating_1500"], yd["ranking_20000"]]
        for c_idx, val in enumerate(row_vals, start=1):
            cell = ws06_yd.cell(row=r_idx, column=c_idx, value=val)
            cell.font = font_regular
            cell.alignment = align_center
            cell.border = grid_border
        r_idx += 1
    auto_fit_columns(ws06_yd)

    # 6. 07_Current_Week Sheet
    ws07 = wb.create_sheet(title="07_Current_Week")
    write_student_table(ws07, "CURRENT WEEK SNAPSHOT", f"Latest Verified Snapshot ({report_date})", std_headers, [extract_std_row(r, i) for i, r in enumerate(cur_all, start=1)])

    # 7. 08_Last_Week Sheet
    ws08 = wb.create_sheet(title="08_Last_Week")
    write_student_table(ws08, "LAST WEEK SNAPSHOT", "Previous Historical Snapshot", std_headers, [extract_std_row(r, i) for i, r in enumerate(data.get("all_students_last_week", []), start=1)])

    # 8. 09_Weekly_Comparison Sheet
    ws09 = wb.create_sheet(title="09_Weekly_Comparison")
    comp_headers = ["S.No", "Register No", "Student Name", "Department", "Year", "Batch", "Last Solved", "Current Solved", "Problems Added", "Last Category", "Current Category", "Category Movement"]
    comp_rows = []
    last_week_recs = {r["reg_no"]: r for r in data.get("all_students_last_week", [])}
    for idx, r in enumerate(cur_all, start=1):
        l_rec = last_week_recs.get(r["reg_no"], {})
        l_sol = l_rec.get("total_solved")
        c_sol = r.get("total_solved")
        l_cat = l_rec.get("category", "Data Unavailable")
        c_cat = r.get("category", "Data Unavailable")
        added = (c_sol - l_sol) if (c_sol is not None and l_sol is not None) else "N/A"
        added_str = f"+{added}" if isinstance(added, int) and added > 0 else str(added)
        mov = "PROMOTED" if (l_cat != c_cat and isinstance(added, int) and added > 0) else ("NO CHANGE" if (isinstance(added, int) and added == 0) else ("ATTENTION" if (isinstance(added, int) and added < 0) else "STABLE"))
        comp_rows.append([idx, r["reg_no"], r["name"], r["dept"], r["year"], r["batch"], l_sol if l_sol is not None else "N/A", c_sol if c_sol is not None else "N/A", added_str, l_cat, c_cat, mov])
    write_student_table(ws09, "STUDENT WEEKLY COMPARISON", "Student-Level Movement & Progression", comp_headers, comp_rows)

    # 9. Category Roster Sheets (10_Above_500, 11_250_500, 12_101_250, 13_Less_Than_100, 14_Not_Started)
    cat_sheets = [
        ("10_Above_500", "STUDENT ROSTER — ABOVE 500 SOLVED", "above_500"),
        ("11_250_500", "STUDENT ROSTER — 250 TO 500 SOLVED", "250_500"),
        ("12_101_250", "STUDENT ROSTER — 101 TO 250 SOLVED", "101_250"),
        ("13_Less_Than_100", "STUDENT ROSTER — LESS THAN 100 SOLVED", "less_100"),
        ("14_Not_Started", "STUDENT ROSTER — NOT YET STARTED", "not_started"),
    ]
    for sheet_t, title_t, cat_key in cat_sheets:
        ws_cat = wb.create_sheet(title=sheet_t)
        cat_recs = data.get("categories", {}).get(cat_key, [])
        write_student_table(ws_cat, title_t, f"Category Roster ({len(cat_recs)} Students)", std_headers, [extract_std_row(r, i) for i, r in enumerate(cat_recs, start=1)])

    # 10. 15_Fetch_Status Sheet
    ws15 = wb.create_sheet(title="15_Fetch_Status")
    add_header_banner(ws15, "FETCH STATUS SUMMARY", "Overall Verification Metrics Breakdown")
    headers15 = ["Fetch Metric Category", "Student Count"]
    style_table_headers(ws15, 4, headers15)
    fs_data = data.get("fetch_status_summary", {})
    r_idx = 5
    for f_k, f_v in fs_data.items():
        c1 = ws15.cell(row=r_idx, column=1, value=f_k)
        c2 = ws15.cell(row=r_idx, column=2, value=f_v)
        c1.font, c2.font = font_bold, font_regular
        c1.border, c2.border = grid_border, grid_border
        c1.alignment, c2.alignment = align_left, align_center
        r_idx += 1
    auto_fit_columns(ws15)

    # 11. 16_Fetch_Errors Sheet
    ws16_err = wb.create_sheet(title="16_Fetch_Errors")
    err_headers = ["S.No", "Register No", "Student Name", "Department", "Year", "Batch", "Username", "LeetCode URL", "Error Type", "Error Message", "Last Successful Fetch", "Latest Attempt", "Previous Total", "Current Attempt Status", "Action Required"]
    err_rows = []
    for err in data.get("fetch_errors", []):
        err_rows.append([err["s_no"], err["reg_no"], err["name"], err["dept"], err["year"], err["batch"], err["username"], err["leetcode_url"], err["error_type"], err["error_message"], err["last_successful_fetch"], err["latest_attempt"], err["previous_total"], err["current_attempt_status"], err["action_required"]])
    write_student_table(ws16_err, "FETCH ERRORS & ATTENTION REPORT", f"Profile Error Log ({len(err_rows)} Records)", err_headers, err_rows)

    # 12. 17_Data_Validation Sheet
    ws17 = wb.create_sheet(title="17_Data_Validation")
    val_headers = ["Issue Type", "Register No", "Student Name", "Field", "Expected Value", "Actual Value", "Severity", "Status"]
    val_rows = []
    for v in data.get("validation_issues", []):
        val_rows.append([v.get("issue_type"), v.get("reg_no"), v.get("student"), v.get("field"), v.get("expected"), v.get("actual"), v.get("severity"), v.get("status")])
    write_student_table(ws17, "DATA VALIDATION & AUDIT LOG", "Integrity Audits & Count Equations", val_headers, val_rows)

    # 13. 18_Snapshot_Audit Sheet
    ws18 = wb.create_sheet(title="18_Snapshot_Audit")
    audit_headers = ["S.No", "Student Name", "Register No", "Department", "Batch", "Previous Snapshot Date", "Previous Total", "Current Snapshot Date", "Current Total", "Change", "Status"]
    audit_rows = []
    for sa in data.get("snapshot_audit", []):
        audit_rows.append([sa["s_no"], sa["student"], sa["reg_no"], sa["dept"], sa["batch"], sa["previous_snapshot_date"], sa["previous_total"], sa["current_snapshot_date"], sa["current_total"], sa["change"], sa["status"]])
    write_student_table(ws18, "HISTORICAL SNAPSHOT AUDIT LOG", "Student Snapshot Progression Audit", audit_headers, audit_rows)

    # 14. 19_Contest_Validation Sheet
    ws19_contest = wb.create_sheet(title="19_Contest_Validation")
    c_val_headers = ["Register No", "Student Name", "Username", "Contest Query Status", "Contest Parse Status", "Contest Name", "Contest Date", "Questions Solved", "Questions Total", "Contest Rating", "Contest Rank", "Profile Rank", "Error Message", "Last Successful Contest Sync"]
    c_val_rows = []
    for cv in data.get("contest_validation", []):
        c_val_rows.append([cv.get("reg_no"), cv.get("name"), cv.get("username"), cv.get("contest_query_status"), cv.get("contest_parse_status"), cv.get("contest_name"), cv.get("contest_date"), cv.get("questions_solved"), cv.get("questions_total"), cv.get("contest_rating"), cv.get("contest_rank"), cv.get("profile_rank"), cv.get("error_message"), cv.get("last_successful_contest_sync")])
    write_student_table(ws19_contest, "CONTEST DATA VALIDATION & AUDIT SHEET", f"Unverified Contest Profiles ({len(c_val_rows)} Records)", c_val_headers, c_val_rows)

    # Remove default empty sheet
    if default_sheet in wb.worksheets and len(wb.worksheets) > 1:
        wb.remove(default_sheet)

    try:
        wb.save(filepath)
        return filepath
    except PermissionError:
        import datetime
        import os
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        base, ext = os.path.splitext(filepath)
        fallback_path = f"{base}_{timestamp}{ext}"
        wb.save(fallback_path)
        print(f"\n⚠️ NOTICE: Destination file '{os.path.basename(filepath)}' is open in Microsoft Excel.")
        print(f"   Saved output workbook to fallback file: {fallback_path}")
        return fallback_path


def build_public_contest_excel(data: Dict[str, Any], filepath: str) -> str:
    """Generates standalone Public_Contest.xlsx containing ONLY Public contest results."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Public_Contest"

    report_date = data.get("report_date", "13-08-2026")
    contest_name = data.get("contest_name", "Weekly Contest")
    contest_number = data.get("contest_number", "")
    contest_date = data.get("contest_date", "")

    navy_header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    brand_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

    font_title = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    font_sub = Font(name="Arial", size=10, italic=True, color="E5E7EB")
    font_header = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    font_bold = Font(name="Arial", size=10, bold=True)
    font_regular = Font(name="Arial", size=10)

    thin_border_side = Side(style='thin', color='D1D5DB')
    grid_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')

    # Banner
    ws.merge_cells('A1:S1')
    ws.merge_cells('A2:S2')
    cell_t = ws['A1']
    cell_t.value = f"NANDHA ENGINEERING COLLEGE (AUTONOMOUS) — Public Contest Performance Report"
    cell_t.font = font_title
    cell_t.fill = navy_header_fill
    cell_t.alignment = align_center

    cell_s = ws['A2']
    cell_s.value = f"Contest: {contest_name} (#{contest_number}) | Contest Date: {contest_date} | Report Date: {report_date} | Total Students: {len(data.get('rows', []))}"
    cell_s.font = font_sub
    cell_s.fill = brand_fill
    cell_s.alignment = align_center

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 20

    # Summary Row
    sum_data = data.get("public_summary", {})
    ws.merge_cells('A4:S4')
    ws['A4'] = f"PUBLIC CONTEST SUMMARY: 4Q Solved: {sum_data.get('q4', 0)} | 3Q Solved: {sum_data.get('q3', 0)} | 2Q Solved: {sum_data.get('q2', 0)} | 1Q Solved: {sum_data.get('q1', 0)} | Not Attended: {sum_data.get('not_attended', 0)} | Fetch Failed: {sum_data.get('fetch_failed', 0)} | Mode Uncertain: {sum_data.get('mode_uncertain', 0)}"
    ws['A4'].font = font_bold
    ws['A4'].alignment = align_center

    headers = [
        "S.No", "Register No", "Student Name", "Department", "Year", "Batch",
        "LeetCode Username", "Contest Name", "Contest Number", "Contest Date",
        "Public Attendance", "Questions Solved", "Questions Total", "Score",
        "Contest Rank", "Contest Rating", "Top %", "Status", "Fetched At"
    ]

    ws.row_dimensions[6].height = 24
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=6, column=col_idx, value=h)
        c.font = font_header
        c.fill = navy_header_fill
        c.alignment = align_center
        c.border = grid_border

    current_row = 7
    for row_data in data.get("rows", []):
        ws.row_dimensions[current_row].height = 20
        fill = zebra_fill if current_row % 2 == 0 else PatternFill(fill_type=None)

        ws.cell(row=current_row, column=1, value=row_data.get("s_no")).alignment = align_center
        ws.cell(row=current_row, column=2, value=row_data.get("reg_no")).alignment = align_center
        ws.cell(row=current_row, column=3, value=row_data.get("student_name")).alignment = align_left
        ws.cell(row=current_row, column=4, value=row_data.get("department")).alignment = align_center
        ws.cell(row=current_row, column=5, value=row_data.get("year")).alignment = align_center
        ws.cell(row=current_row, column=6, value=row_data.get("batch")).alignment = align_center
        ws.cell(row=current_row, column=7, value=row_data.get("username")).alignment = align_left
        ws.cell(row=current_row, column=8, value=row_data.get("contest_name", contest_name)).alignment = align_left
        ws.cell(row=current_row, column=9, value=row_data.get("contest_number", contest_number)).alignment = align_center
        ws.cell(row=current_row, column=10, value=row_data.get("contest_date", contest_date)).alignment = align_center
        ws.cell(row=current_row, column=11, value="ATTENDED" if row_data.get("attended") else "NOT ATTENDED").alignment = align_center
        ws.cell(row=current_row, column=12, value=row_data.get("questions_solved", 0)).alignment = align_center
        ws.cell(row=current_row, column=13, value=row_data.get("questions_total", 4)).alignment = align_center
        ws.cell(row=current_row, column=14, value=row_data.get("score_display", "Not Attended")).alignment = align_center
        ws.cell(row=current_row, column=15, value=row_data.get("contest_rank") or "—").alignment = align_center
        ws.cell(row=current_row, column=16, value=row_data.get("contest_rating") or "—").alignment = align_center
        ws.cell(row=current_row, column=17, value=row_data.get("top_percentage") or "—").alignment = align_center
        ws.cell(row=current_row, column=18, value=row_data.get("status", "NOT_ATTENDED")).alignment = align_center
        ws.cell(row=current_row, column=19, value=row_data.get("fetched_at") or "—").alignment = align_center

        for col_idx in range(1, 20):
            c = ws.cell(row=current_row, column=col_idx)
            c.font = font_regular
            c.border = grid_border
            if fill.fill_type:
                c.fill = fill
        current_row += 1

    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(len(str(cell.value or '')) for cell in col if cell.row > 2)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    wb.save(filepath)
    return filepath


def build_virtual_contest_excel(data: Dict[str, Any], filepath: str) -> str:
    """Generates standalone Virtual_Contest.xlsx containing ONLY Virtual contest results."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Virtual_Contest"

    report_date = data.get("report_date", "13-08-2026")
    contest_name = data.get("contest_name", "Weekly Contest")
    contest_number = data.get("contest_number", "")
    contest_date = data.get("contest_date", "")

    navy_header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    brand_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

    font_title = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    font_sub = Font(name="Arial", size=10, italic=True, color="E5E7EB")
    font_header = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    font_bold = Font(name="Arial", size=10, bold=True)
    font_regular = Font(name="Arial", size=10)

    thin_border_side = Side(style='thin', color='D1D5DB')
    grid_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')

    # Banner
    ws.merge_cells('A1:S1')
    ws.merge_cells('A2:S2')
    cell_t = ws['A1']
    cell_t.value = f"NANDHA ENGINEERING COLLEGE (AUTONOMOUS) — Virtual Contest Performance Report"
    cell_t.font = font_title
    cell_t.fill = navy_header_fill
    cell_t.alignment = align_center

    cell_s = ws['A2']
    cell_s.value = f"Contest: {contest_name} (#{contest_number}) | Contest Date: {contest_date} | Report Date: {report_date} | Total Students: {len(data.get('rows', []))}"
    cell_s.font = font_sub
    cell_s.fill = brand_fill
    cell_s.alignment = align_center

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 20

    # Summary Row
    sum_data = data.get("virtual_summary", {})
    ws.merge_cells('A4:S4')
    ws['A4'] = f"VIRTUAL CONTEST SUMMARY: 4Q Solved: {sum_data.get('q4', 0)} | 3Q Solved: {sum_data.get('q3', 0)} | 2Q Solved: {sum_data.get('q2', 0)} | 1Q Solved: {sum_data.get('q1', 0)} | Not Attended: {sum_data.get('not_attended', 0)} | Fetch Failed: {sum_data.get('fetch_failed', 0)} | Mode Uncertain: {sum_data.get('mode_uncertain', 0)}"
    ws['A4'].font = font_bold
    ws['A4'].alignment = align_center

    headers = [
        "S.No", "Register No", "Student Name", "Department", "Year", "Batch",
        "LeetCode Username", "Contest Name", "Contest Number", "Contest Date",
        "Virtual Attendance", "Questions Solved", "Questions Total", "Score",
        "Contest Rank", "Contest Rating", "Top %", "Status", "Fetched At"
    ]

    ws.row_dimensions[6].height = 24
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=6, column=col_idx, value=h)
        c.font = font_header
        c.fill = navy_header_fill
        c.alignment = align_center
        c.border = grid_border

    current_row = 7
    for row_data in data.get("rows", []):
        ws.row_dimensions[current_row].height = 20
        fill = zebra_fill if current_row % 2 == 0 else PatternFill(fill_type=None)

        ws.cell(row=current_row, column=1, value=row_data.get("s_no")).alignment = align_center
        ws.cell(row=current_row, column=2, value=row_data.get("reg_no")).alignment = align_center
        ws.cell(row=current_row, column=3, value=row_data.get("student_name")).alignment = align_left
        ws.cell(row=current_row, column=4, value=row_data.get("department")).alignment = align_center
        ws.cell(row=current_row, column=5, value=row_data.get("year")).alignment = align_center
        ws.cell(row=current_row, column=6, value=row_data.get("batch")).alignment = align_center
        ws.cell(row=current_row, column=7, value=row_data.get("username")).alignment = align_left
        ws.cell(row=current_row, column=8, value=row_data.get("contest_name", contest_name)).alignment = align_left
        ws.cell(row=current_row, column=9, value=row_data.get("contest_number", contest_number)).alignment = align_center
        ws.cell(row=current_row, column=10, value=row_data.get("contest_date", contest_date)).alignment = align_center
        ws.cell(row=current_row, column=11, value="ATTENDED" if row_data.get("attended") else "NOT ATTENDED").alignment = align_center
        ws.cell(row=current_row, column=12, value=row_data.get("questions_solved", 0)).alignment = align_center
        ws.cell(row=current_row, column=13, value=row_data.get("questions_total", 4)).alignment = align_center
        ws.cell(row=current_row, column=14, value=row_data.get("score_display", "Not Attended")).alignment = align_center
        ws.cell(row=current_row, column=15, value=row_data.get("contest_rank") or "—").alignment = align_center
        ws.cell(row=current_row, column=16, value=row_data.get("contest_rating") or "—").alignment = align_center
        ws.cell(row=current_row, column=17, value=row_data.get("top_percentage") or "—").alignment = align_center
        ws.cell(row=current_row, column=18, value=row_data.get("status", "NOT_ATTENDED")).alignment = align_center
        ws.cell(row=current_row, column=19, value=row_data.get("fetched_at") or "—").alignment = align_center

        for col_idx in range(1, 20):
            c = ws.cell(row=current_row, column=col_idx)
            c.font = font_regular
            c.border = grid_border
            if fill.fill_type:
                c.fill = fill
        current_row += 1

    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(len(str(cell.value or '')) for cell in col if cell.row > 2)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    wb.save(filepath)
    return filepath


def build_contest_combined_excel(data: Dict[str, Any], filepath: str) -> str:
    """Generates Contest_Combined.xlsx containing side-by-side Public and Virtual contest comparison + Data Validation sheet."""
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Contest_Combined"

    report_date = data.get("report_date", "13-08-2026")
    contest_name = data.get("contest_name", "Weekly Contest")
    contest_number = data.get("contest_number", "")
    contest_date = data.get("contest_date", "")

    navy_header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    brand_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

    font_title = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    font_sub = Font(name="Arial", size=10, italic=True, color="E5E7EB")
    font_header = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    font_regular = Font(name="Arial", size=10)

    thin_border_side = Side(style='thin', color='D1D5DB')
    grid_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')

    # Banner
    ws1.merge_cells('A1:V1')
    ws1.merge_cells('A2:V2')
    ws1['A1'] = "NANDHA ENGINEERING COLLEGE (AUTONOMOUS) — Public & Virtual Contest Combined Final Report"
    ws1['A1'].font = font_title
    ws1['A1'].fill = navy_header_fill
    ws1['A1'].alignment = align_center

    ws1['A2'] = f"Contest: {contest_name} (#{contest_number}) | Contest Date: {contest_date} | Report Date: {report_date} | Total Students: {len(data.get('rows', []))}"
    ws1['A2'].font = font_sub
    ws1['A2'].fill = brand_fill
    ws1['A2'].alignment = align_center

    headers = [
        "Register No", "Student Name", "Department", "Year", "Batch", "Username",
        "Public Contest", "Public Score", "Public Questions Solved", "Public Questions Total", "Public Rank", "Public Rating",
        "Virtual Contest", "Virtual Score", "Virtual Questions Solved", "Virtual Questions Total", "Virtual Rank", "Virtual Rating",
        "Overall Participation Mode", "Public Status", "Virtual Status", "Last Fetched"
    ]

    ws1.row_dimensions[4].height = 24
    for col_idx, h in enumerate(headers, start=1):
        c = ws1.cell(row=4, column=col_idx, value=h)
        c.font = font_header
        c.fill = navy_header_fill
        c.alignment = align_center
        c.border = grid_border

    current_row = 5
    for r in data.get("rows", []):
        ws1.row_dimensions[current_row].height = 20
        fill = zebra_fill if current_row % 2 == 0 else PatternFill(fill_type=None)

        pub = r.get("public_contest_result", {})
        vir = r.get("virtual_contest_result", {})

        ws1.cell(row=current_row, column=1, value=r.get("reg_no")).alignment = align_center
        ws1.cell(row=current_row, column=2, value=r.get("student_name")).alignment = align_left
        ws1.cell(row=current_row, column=3, value=r.get("department")).alignment = align_center
        ws1.cell(row=current_row, column=4, value=r.get("year")).alignment = align_center
        ws1.cell(row=current_row, column=5, value=r.get("batch")).alignment = align_center
        ws1.cell(row=current_row, column=6, value=r.get("username")).alignment = align_left

        ws1.cell(row=current_row, column=7, value=pub.get("contest_name", contest_name)).alignment = align_left
        ws1.cell(row=current_row, column=8, value=pub.get("score_display", "Not Attended")).alignment = align_center
        ws1.cell(row=current_row, column=9, value=pub.get("questions_solved", 0)).alignment = align_center
        ws1.cell(row=current_row, column=10, value=pub.get("questions_total", 4)).alignment = align_center
        ws1.cell(row=current_row, column=11, value=pub.get("contest_rank") or "—").alignment = align_center
        ws1.cell(row=current_row, column=12, value=pub.get("contest_rating") or "—").alignment = align_center

        ws1.cell(row=current_row, column=13, value=vir.get("contest_name", contest_name)).alignment = align_left
        ws1.cell(row=current_row, column=14, value=vir.get("score_display", "Not Attended")).alignment = align_center
        ws1.cell(row=current_row, column=15, value=vir.get("questions_solved", 0)).alignment = align_center
        ws1.cell(row=current_row, column=16, value=vir.get("questions_total", 4)).alignment = align_center
        ws1.cell(row=current_row, column=17, value=vir.get("contest_rank") or "—").alignment = align_center
        ws1.cell(row=current_row, column=18, value=vir.get("contest_rating") or "—").alignment = align_center

        ws1.cell(row=current_row, column=19, value=r.get("overall_participation_mode", "NONE")).alignment = align_center
        ws1.cell(row=current_row, column=20, value=pub.get("status", "NOT_ATTENDED")).alignment = align_center
        ws1.cell(row=current_row, column=21, value=vir.get("status", "NOT_ATTENDED")).alignment = align_center
        ws1.cell(row=current_row, column=22, value=r.get("fetched_at") or "—").alignment = align_center

        for col_idx in range(1, 23):
            c = ws1.cell(row=current_row, column=col_idx)
            c.font = font_regular
            c.border = grid_border
            if fill.fill_type:
                c.fill = fill
        current_row += 1

    for col in ws1.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(len(str(cell.value or '')) for cell in col if cell.row > 2)
        ws1.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Add Contest_Data_Validation sheet per Section 19
    ws2 = wb.create_sheet(title="Contest_Data_Validation")
    v_headers = [
        "Register No", "Student Name", "Username", "Contest", "Contest Number",
        "Participation Mode", "Questions Solved", "Questions Total", "Contest Rank",
        "Contest Rating", "Status", "Error Message", "Fetched At"
    ]
    ws2.row_dimensions[1].height = 24
    for col_idx, h in enumerate(v_headers, start=1):
        c = ws2.cell(row=1, column=col_idx, value=h)
        c.font = font_header
        c.fill = navy_header_fill
        c.alignment = align_center

    c_row = 2
    for err_item in data.get("validation_logs", []):
        ws2.cell(row=c_row, column=1, value=err_item.get("reg_no")).alignment = align_center
        ws2.cell(row=c_row, column=2, value=err_item.get("student_name")).alignment = align_left
        ws2.cell(row=c_row, column=3, value=err_item.get("username")).alignment = align_left
        ws2.cell(row=c_row, column=4, value=err_item.get("contest_name", contest_name)).alignment = align_left
        ws2.cell(row=c_row, column=5, value=err_item.get("contest_number", contest_number)).alignment = align_center
        ws2.cell(row=c_row, column=6, value=err_item.get("participation_mode", "UNKNOWN")).alignment = align_center
        ws2.cell(row=c_row, column=7, value=err_item.get("questions_solved", 0)).alignment = align_center
        ws2.cell(row=c_row, column=8, value=err_item.get("questions_total", 4)).alignment = align_center
        ws2.cell(row=c_row, column=9, value=err_item.get("contest_rank") or "—").alignment = align_center
        ws2.cell(row=c_row, column=10, value=err_item.get("contest_rating") or "—").alignment = align_center
        ws2.cell(row=c_row, column=11, value=err_item.get("status", "VERIFIED")).alignment = align_center
        ws2.cell(row=c_row, column=12, value=err_item.get("error_message") or "—").alignment = align_left
        ws2.cell(row=c_row, column=13, value=err_item.get("fetched_at") or "—").alignment = align_center
        c_row += 1

    wb.save(filepath)
    return filepath

