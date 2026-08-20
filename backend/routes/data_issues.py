import datetime
import io
import re
from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, Query, Response
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import desc, func

from backend.database import get_db
from backend.models import Student, LeetCodeProfileStats, Department, Section, AuditLog
from backend.leetcode_client import fetch_leetcode_profile, extract_leetcode_username
from backend.logger import logger
from backend.routes.auth import get_current_user

router = APIRouter(prefix="/api/data-issues", tags=["Student Data Issues & Recovery"])

# Helper: Classify a student's data issue using mutually exclusive strict priority
def classify_student_issue(student: Student) -> dict:
    stats = student.stats
    username = (student.username or "").strip()
    leetcode_url = (student.leetcode_url or "").strip()
    
    # 1. Canonical LeetCode Profile URL generator
    canonical_url = None
    if username:
        canonical_url = f"https://leetcode.com/u/{username}/"
    elif leetcode_url and ("leetcode.com" in leetcode_url or "leetcode.cn" in leetcode_url):
        extracted = extract_leetcode_username(leetcode_url)
        if extracted:
            canonical_url = f"https://leetcode.com/u/{extracted}/"

    # Default issue classification
    issue_category = "HEALTHY"
    issue_label = "Healthy Record"
    severity = "HEALTHY"
    error_desc = "Profile verified, synced, and active in evaluation."
    recommended_action = "Routine monitoring."
    url_status = "VERIFIED"

    # Check 1: Missing Username
    if not username and not leetcode_url:
        issue_category = "MISSING_USERNAME"
        issue_label = "Missing Username"
        severity = "CRITICAL"
        error_desc = "LeetCode username is not configured for this student."
        recommended_action = "Add valid LeetCode username in profile."
        url_status = "NO_USERNAME"
        canonical_url = None

    # Check 2: Invalid / Unverifiable URL
    elif not canonical_url:
        issue_category = "INVALID_URL"
        issue_label = "Invalid Profile URL"
        severity = "CRITICAL"
        error_desc = "Stored profile URL is malformed or invalid."
        recommended_action = "Repair stored LeetCode URL."
        url_status = "INVALID"

    # Check 3: Profile Not Found on LeetCode
    elif stats and (stats.status == "PROFILE NOT FOUND" or stats.error_code == "PROFILE_NOT_FOUND"):
        issue_category = "INVALID_USERNAME"
        issue_label = "Profile Not Found"
        severity = "CRITICAL"
        error_desc = stats.error_message or f"LeetCode user '{username}' could not be resolved on LeetCode servers."
        recommended_action = "Verify and correct LeetCode username."
        url_status = "INVALID"

    # Check 4: Data Mismatch / Integrity Conflict
    elif stats and (stats.sync_status == "mismatch" or stats.validation_status == "mismatch"):
        issue_category = "DATA_MISMATCH"
        issue_label = "Data Mismatch"
        severity = "WARNING"
        error_desc = stats.error_message or "Stored solved count differs from verified public profile."
        recommended_action = "Trigger deep reconciliation sync."
        url_status = "NEEDS_CHECK"

    # Check 5: Sync Failed (only if current sync is actually in failed state, not leftover error_code from past)
    elif stats and (
        stats.sync_status == "failed" or
        (stats.status and stats.status.startswith("INVALID")) or
        (stats.error_code and stats.error_code == "PENDING_USERNAME") or
        (stats.error_code and stats.sync_status != "success" and stats.status != "verified")
    ):
        issue_category = "SYNC_FAILED"
        issue_label = "Sync Failed"
        severity = "CRITICAL"
        error_desc = stats.error_message or "LeetCode profile fetch failed during the last synchronization."
        recommended_action = "Retry sync or check LeetCode API status."
        url_status = "NEEDS_CHECK"

    # Check 6: Never Synced
    elif not stats or stats.sync_status in ("not_started", "pending", None) or (stats.last_successful_sync is None and (stats.total_solved is None)):
        issue_category = "NEVER_SYNCED"
        issue_label = "Never Synced"
        severity = "CRITICAL"
        error_desc = "No successful synchronization has ever been recorded for this student."
        recommended_action = "Execute initial sync for this student."
        url_status = "NEEDS_CHECK"

    # Check 7: Stale Data (> 7 days since last successful sync)
    elif stats.last_successful_sync:
        now_dt = datetime.datetime.utcnow()
        sync_dt = stats.last_successful_sync
        if isinstance(sync_dt, datetime.datetime):
            age_days = (now_dt - sync_dt).days
            if age_days >= 7:
                issue_category = "STALE_DATA"
                issue_label = "Stale Data"
                severity = "WARNING"
                error_desc = f"Last successful verification was {age_days} days ago (exceeds 7-day threshold)."
                recommended_action = "Trigger on-demand stats refresh."
                url_status = "VERIFIED"

    # Check 8: Not Started (0 Solved, but profile is verified)
    if issue_category == "HEALTHY" and stats and stats.total_solved == 0:
        issue_category = "NOT_STARTED"
        issue_label = "Not Started (0 Solved)"
        severity = "INFO"
        error_desc = "Student has a valid, verified LeetCode profile but has solved 0 problems."
        recommended_action = "Faculty mentoring and introductory problem assignment."
        url_status = "VERIFIED"

    # Format department names consistently
    dept_code = student.department.code if student.department else "CSE"
    dept_name = student.department.name if student.department else "Computer Science and Engineering"
    if "IOT" in dept_code.upper():
        canonical_dept = "Computer Science and Engineering (IoT)"
        short_dept = "CSE (IoT)"
    else:
        canonical_dept = "Computer Science and Engineering (Cyber Security)"
        short_dept = "CSE (Cyber Security)"

    last_sync_str = "Never"
    if stats and stats.last_successful_sync:
        last_sync_str = stats.last_successful_sync.strftime("%d %b %Y, %I:%M %p")
    elif stats and stats.last_attempt_at:
        last_sync_str = f"Attempted: {stats.last_attempt_at.strftime('%d %b %Y, %I:%M %p')}"

    return {
        "id": student.id,
        "name": student.name,
        "reg_no": student.reg_no,
        "department_code": dept_code,
        "department_name": canonical_dept,
        "department_short": short_dept,
        "year_level": student.year_level or "II Year",
        "username": username or None,
        "leetcode_url": canonical_url,
        "url_status": url_status,
        "issue_category": issue_category,
        "issue_label": issue_label,
        "severity": severity,
        "error_description": error_desc,
        "recommended_action": recommended_action,
        "total_solved": stats.total_solved if (stats and stats.total_solved is not None) else 0,
        "contest_rating": stats.contest_rating if (stats and stats.contest_rating is not None) else None,
        "sync_status": stats.sync_status if stats else "not_started",
        "last_sync": last_sync_str,
        "last_sync_raw": stats.last_successful_sync.isoformat() if (stats and stats.last_successful_sync) else None,
        "is_active": student.is_active
    }


@router.get("/summary")
def get_data_issues_summary(db: Session = Depends(get_db)):
    """
    Computes real, ground-truth issue counts across all students.
    """
    students = db.query(Student).options(
        joinedload(Student.stats),
        joinedload(Student.department)
    ).all()

    classified = [classify_student_issue(s) for s in students]

    counts = {
        "total_students": len(classified),
        "not_started": sum(1 for c in classified if c["issue_category"] == "NOT_STARTED"),
        "sync_failed": sum(1 for c in classified if c["issue_category"] == "SYNC_FAILED"),
        "never_synced": sum(1 for c in classified if c["issue_category"] == "NEVER_SYNCED"),
        "missing_username": sum(1 for c in classified if c["issue_category"] == "MISSING_USERNAME"),
        "invalid_username": sum(1 for c in classified if c["issue_category"] == "INVALID_USERNAME"),
        "invalid_url": sum(1 for c in classified if c["issue_category"] == "INVALID_URL"),
        "stale_data": sum(1 for c in classified if c["issue_category"] == "STALE_DATA"),
        "data_mismatch": sum(1 for c in classified if c["issue_category"] == "DATA_MISMATCH"),
        "healthy": sum(1 for c in classified if c["issue_category"] == "HEALTHY"),
        "critical_issues": sum(1 for c in classified if c["severity"] == "CRITICAL"),
        "warning_issues": sum(1 for c in classified if c["severity"] == "WARNING"),
    }

    # Department breakdown matrix
    dept_matrix = {}
    for c in classified:
        d = c["department_short"]
        if d not in dept_matrix:
            dept_matrix[d] = {
                "department": d,
                "total": 0,
                "sync_failed": 0,
                "missing_username": 0,
                "not_started": 0,
                "never_synced": 0,
                "stale_data": 0,
                "healthy": 0
            }
        dept_matrix[d]["total"] += 1
        cat = c["issue_category"].lower()
        if cat in dept_matrix[d]:
            dept_matrix[d][cat] += 1

    # Year breakdown matrix
    year_matrix = {}
    for c in classified:
        y = c["year_level"]
        if y not in year_matrix:
            year_matrix[y] = {
                "year": y,
                "total": 0,
                "sync_failed": 0,
                "missing_username": 0,
                "not_started": 0,
                "never_synced": 0,
                "stale_data": 0,
                "healthy": 0
            }
        year_matrix[y]["total"] += 1
        cat = c["issue_category"].lower()
        if cat in year_matrix[y]:
            year_matrix[y][cat] += 1

    return {
        "counts": counts,
        "dept_breakdown": list(dept_matrix.values()),
        "year_breakdown": sorted(list(year_matrix.values()), key=lambda x: x["year"])
    }


@router.get("/students")
def get_data_issues_students(
    department: Optional[str] = Query("all", description="All, CSE(CS), or CSE(IOT)"),
    year_level: Optional[str] = Query("all", description="All, II Year, III Year, IV Year"),
    issue_type: Optional[str] = Query("all", description="Specific issue category or critical"),
    search: Optional[str] = Query(None, description="Search term across name, reg_no, username, issue"),
    limit: int = Query(500, ge=1, le=1000),
    db: Session = Depends(get_db)
):
    """
    Returns filtered student issues list with full URL validation metadata.
    """
    students = db.query(Student).options(
        joinedload(Student.stats),
        joinedload(Student.department)
    ).all()

    classified = [classify_student_issue(s) for s in students]

    # 1. Filter by Department
    if department and department.lower() != "all":
        dept_lower = department.lower()
        classified = [
            c for c in classified
            if dept_lower in c["department_code"].lower() or dept_lower in c["department_name"].lower()
        ]

    # 2. Filter by Academic Year
    if year_level and year_level.lower() != "all":
        y_clean = year_level.replace("Year", "").strip().lower()
        classified = [
            c for c in classified
            if y_clean in c["year_level"].lower()
        ]

    # 3. Filter by Issue Category
    if issue_type and issue_type.upper() != "ALL":
        i_upper = issue_type.upper()
        if i_upper == "CRITICAL":
            classified = [c for c in classified if c["severity"] == "CRITICAL"]
        elif i_upper == "WARNING":
            classified = [c for c in classified if c["severity"] == "WARNING"]
        elif i_upper == "ISSUES":
            classified = [c for c in classified if c["issue_category"] != "HEALTHY"]
        else:
            classified = [c for c in classified if c["issue_category"] == i_upper]

    # 4. Search Filter
    if search and search.strip():
        q = search.strip().lower()
        classified = [
            c for c in classified
            if q in c["name"].lower() or
               q in c["reg_no"].lower() or
               (c["username"] and q in c["username"].lower()) or
               q in c["issue_label"].lower() or
               q in c["error_description"].lower()
        ]

    return {
        "total_matched": len(classified),
        "students": classified[:limit]
    }


class VerifyUrlRequest(BaseModel):
    username: str

@router.post("/verify-url")
def verify_leetcode_url(req: VerifyUrlRequest):
    """
    Live on-demand validation of a LeetCode username.
    """
    clean_user = extract_leetcode_username(req.username) or req.username.strip()
    if not clean_user:
        raise HTTPException(status_code=400, detail="LeetCode username cannot be empty.")

    try:
        profile = fetch_leetcode_profile(clean_user)
        if not profile or not profile.get("exists"):
            return {
                "valid": False,
                "username": clean_user,
                "canonical_url": f"https://leetcode.com/u/{clean_user}/",
                "message": f"LeetCode user '{clean_user}' does not exist on LeetCode.",
                "total_solved": 0
            }

        return {
            "valid": True,
            "username": clean_user,
            "canonical_url": f"https://leetcode.com/u/{clean_user}/",
            "message": "LeetCode profile exists and is active.",
            "total_solved": profile.get("total_solved", 0),
            "easy_solved": profile.get("easy_solved", 0),
            "medium_solved": profile.get("medium_solved", 0),
            "hard_solved": profile.get("hard_solved", 0),
            "contest_rating": profile.get("contest_rating")
        }
    except Exception as e:
        logger.error(f"Verify URL failed: {e}")
        return {
            "valid": False,
            "username": clean_user,
            "canonical_url": f"https://leetcode.com/u/{clean_user}/",
            "message": f"Verification error: {str(e)}",
            "total_solved": 0
        }


class RepairProfileRequest(BaseModel):
    new_username: str
    admin_name: Optional[str] = "Administrator"

@router.put("/repair-profile/{student_id}")
def repair_student_profile(
    student_id: int,
    req: RepairProfileRequest,
    db: Session = Depends(get_db)
):
    """
    Repairs a student's LeetCode profile after live verification and logs an audit record.
    """
    student = db.query(Student).filter(Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student record not found.")

    clean_user = extract_leetcode_username(req.new_username) or req.new_username.strip()
    if not clean_user:
        raise HTTPException(status_code=400, detail="Valid LeetCode username required.")

    # 1. Live verify profile
    profile = fetch_leetcode_profile(clean_user)
    if not profile or not profile.get("exists"):
        raise HTTPException(status_code=400, detail=f"Cannot verify '{clean_user}' on LeetCode. Please check spelling.")

    old_user = student.username
    old_url = student.leetcode_url

    # 2. Update Student record
    student.username = clean_user
    student.leetcode_url = f"https://leetcode.com/u/{clean_user}/"

    # 3. Update or create LeetCodeProfileStats
    if not student.stats:
        student.stats = LeetCodeProfileStats(student_id=student.id)

    stats = student.stats
    stats.total_solved = profile.get("total_solved", 0)
    stats.easy_solved = profile.get("easy_solved", 0)
    stats.medium_solved = profile.get("medium_solved", 0)
    stats.hard_solved = profile.get("hard_solved", 0)
    stats.contest_rating = profile.get("contest_rating")
    stats.status = "OK"
    stats.sync_status = "success"
    stats.validation_status = "verified"
    stats.error_message = None
    stats.error_code = None
    stats.last_successful_sync = datetime.datetime.utcnow()
    stats.last_verified_at = datetime.datetime.utcnow()

    # 4. Write to Audit Log
    audit = AuditLog(
        action="REPAIR_LEETCODE_PROFILE",
        performed_by=req.admin_name or "Administrator",
        target_entity=f"Student {student.reg_no}",
        details=f"Updated LeetCode username from '{old_user}' to '{clean_user}'. Solved: {stats.total_solved}.",
        timestamp=datetime.datetime.utcnow()
    )
    db.add(audit)
    db.commit()
    db.refresh(student)

    return {
        "success": True,
        "message": f"Successfully updated and verified profile for {student.name} ({student.reg_no}).",
        "student": classify_student_issue(student)
    }


class BulkSyncRequest(BaseModel):
    student_ids: List[int]

@router.post("/bulk-sync")
def bulk_sync_issues(
    req: BulkSyncRequest,
    db: Session = Depends(get_db)
):
    """
    Triggers on-demand sync for a selected batch of student IDs.
    """
    students = db.query(Student).filter(Student.id.in_(req.student_ids)).all()
    results = []

    for s in students:
        if not s.username:
            results.append({"id": s.id, "name": s.name, "status": "SKIPPED_NO_USERNAME"})
            continue

        try:
            profile = fetch_leetcode_profile(s.username)
            if profile and profile.get("exists"):
                if not s.stats:
                    s.stats = LeetCodeProfileStats(student_id=s.id)
                s.stats.total_solved = profile.get("total_solved", 0)
                s.stats.easy_solved = profile.get("easy_solved", 0)
                s.stats.medium_solved = profile.get("medium_solved", 0)
                s.stats.hard_solved = profile.get("hard_solved", 0)
                s.stats.contest_rating = profile.get("contest_rating")
                s.stats.status = "OK"
                s.stats.sync_status = "success"
                s.stats.validation_status = "verified"
                s.stats.last_successful_sync = datetime.datetime.utcnow()
                results.append({"id": s.id, "name": s.name, "status": "SYNCED", "solved": s.stats.total_solved})
            else:
                if s.stats:
                    s.stats.sync_status = "failed"
                    s.stats.status = "PROFILE NOT FOUND"
                results.append({"id": s.id, "name": s.name, "status": "NOT_FOUND"})
        except Exception as e:
            results.append({"id": s.id, "name": s.name, "status": "ERROR", "error": str(e)})

    db.commit()
    return {"total": len(req.student_ids), "synced": len([r for r in results if r["status"] == "SYNCED"]), "results": results}


@router.get("/export-excel")
def export_issues_excel(
    department: Optional[str] = Query("all"),
    year_level: Optional[str] = Query("all"),
    issue_type: Optional[str] = Query("all"),
    search: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    """
    Generates official XLSX export containing exact filtered issue records with all 15 columns.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    res = get_data_issues_students(department, year_level, issue_type, search, limit=1000, db=db)
    students_data = res["students"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Data Issues"

    # Title Block
    ws.merge_cells("A1:O1")
    title_cell = ws["A1"]
    title_cell.value = "NANDHA ENGINEERING COLLEGE (AUTONOMOUS)"
    title_cell.font = Font(name="Calibri", size=15, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="0B192C", end_color="0B192C", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Subtitle Block
    ws.merge_cells("A2:O2")
    sub_cell = ws["A2"]
    sub_cell.value = f"Student Data Quality & LeetCode Sync Issues Report • Generated: {datetime.datetime.now().strftime('%d %b %Y, %I:%M %p')} IST • Filter: Dept={department}, Year={year_level}, Issue={issue_type} ({len(students_data)} Records)"
    sub_cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    sub_cell.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # Headers
    headers = [
        "S.No", "Student Name", "Register Number", "Department", "Academic Year",
        "LeetCode Username", "LeetCode Profile URL", "URL Verification Status",
        "Issue Category", "Issue Severity", "Exact Issue Description",
        "Solved Count", "Contest Rating", "Last Successful Sync", "Recommended Action"
    ]
    ws.append([]) # Row 3 empty

    ws.append(headers) # Row 4
    header_row = ws[4]
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(name="Calibri", size=10, bold=True, color="F8FAFC")
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0")
    )

    for cell in header_row:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[4].height = 28

    # Populate Data Rows
    for idx, st in enumerate(students_data, start=1):
        row_data = [
            idx,
            st["name"],
            st["reg_no"],
            st["department_short"],
            st["year_level"],
            st["username"] or "—",
            st["leetcode_url"] or "—",
            st["url_status"],
            st["issue_label"],
            st["severity"],
            st["error_description"],
            st["total_solved"],
            st["contest_rating"] if st["contest_rating"] else "—",
            st["last_sync"],
            st["recommended_action"]
        ]
        ws.append(row_data)
        current_row = ws[ws.max_row]

        # Row styling & Zebra striping
        is_even = idx % 2 == 0
        bg_color = "F8FAFC" if is_even else "FFFFFF"
        if st["severity"] == "CRITICAL":
            bg_color = "FFF1F2" if is_even else "FFE4E6"
        elif st["severity"] == "WARNING":
            bg_color = "FFFBEB" if is_even else "FEF3C7"

        row_fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        for cell in current_row:
            cell.fill = row_fill
            cell.font = Font(name="Calibri", size=9.5)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

        ws.row_dimensions[ws.max_row].height = 22

    # Column Width Auto-Fitting
    col_widths = {
        "A": 6, "B": 24, "C": 15, "D": 22, "E": 12, "F": 18,
        "G": 35, "H": 20, "I": 22, "J": 14, "K": 42,
        "L": 12, "M": 14, "N": 24, "O": 38
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"Student_Data_Issues_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/export-csv")
def export_issues_csv(
    department: Optional[str] = Query("all"),
    year_level: Optional[str] = Query("all"),
    issue_type: Optional[str] = Query("all"),
    search: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    """
    Generates standard CSV export for filtered issue records.
    """
    import csv

    res = get_data_issues_students(department, year_level, issue_type, search, limit=1000, db=db)
    students_data = res["students"]

    buf = io.StringIO()
    writer = csv.writer(buf)

    writer.writerow([
        "S.No", "Student Name", "Register Number", "Department", "Academic Year",
        "LeetCode Username", "LeetCode Profile URL", "URL Verification Status",
        "Issue Category", "Issue Severity", "Exact Issue Description",
        "Solved Count", "Contest Rating", "Last Successful Sync", "Recommended Action"
    ])

    for idx, st in enumerate(students_data, start=1):
        writer.writerow([
            idx,
            st["name"],
            st["reg_no"],
            st["department_name"],
            st["year_level"],
            st["username"] or "",
            st["leetcode_url"] or "",
            st["url_status"],
            st["issue_label"],
            st["severity"],
            st["error_description"],
            st["total_solved"],
            st["contest_rating"] or "",
            st["last_sync"],
            st["recommended_action"]
        ])

    buf.seek(0)
    filename = f"Student_Data_Issues_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return StreamingResponse(
        io.BytesIO(buf.getvalue().encode("utf-8-sig")),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )
