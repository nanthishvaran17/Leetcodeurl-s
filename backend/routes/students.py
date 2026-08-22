from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, BackgroundTasks, Query
from sqlalchemy.orm import Session
from typing import List, Optional
import asyncio
import datetime

from backend.database import get_db
from backend.models import Student, LeetCodeProfileStats, Department, Section, AuditLog, WeeklyStudentProgress
from backend.schemas import StudentOut, StudentCreate, StudentUpdate, ContestResultOut
from backend.routes.auth import get_current_user
from backend.security import require_security_access
from backend.leetcode_client import fetch_leetcode_profile, extract_leetcode_username
from backend.excel_handler import validate_excel_import, commit_excel_import
from backend.ranking import update_all_rankings_and_badges
from backend.logger import logger

router = APIRouter(prefix="/api/students", tags=["Students"])

from sqlalchemy import func

from sqlalchemy.orm import joinedload

from backend.cache import cache
from sqlalchemy import desc, asc, nullslast

@router.get("/leaderboard-fast")
def get_leaderboard_fast(
    dept_id: Optional[int] = None,
    year_level: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """
    Ultra-fast leaderboard endpoint for the LandingPage.
    Returns slim pre-serialized JSON — ~10x smaller payload and ~10x faster than /students.
    Cache TTL: 120s. Pre-serializes to dict to avoid re-serialization on every call.
    """
    cache_key = f"leaderboard_fast:{dept_id}:{year_level}"
    cached_bytes = cache.get(cache_key)
    if cached_bytes is not None:
        from starlette.responses import Response
        return Response(content=cached_bytes, media_type="application/json")

    from sqlalchemy import text
    from backend.models import (
        WeeklyPublicResult, WeeklyVirtualResult, WeeklySession,
        LeetCodeContestRatingHistory, LeetCodeProfileStats,
        LeetCodeProfile, LeetCodeActivity, WeeklyStudentProgress
    )
    import re

    # --- Step 1: Find active target session (excludes future upcoming sessions until contest day) ---
    def _parse_session_date(d_str):
        if not d_str:
            return datetime.date.min
        try:
            parts = str(d_str).strip().split('.')
            if len(parts) == 3:
                return datetime.date(int(parts[2]), int(parts[1]), int(parts[0]))
        except Exception:
            pass
        return datetime.date.min

    today = datetime.datetime.utcnow().date()
    sessions = db.query(WeeklySession).all()
    eligible = []
    for s in sessions:
        s_date = _parse_session_date(s.session_date)
        if s.status in ["FINALIZED", "COMPLETED"]:
            eligible.append((s, s_date))
        elif s.status in ["LIVE", "ACTIVE"] and s_date <= today:
            eligible.append((s, s_date))

    def _get_c_num(item):
        s = item[0]
        m = re.search(r'\d+', s.contest_name or '')
        return int(m.group(0)) if m else (s.id or 0)

    eligible_sorted = sorted(eligible, key=_get_c_num, reverse=True)
    target_session = eligible_sorted[0][0] if eligible_sorted else db.query(WeeklySession).filter(WeeklySession.status.in_(["FINALIZED", "COMPLETED"])).order_by(WeeklySession.id.desc()).first()

    target_session_id = target_session.id if target_session else None
    target_contest_name = target_session.contest_name if target_session else "Weekly Contest"
    target_contest_date = str(target_session.session_date) if (target_session and target_session.session_date) else None
    c_num = None
    if target_session and target_session.contest_name:
        m = re.search(r'\d+', target_session.contest_name)
        if m:
            c_num = int(m.group(0))

    # --- Step 2: Load students with a single joined query ---
    query = (
        db.query(Student)
        .outerjoin(Student.stats)
        .options(
            joinedload(Student.department),
            joinedload(Student.stats),
            joinedload(Student.lc_activity),
        )
        .filter((Student.is_active == True) | (Student.is_active.is_(None)))
    )
    if dept_id:
        query = query.filter(Student.department_id == dept_id)
    if year_level and year_level.strip().upper() not in ('ALL', 'ALL YEARS', ''):
        clean_yr = year_level.strip().upper().replace('YEAR', '').strip()
        query = query.filter(func.upper(Student.year_level) == clean_yr)

    # Sort by solved desc for leaderboard
    query = query.order_by(nullslast(desc(LeetCodeProfileStats.total_solved)), Student.name.asc())
    students = query.all()

    if not students:
        empty_bytes = b'[]'
        cache.set(cache_key, empty_bytes, ttl_seconds=60, tags=["students", "leaderboard"])
        from starlette.responses import Response
        return Response(content=empty_bytes, media_type="application/json")

    student_ids = [st.id for st in students]

    # --- Step 3: Batch load all lookup tables in parallel bulk queries ---
    prog_map: dict = {}
    for p in db.query(WeeklyStudentProgress).filter(WeeklyStudentProgress.student_id.in_(student_ids)).all():
        if p.student_id not in prog_map or p.id > prog_map[p.student_id].id:
            prog_map[p.student_id] = p

    pub_map: dict = {}
    vir_map: dict = {}
    hist_map: dict = {}
    if target_session_id:
        for pr in db.query(WeeklyPublicResult).filter(
            WeeklyPublicResult.session_id == target_session_id,
            WeeklyPublicResult.student_id.in_(student_ids)
        ).all():
            pub_map[pr.student_id] = pr

        for vr in db.query(WeeklyVirtualResult).filter(
            WeeklyVirtualResult.session_id == target_session_id,
            WeeklyVirtualResult.student_id.in_(student_ids)
        ).all():
            vir_map[vr.student_id] = vr

    if c_num:
        for hr in db.query(LeetCodeContestRatingHistory).filter(
            LeetCodeContestRatingHistory.contest_name.ilike(f"%{c_num}%"),
            LeetCodeContestRatingHistory.student_id.in_(student_ids)
        ).all():
            hist_map[hr.student_id] = hr

    # --- Step 4: Build slim response dicts (no Pydantic overhead) ---
    results = []
    for st in students:
        s = st.stats
        is_verified = bool(s and s.sync_status in ("success", "verified") and s.status == "verified" and s.total_solved is not None)
        is_invalid = bool(s and (s.sync_status == "invalid_username" or s.status == "INVALID_USERNAME"))
        is_pending = bool(not st.username or not str(st.username).strip() or
                          (s and (s.sync_status == "pending_username" or s.status == "PENDING_USERNAME")))

        sync_state = "SYNCED" if is_verified else ("INVALID_USERNAME" if is_invalid else "PENDING_USERNAME")
        total_solved = s.total_solved if (is_verified and s) else None
        easy_solved = s.easy_solved if (is_verified and s) else None
        medium_solved = s.medium_solved if (is_verified and s) else None
        hard_solved = s.hard_solved if (is_verified and s) else None
        contest_rating = round(s.contest_rating, 1) if (is_verified and s and s.contest_rating) else None

        streak = 0
        if st.lc_activity and st.lc_activity.current_streak is not None:
            streak = st.lc_activity.current_streak
        elif s and s.max_streak is not None:
            streak = s.max_streak

        prog = prog_map.get(st.id)
        college_rank = prog.college_rank if (prog and is_verified) else None
        dept_rank = prog.dept_rank if (prog and is_verified) else None
        weekly_progress = prog.weekly_progress if (prog and is_verified) else 0
        if streak == 0 and prog and is_verified:
            streak = prog.streak_count or 0

        # Contest status — priority: rating history > public result
        contest_status = "NOT_ATTENDED"
        contest_solved = 0
        contest_score_display = "Not Attended"

        h = hist_map.get(st.id)
        pub = pub_map.get(st.id)
        vir = vir_map.get(st.id)

        if h and h.attended:
            contest_status = "PUBLIC_ATTENDED"
            contest_solved = h.problems_solved or 0
            contest_score_display = f"{contest_solved} / 4"
        elif h and not h.attended and (h.problems_solved or 0) > 0:
            contest_status = "VIRTUAL_ATTENDED"
            contest_solved = h.problems_solved or 0
            contest_score_display = f"{contest_solved} / 4"
        elif pub:
            is_att = pub.participation_status in ("PUBLIC", "PUBLIC_ATTENDED", "ATTENDED")
            contest_status = pub.participation_status or "NOT_ATTENDED"
            if is_att:
                contest_solved = pub.total_contest_solved or (pub.q1 + pub.q2 + pub.q3 + pub.q4)
                contest_score_display = f"{contest_solved} / 4"
        elif not st.username or not str(st.username).strip():
            contest_status = "PENDING_USERNAME"
            contest_score_display = "Data Unavailable"

        results.append({
            "id": st.id,
            "name": st.name,
            "reg_no": st.reg_no,
            "username": st.username,
            "year_level": st.year_level,
            "department_id": st.department_id,
            "department": {"id": st.department.id, "name": st.department.name, "code": st.department.code} if st.department else None,
            "sync_state": sync_state,
            "profile_url": f"https://leetcode.com/u/{st.username}/" if (is_verified and st.username) else None,
            "stats": {
                "total_solved": total_solved,
                "easy_solved": easy_solved,
                "medium_solved": medium_solved,
                "hard_solved": hard_solved,
                "contest_rating": contest_rating,
                "sync_status": s.sync_status if s else "pending_username",
                "status": s.status if s else "PENDING_USERNAME",
                "last_verified_at": s.last_verified_at.isoformat() if (s and s.last_verified_at) else None,
            } if s else None,
            "streak_count": streak,
            "college_rank": college_rank,
            "dept_rank": dept_rank,
            "weekly_progress": weekly_progress,
            "contest_status": contest_status,
            "contest_solved": contest_solved,
            "contest_score_display": contest_score_display,
            "contest_name": target_contest_name,
            "contest_number": c_num,
            "has_virtual": vir is not None and vir.participation_status in ("VIRTUAL_ATTENDED", "VIRTUAL"),
        })

    import json
    json_bytes = json.dumps(results, ensure_ascii=False, separators=(',', ':')).encode('utf-8')
    cache.set(cache_key, json_bytes, ttl_seconds=120, tags=["students", "leaderboard"])
    from starlette.responses import Response
    return Response(content=json_bytes, media_type="application/json")


@router.get("", response_model=List[StudentOut])
def get_students(
    dept_id: Optional[int] = None,
    year_level: Optional[str] = None,
    section_id: Optional[int] = None,
    search: Optional[str] = None,
    session_id: Optional[int] = None,
    sort_by: Optional[str] = Query(None, description="solved_desc, solved_asc, name_asc, name_desc, rating_desc, streak_desc"),
    min_solved: Optional[int] = None,
    max_solved: Optional[int] = None,
    verified_only: Optional[bool] = False,
    page: Optional[int] = Query(None, ge=1),
    limit: Optional[int] = Query(None, ge=1, le=500),
    db: Session = Depends(get_db)
):
    cache_key = f"students_list:{dept_id}:{year_level}:{section_id}:{search}:{session_id}:{sort_by}:{min_solved}:{max_solved}:{verified_only}:{page}:{limit}"
    cached_data = cache.get(cache_key)
    if cached_data is not None:
        from starlette.responses import Response
        return Response(content=cached_data, media_type="application/json")

    query = db.query(Student).outerjoin(Student.stats).options(
        joinedload(Student.department),
        joinedload(Student.section),
        joinedload(Student.stats),
        joinedload(Student.lc_profile),
        joinedload(Student.lc_problem_stats),
        joinedload(Student.lc_contest_standing),
        joinedload(Student.lc_activity)
    ).filter((Student.is_active == True) | (Student.is_active.is_(None)))

    if dept_id:
        query = query.filter(Student.department_id == dept_id)
    if year_level and year_level.strip().upper() not in ['ALL', 'ALL YEARS', '']:
        clean_yr = year_level.strip().upper().replace('YEAR', '').strip()
        query = query.filter(func.upper(Student.year_level) == clean_yr)

    if section_id:
        query = query.filter(Student.section_id == section_id)

    if min_solved is not None:
        query = query.filter(LeetCodeProfileStats.total_solved >= min_solved)
    if max_solved is not None:
        query = query.filter(LeetCodeProfileStats.total_solved <= max_solved)
    if verified_only:
        query = query.filter(LeetCodeProfileStats.sync_status.in_(["success", "OK", "verified"]))

    if search:
        s = f"%{search.strip()}%"
        query = query.filter(
            (Student.name.ilike(s)) |
            (Student.reg_no.ilike(s)) |
            (Student.username.ilike(s))
        )

    # Server-side sorting
    if sort_by == "solved_desc" or sort_by == "solved":
        query = query.order_by(nullslast(desc(LeetCodeProfileStats.total_solved)), Student.name.asc())
    elif sort_by == "solved_asc":
        query = query.order_by(nullslast(asc(LeetCodeProfileStats.total_solved)), Student.name.asc())
    elif sort_by == "name_desc":
        query = query.order_by(Student.name.desc())
    elif sort_by == "rating_desc" or sort_by == "rating":
        query = query.order_by(nullslast(desc(LeetCodeProfileStats.contest_rating)), Student.name.asc())
    elif sort_by == "streak_desc" or sort_by == "streak":
        query = query.order_by(nullslast(desc(LeetCodeProfileStats.max_streak)), Student.name.asc())
    else:
        query = query.order_by(Student.name.asc())

    # Pagination if page and limit provided
    if isinstance(page, int) and isinstance(limit, int) and page >= 1 and limit >= 1:
        offset = (page - 1) * limit
        students = query.offset(offset).limit(limit).all()
    elif isinstance(limit, int) and limit >= 1:
        students = query.limit(limit).all()
    else:
        students = query.all()
    
    if not students:
        return []

    # Batch fetch all student progress in 1 single query
    student_ids = [st.id for st in students]
    progs = db.query(WeeklyStudentProgress).filter(
        WeeklyStudentProgress.student_id.in_(student_ids)
    ).all()

    prog_map = {}
    for p in progs:
        if p.student_id not in prog_map or p.id > prog_map[p.student_id].id:
            prog_map[p.student_id] = p

    # Determine target session ID
    from backend.models import WeeklyPublicResult, WeeklyVirtualResult, WeeklySession, LeetCodeContestRatingHistory
    import re
    
    target_session_id = session_id
    target_session = None
    if not target_session_id:
        def _parse_session_date(d_str):
            if not d_str:
                return datetime.date.min
            try:
                parts = str(d_str).strip().split('.')
                if len(parts) == 3:
                    return datetime.date(int(parts[2]), int(parts[1]), int(parts[0]))
            except Exception:
                pass
            return datetime.date.min

        today = datetime.datetime.utcnow().date()
        sessions = db.query(WeeklySession).all()
        eligible = []
        for s in sessions:
            s_date = _parse_session_date(s.session_date)
            if s.status in ["FINALIZED", "COMPLETED"]:
                eligible.append((s, s_date))
            elif s.status in ["LIVE", "ACTIVE"] and s_date <= today:
                eligible.append((s, s_date))

        def _get_c_num(item):
            s = item[0]
            m = re.search(r'\d+', s.contest_name or '')
            return int(m.group(0)) if m else (s.id or 0)

        eligible_sorted = sorted(eligible, key=_get_c_num, reverse=True)
        target_session = eligible_sorted[0][0] if eligible_sorted else db.query(WeeklySession).filter(WeeklySession.status.in_(["FINALIZED", "COMPLETED"])).order_by(WeeklySession.id.desc()).first()
        if target_session:
            target_session_id = target_session.id
    else:
        target_session = db.query(WeeklySession).filter(WeeklySession.id == target_session_id).first()

    c_num = None
    target_contest_name = target_session.contest_name if (target_session and target_session.contest_name) else "Weekly Contest"
    if target_session and target_session.contest_name:
        m = re.search(r'\d+', target_session.contest_name)
        if m:
            c_num = int(m.group(0))

    pub_map = {}
    vir_map = {}
    hist_map = {}
    if target_session_id:
        pub_results = db.query(WeeklyPublicResult).filter(
            WeeklyPublicResult.session_id == target_session_id,
            WeeklyPublicResult.student_id.in_(student_ids)
        ).all()
        for pr in pub_results:
            pub_map[pr.student_id] = pr

        vir_results = db.query(WeeklyVirtualResult).filter(
            WeeklyVirtualResult.session_id == target_session_id,
            WeeklyVirtualResult.student_id.in_(student_ids)
        ).all()
        for vr in vir_results:
            vir_map[vr.student_id] = vr

    if c_num:
        hist_rows = db.query(LeetCodeContestRatingHistory).filter(
            LeetCodeContestRatingHistory.contest_name.ilike(f"%{c_num}%"),
            LeetCodeContestRatingHistory.student_id.in_(student_ids)
        ).all()
        for hr in hist_rows:
            hist_map[hr.student_id] = hr

    # Pre-compute canonical college ranks across all verified solvers
    verified_solvers = [st for st in students if st.stats and st.stats.total_solved is not None and st.stats.sync_status in ("success", "verified")]
    verified_solvers_sorted = sorted(verified_solvers, key=lambda x: (x.stats.total_solved or 0, x.stats.contest_rating or 0), reverse=True)
    rank_map = {st.id: r + 1 for r, st in enumerate(verified_solvers_sorted)}

    results = []
    for st in students:
        st_out = StudentOut.from_orm(st)

        # Rule 1 & 2: Canonical accuracy check — zero out fake/guessed data on invalid/pending profiles
        is_verified = bool(st.stats and st.stats.sync_status in ("success", "verified") and st.stats.status == "verified" and st.stats.total_solved is not None)
        is_invalid = bool(st.stats and (st.stats.sync_status == "invalid_username" or st.stats.status == "INVALID_USERNAME"))
        is_pending = bool(not st.username or not str(st.username).strip() or (st.stats and (st.stats.sync_status == "pending_username" or st.stats.status == "PENDING_USERNAME")))

        if is_invalid:
            st_out.leetcode_url = None
            if st_out.stats:
                st_out.stats.total_solved = None
                st_out.stats.easy_solved = None
                st_out.stats.medium_solved = None
                st_out.stats.hard_solved = None
                st_out.stats.contest_rating = None
                st_out.stats.contest_global_ranking = None
                st_out.stats.public_profile_ranking = None
                st_out.stats.status = "INVALID_USERNAME"
                st_out.stats.sync_status = "invalid_username"
                st_out.stats.validation_status = "invalid_username"
        elif is_pending:
            st_out.leetcode_url = None
            if st_out.stats:
                st_out.stats.total_solved = None
                st_out.stats.easy_solved = None
                st_out.stats.medium_solved = None
                st_out.stats.hard_solved = None
                st_out.stats.contest_rating = None
                st_out.stats.contest_global_ranking = None
                st_out.stats.public_profile_ranking = None
                st_out.stats.status = "PENDING_USERNAME"
                st_out.stats.sync_status = "pending_username"
                st_out.stats.validation_status = "pending_username"

        # Canonical fields from normalized tables if available
        if st.lc_profile:
            st_out.canonical_username = st.lc_profile.canonical_username if is_verified else None
            st_out.profile_url = st.lc_profile.profile_url if is_verified else None
            st_out.real_name = st.lc_profile.real_name if is_verified else None
            st_out.avatar_url = st.lc_profile.avatar_url if is_verified else None
            st_out.sync_state = st.lc_profile.sync_state
        else:
            st_out.canonical_username = st.username if is_verified else None
            st_out.profile_url = f"https://leetcode.com/u/{st.username}/" if (is_verified and st.username) else None
            st_out.sync_state = "SYNCED" if is_verified else ("INVALID_USERNAME" if is_invalid else "PENDING_USERNAME")

        if st.lc_activity and st.lc_activity.current_streak is not None:
            st_out.streak_count = st.lc_activity.current_streak or 0
            st_out.longest_streak = st.lc_activity.longest_streak or 0
            st_out.total_active_days = st.lc_activity.total_active_days or 0
        elif st.stats and st.stats.max_streak is not None:
            st_out.streak_count = st.stats.max_streak
            st_out.longest_streak = st.stats.max_streak
            st_out.total_active_days = st.stats.active_days or 0

        latest_prog = prog_map.get(st.id)
        if latest_prog:
            st_out.college_rank = latest_prog.college_rank if (latest_prog.college_rank and is_verified) else rank_map.get(st.id)
            st_out.dept_rank = latest_prog.dept_rank if is_verified else None
            st_out.year_rank = latest_prog.year_rank if is_verified else None
            st_out.section_rank = latest_prog.section_rank if is_verified else None
            st_out.weekly_progress = latest_prog.weekly_progress if is_verified else 0
            if (st_out.streak_count is None or st_out.streak_count == 0) and latest_prog.streak_count:
                st_out.streak_count = latest_prog.streak_count if is_verified else 0
            st_out.consistency_score = latest_prog.consistency_score if is_verified else 0.0
            st_out.badge_list = latest_prog.badge_list or []
        else:
            st_out.college_rank = rank_map.get(st.id) if is_verified else None

        pub_res = pub_map.get(st.id)
        vir_res = vir_map.get(st.id)
        h_res = hist_map.get(st.id)

        target_contest_name = target_session.contest_name if target_session else "Weekly Contest"
        target_contest_date = target_session.session_date if target_session else None

        if h_res and h_res.attended:
            tot_solved = h_res.problems_solved or 0
            st_out.overall_participation_mode = "PUBLIC"
            st_out.contest_status = "PUBLIC_ATTENDED"
            st_out.public_contest_result = ContestResultOut(
                contest_name=target_contest_name,
                contest_number=c_num,
                contest_date=target_contest_date,
                questions_solved=tot_solved,
                questions_total=4,
                score_display=f"{tot_solved} / 4",
                contest_rank=h_res.contest_rank,
                contest_rating=round(h_res.rating_after, 1) if h_res.rating_after else None,
                top_percentage=None,
                status="PUBLIC_ATTENDED",
                fetched_at=datetime.datetime.utcnow().isoformat()
            )
        elif h_res and not h_res.attended and (h_res.problems_solved or 0) > 0:
            tot_solved = h_res.problems_solved or 0
            st_out.overall_participation_mode = "VIRTUAL"
            st_out.contest_status = "VIRTUAL_ATTENDED"
            st_out.public_contest_result = ContestResultOut(
                contest_name=target_contest_name,
                contest_number=c_num,
                contest_date=target_contest_date,
                questions_solved=tot_solved,
                questions_total=4,
                score_display=f"{tot_solved} / 4",
                contest_rank=None,
                contest_rating=None,
                top_percentage=None,
                status="VIRTUAL_ATTENDED",
                fetched_at=datetime.datetime.utcnow().isoformat()
            )
        elif pub_res:
            tot_solved = pub_res.total_contest_solved or (pub_res.q1 + pub_res.q2 + pub_res.q3 + pub_res.q4)
            is_att = pub_res.participation_status in ("PUBLIC", "PUBLIC_ATTENDED", "ATTENDED")
            is_not_att = pub_res.participation_status in ("NOT_ATTENDED", "PUBLIC_NOT_ATTENDED")
            score_disp = f"{tot_solved} / 4" if is_att else ("Not Attended" if is_not_att else "Data Unavailable")
            st_out.overall_participation_mode = "PUBLIC" if is_att else "NONE"
            st_out.contest_status = pub_res.participation_status or "NOT_ATTENDED"
            st_out.public_contest_result = ContestResultOut(
                contest_name=pub_res.session.contest_name if pub_res.session else target_contest_name,
                contest_number=c_num,
                contest_date=pub_res.session.session_date if pub_res.session else target_contest_date,
                questions_solved=tot_solved if is_att else 0,
                questions_total=4,
                score_display=score_disp,
                contest_rank=pub_res.contest_rank,
                contest_rating=pub_res.contest_rating,
                top_percentage=None,
                status=pub_res.participation_status or "NOT_ATTENDED",
                fetched_at=pub_res.last_fetched_at.isoformat() if pub_res.last_fetched_at else None
            )
        else:
            has_uname = bool(st.username and st.username.strip())
            st_out.overall_participation_mode = "NONE"
            st_out.contest_status = "NOT_ATTENDED" if has_uname else "PENDING_USERNAME"
            st_out.public_contest_result = ContestResultOut(
                contest_name=target_contest_name,
                contest_number=c_num,
                contest_date=target_contest_date,
                questions_solved=0,
                questions_total=4,
                score_display="Not Attended" if has_uname else "Data Unavailable",
                contest_rank=None,
                contest_rating=None,
                top_percentage=None,
                status="NOT_ATTENDED" if has_uname else "UNKNOWN",
                fetched_at=None
            )

        if vir_res:
            tot_solved_v = vir_res.total_contest_solved or (vir_res.q1 + vir_res.q2 + vir_res.q3 + vir_res.q4)
            st_out.virtual_contest_result = ContestResultOut(
                contest_name=vir_res.session.contest_name if vir_res.session else "Weekly Contest",
                contest_number=None,
                contest_date=vir_res.session.session_date if vir_res.session else None,
                questions_solved=tot_solved_v,
                questions_total=4,
                score_display=f"{tot_solved_v} / 4" if vir_res.participation_status in ("VIRTUAL_ATTENDED", "VIRTUAL") else "Not Attended",
                contest_rank=getattr(vir_res, 'contest_rank', None),
                contest_rating=getattr(vir_res, 'contest_rating', None),
                top_percentage=getattr(vir_res, 'top_percentage', None),
                status=vir_res.participation_status or "NO_VIRTUAL_RECORD",
                fetched_at=getattr(vir_res, 'completed_at', None).isoformat() if getattr(vir_res, 'completed_at', None) else None
            )
        else:
            st_out.virtual_contest_result = ContestResultOut(
                contest_name="Weekly Contest",
                contest_number=None,
                contest_date=None,
                questions_solved=0,
                questions_total=4,
                score_display="Not Attended",
                contest_rank=None,
                contest_rating=None,
                top_percentage=None,
                status="NO_VIRTUAL_RECORD",
                fetched_at=None
            )

        results.append(st_out)

    cache.set(cache_key, results, ttl_seconds=30, tags=["students"])
    return results

@router.get("/sample-excel")
def download_sample_student_excel():
    """
    Generates and returns Student_Import_Sample.xlsx with exact required columns:
    REG NO | NAME | DEPT | YEAR | LEETCODE PROFILE LINK
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    ws.sheet_view.showGridLines = True

    headers = ["REG NO", "NAME", "DEPT", "YEAR", "LEETCODE PROFILE LINK"]
    col_widths = [18, 28, 14, 10, 45]

    navy_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    font_header = Font(name="Times New Roman", size=11, bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    for col_idx, (h_text, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h_text)
        cell.fill = navy_fill
        cell.font = font_header
        cell.alignment = center
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = w

    ws.row_dimensions[1].height = 26

    sample_rows = [
        ["732224CC001", "AJAY A", "CSE(CS)", "III", "https://leetcode.com/u/example_student/"],
        ["732224CC002", "AMRUTHA M", "CSE(CS)", "III", "https://leetcode.com/u/example_student2/"],
        ["732224CI001", "BHARATH K", "CSE(IOT)", "III", "https://leetcode.com/u/example_student3/"],
    ]

    for row_idx, r_data in enumerate(sample_rows, start=2):
        ws.row_dimensions[row_idx].height = 20
        for col_idx, val in enumerate(r_data, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.font = Font(name="Times New Roman", size=10)
            c.alignment = center if col_idx in (1, 3, 4) else left
            c.border = thin_border

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{len(sample_rows)+1}"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from fastapi.responses import Response
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="Student_Import_Sample.xlsx"'}
    )

@router.get("/{student_id}", response_model=StudentOut)
def get_student_detail(student_id: int, db: Session = Depends(get_db)):
    student = db.query(Student).filter(Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    
    st_out = StudentOut.from_orm(student)
    latest_prog = db.query(WeeklyStudentProgress).filter(WeeklyStudentProgress.student_id == student.id).order_by(WeeklyStudentProgress.id.desc()).first()
    if latest_prog:
        st_out.college_rank = latest_prog.college_rank
        st_out.dept_rank = latest_prog.dept_rank
        st_out.year_rank = latest_prog.year_rank
        st_out.section_rank = latest_prog.section_rank
        st_out.weekly_progress = latest_prog.weekly_progress
        st_out.streak_count = latest_prog.streak_count
        st_out.consistency_score = latest_prog.consistency_score
        st_out.badge_list = latest_prog.badge_list or []
        
    return st_out

@router.post("", response_model=StudentOut)
def create_student(
    student_in: StudentCreate,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user=Depends(require_security_access(resource_name="Create Student", required_roles=["admin", "super admin", "hod"]))
):
    existing = db.query(Student).filter(Student.reg_no == student_in.reg_no.upper()).first()
    if existing:
        raise HTTPException(status_code=400, detail=f"Student with Register No '{student_in.reg_no}' already exists.")

    username, std_url, url_status = extract_leetcode_username(student_in.leetcode_url)
    # Normalise username to lowercase for consistent lookup
    if username:
        username = username.lower()
        std_url = f"https://leetcode.com/u/{username}/"

    student = Student(
        reg_no=student_in.reg_no.upper(),
        name=student_in.name,
        department_id=student_in.department_id,
        year_level=student_in.year_level,
        section_id=student_in.section_id,
        email=student_in.email,
        leetcode_url=std_url if std_url else student_in.leetcode_url,
        username=username,
        codeforces_username=student_in.codeforces_username,
        hackerrank_username=student_in.hackerrank_username
    )
    db.add(student)
    db.commit()
    db.refresh(student)

    # Init stats row — sync_status starts as "pending" until background sync runs
    stats = LeetCodeProfileStats(
        student_id=student.id,
        status=url_status,
        sync_status="pending" if username else "not_started",
        validation_status="pending" if username else "not_started",
    )
    db.add(stats)

    audit = AuditLog(
        user_id=current_user.id,
        user_name=current_user.username,
        action="CREATE_STUDENT",
        details=f"Created student {student.reg_no} ({student.name}) username={username or 'none'}"
    )
    db.add(audit)
    db.commit()

    # Trigger background verification+sync immediately after creation so the profile
    # is verified without blocking the HTTP response.  The background task opens its
    # own DB session — it must NEVER reuse the request session.
    if username:
        student_id_for_sync = student.id
        def _bg_sync_new_student():
            from backend.database import SessionLocal as _SL
            from backend.services.live_sync_service import sync_single_student as _sss
            _bg_db = _SL()
            try:
                _sss(student_id_for_sync, _bg_db)
            except Exception as _e:
                logger.warning(f"[CREATE_STUDENT_SYNC] Background sync note for student_id={student_id_for_sync}: {_e}")
            finally:
                _bg_db.close()
        background_tasks.add_task(_bg_sync_new_student)
        logger.info(f"[CREATE_STUDENT] Background sync queued for new student {student.reg_no} (username={username})")

    return StudentOut.from_orm(student)

from pydantic import BaseModel

class BulkDeleteRequest(BaseModel):
    student_ids: List[int]
    soft_delete: Optional[bool] = False

@router.post("/bulk-delete")
def bulk_delete_students(
    req: BulkDeleteRequest,
    db: Session = Depends(get_db),
    current_user=Depends(require_security_access(resource_name="Bulk Delete Students", required_roles=["admin", "super admin"]))
):
    if not req.student_ids:
        raise HTTPException(status_code=400, detail="No student IDs provided for deletion.")

    count = len(req.student_ids)

    if req.soft_delete:
        db.query(Student).filter(Student.id.in_(req.student_ids)).update({"is_active": False}, synchronize_session=False)
        action_name = "BULK_DEACTIVATE_STUDENTS"
        msg = f"Successfully deactivated {count} student records."
    else:
        db.query(LeetCodeProfileStats).filter(LeetCodeProfileStats.student_id.in_(req.student_ids)).delete(synchronize_session=False)
        db.query(WeeklyStudentProgress).filter(WeeklyStudentProgress.student_id.in_(req.student_ids)).delete(synchronize_session=False)
        db.query(Student).filter(Student.id.in_(req.student_ids)).delete(synchronize_session=False)
        action_name = "BULK_DELETE_STUDENTS"
        msg = f"Successfully deleted {count} student records."

    audit = AuditLog(
        user_id=current_user.id,
        user_name=current_user.username,
        action=action_name,
        details=msg
    )
    db.add(audit)
    db.commit()

    update_all_rankings_and_badges(db)
    cache.clear()

    return {"message": msg, "count": count}

class StudentUpdateSchema(BaseModel):
    name: Optional[str] = None
    department_id: Optional[int] = None
    year_level: Optional[str] = None
    section_id: Optional[int] = None
    email: Optional[str] = None
    leetcode_url: Optional[str] = None
    username: Optional[str] = None
    is_active: Optional[bool] = True


@router.patch("/{student_id}")
@router.put("/{student_id}")
def update_student(
    student_id: int,
    payload: StudentUpdateSchema,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user=Depends(require_security_access(resource_name="Update Student", required_roles=["admin", "super admin", "hod"]))
):
    student = db.query(Student).filter(Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student record not found.")

    old_username = student.username

    if payload.name and payload.name.strip():
        student.name = payload.name.strip()
    if payload.department_id is not None:
        student.department_id = payload.department_id
    if payload.year_level and payload.year_level.strip():
        student.year_level = payload.year_level.strip().upper()
    if payload.section_id is not None:
        student.section_id = payload.section_id
    if payload.email is not None:
        student.email = payload.email.strip().lower() if payload.email else None

    # ── LeetCode URL / username normalisation ─────────────────────────────────
    url_changed = False
    old_url = student.leetcode_url
    if payload.leetcode_url is not None:
        raw_lc = payload.leetcode_url.strip() if payload.leetcode_url else None
        if raw_lc != old_url:
            url_changed = True
            student.leetcode_url = raw_lc
            if raw_lc:
                _parsed_u, _parsed_url, _u_status = extract_leetcode_username(raw_lc)
                if _parsed_u:
                    student.username = _parsed_u.lower()
                    student.leetcode_url = _parsed_url  # canonical URL
                else:
                    student.username = None

    # Direct username override (e.g. from the username field in the edit form)
    if payload.username and payload.username.strip():
        _direct_u = payload.username.strip().lower()
        if _direct_u != (old_username or "").strip().lower():
            url_changed = True
            student.username = _direct_u
            student.leetcode_url = f"https://leetcode.com/u/{_direct_u}/"

    if payload.is_active is not None:
        student.is_active = payload.is_active

    old_u_norm = (old_username or "").strip().lower()
    new_u_norm = (student.username or "").strip().lower()
    username_changed = bool(url_changed or (new_u_norm and old_u_norm != new_u_norm))

    if username_changed:
        from backend.leetcode_fetcher import clear_leetcode_cache
        if old_username:
            clear_leetcode_cache(old_username)
        if student.username:
            clear_leetcode_cache(student.username)

        if student.stats:
            student.stats.sync_status = "fetching"
            student.stats.status = "FETCHING"
            student.stats.validation_status = "fetching"
            student.stats.error_message = None

    # Commit updated URL/username to DB first before starting fetch (Requirement 17)
    db.commit()
    db.refresh(student)

    # Perform synchronous fresh fetch for immediate UI update & response
    if username_changed:
        from backend.services.live_sync_service import sync_single_student as _sss
        try:
            _sss(student.id, db, force_refresh=True)
            db.refresh(student)
        except Exception as _sync_err:
            logger.warning(f"[UPDATE_STUDENT_SYNC] Direct sync note for student_id={student.id}: {_sync_err}")

    # ── Cloud Firestore sync (best-effort) ────────────────────────────────────
    try:
        from backend.services.firestore_service import update_firestore_doc
        update_firestore_doc("students", student.reg_no, {
            "reg_no": student.reg_no,
            "name": student.name,
            "username": student.username,
            "leetcode_url": student.leetcode_url,
            "year_level": student.year_level,
            "is_active": student.is_active,
            "updated_at": datetime.datetime.utcnow().isoformat() + "Z"
        })
    except Exception as fs_err:
        logger.warning(f"[FIRESTORE UPDATE NOTE] {fs_err}")

    audit = AuditLog(
        user_id=current_user.id,
        user_name=current_user.username,
        action="UPDATE_STUDENT",
        details=(
            f"Updated student {student.reg_no} ({student.name})"
            + (f" | username: '{old_username}' -> '{student.username}' (re-sync queued)" if username_changed else "")
        )
    )
    db.add(audit)
    db.commit()

    update_all_rankings_and_badges(db)
    cache.clear()

    return StudentOut.from_orm(student)


@router.delete("/{student_id}")
def delete_student(
    student_id: int,
    soft_delete: bool = Query(True),
    db: Session = Depends(get_db),
    current_user=Depends(require_security_access(resource_name="Delete Student", required_roles=["admin", "super admin"]))
):
    student = db.query(Student).filter(Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student record not found.")

    reg_no = student.reg_no
    name = student.name

    if soft_delete:
        student.is_active = False
        db.commit()
        logger.info(f"[SOFT_DELETE_STUDENT] Soft-deleted student roster record {reg_no} ({name})")
    else:
        db.query(LeetCodeProfileStats).filter(LeetCodeProfileStats.student_id == student_id).delete()
        db.query(WeeklyStudentProgress).filter(WeeklyStudentProgress.student_id == student_id).delete()
        db.delete(student)
        db.commit()

    # Sync status to Cloud Firestore
    try:
        from backend.services.firestore_service import update_firestore_doc
        update_firestore_doc("students", reg_no, {
            "is_active": False,
            "deactivated_at": datetime.datetime.utcnow().isoformat() + "Z"
        })
    except Exception as fs_err:
        logger.warning(f"[FIRESTORE DELETE NOTE] {fs_err}")

    audit = AuditLog(
        user_id=current_user.id,
        user_name=current_user.username,
        action="SOFT_DELETE_STUDENT" if soft_delete else "DELETE_STUDENT",
        details=f"Deactivated student roster record {reg_no} ({name})"
    )
    db.add(audit)
    db.commit()

    update_all_rankings_and_badges(db)
    cache.clear()

    return {"message": f"Successfully deactivated student roster record {reg_no} ({name})", "reg_no": reg_no}


@router.post("/import-preview")
async def import_preview(
    file: UploadFile = File(...), 
    db: Session = Depends(get_db),
    current_user=Depends(require_security_access(resource_name="Import Preview", required_roles=["admin", "super admin", "hod"]))
):
    content = await file.read()
    report = validate_excel_import(db, content)
    return report

@router.post("/import-commit")
def import_commit(
    valid_rows: List[dict],
    db: Session = Depends(get_db),
    current_user=Depends(require_security_access(resource_name="Import Commit", required_roles=["admin", "super admin", "hod"]))
):
    imported_count = commit_excel_import(db, valid_rows)
    audit = AuditLog(user_id=current_user.id, user_name=current_user.username, action="EXCEL_IMPORT", details=f"Imported {imported_count} students from Excel.")
    db.add(audit)
    db.commit()
    
    # Recalculate ranks
    update_all_rankings_and_badges(db)
    
    return {"message": f"Successfully imported {imported_count} students.", "count": imported_count}

from backend.sync_engine import run_batch_sync, sync_single_student_by_id, sync_tracker

@router.get("/sync-status")
@router.get("/admin/sync/status/{run_id}")
def get_students_sync_status(run_id: Optional[str] = None):
    return sync_tracker.to_dict()

@router.post("/{student_id}/refresh")
async def refresh_single_student(student_id: int):
    """
    Refreshes single student statistics within target 30-second limit.
    """
    try:
        result = await sync_single_student_by_id(student_id, timeout=30.0)
        if result.get("status") == "failed":
            raise HTTPException(status_code=400, detail=result.get("error", "Sync failed"))
        return {
            "message": f"Refreshed stats for {result.get('name')}",
            "status": result.get("status"),
            "last_verified_at": result.get("last_verified_at"),
            "stats": result.get("stats")
        }
    except ValueError as val_err:
        raise HTTPException(status_code=404, detail=str(val_err))
    except Exception as err:
        raise HTTPException(status_code=500, detail=f"Sync error: {err}")

@router.post("/refresh-all")
@router.post("/sync-all")
@router.post("/admin/sync/start")
async def refresh_all_students(
    background_tasks: BackgroundTasks,
    limit: Optional[int] = None
):
    """
    Starts async background sync worker for all 273 students without blocking browser.
    Returns immediately with runId. Frontend subscribes to Firestore syncRuns/{runId} for progress.
    """
    if sync_tracker.is_running:
        existing_run_id = sync_tracker.run_id or "current"
        return {
            "runId": existing_run_id,
            "message": "Live stats refresh is already running in background.",
            "status": "busy",
            "progress": sync_tracker.to_dict()
        }

    # Pre-generate a deterministic runId so the frontend can subscribe to Firestore immediately
    run_id = f"sync_{datetime.datetime.utcnow().strftime('%Y%m%d_%H%M%S')}"
    background_tasks.add_task(run_batch_sync, limit=limit, pre_run_id=run_id)
    db = SessionLocal()
    try:
        active_count = db.query(Student).filter((Student.is_active == True) | (Student.is_active.is_(None))).count()
    except Exception:
        active_count = 300
    finally:
        db.close()

    return {
        "runId": run_id,
        "status": "started",
        "total": active_count,
        "message": f"Live stats batch sync started in background for {active_count} active students!",
        "sync_status_url": f"/api/students/admin/sync/status/{run_id}"
    }


# ─────────────────────────────────────────────────────────────────────────────
# LEETCODE ACCOUNT VALIDATION ENDPOINT
# Read-only: validates the account exists + identity matches, no DB writes.
# ─────────────────────────────────────────────────────────────────────────────

class LeetCodeValidateRequest(BaseModel):
    leetcode_url: Optional[str] = None
    username: Optional[str] = None

@router.post("/{student_id}/validate-leetcode")
async def validate_leetcode_account(
    student_id: int,
    payload: LeetCodeValidateRequest,
    db: Session = Depends(get_db),
    current_user=Depends(require_security_access(
        resource_name="Validate LeetCode Account",
        required_roles=["admin", "super admin", "hod", "faculty"]
    ))
):
    """
    Validates a LeetCode username/URL against the live LeetCode API.

    Flow:
      1. Normalise + validate the format.
      2. Fetch the public profile via LeetCode GraphQL (same as the sync engine).
      3. Verify the returned canonical username matches the requested one.
      4. Return a structured validation status — does NOT write to the database.

    Use this before saving a new/changed LeetCode username so the UI can surface
    'Validating → Fetching → Verified ✓' / 'Username not found' feedback without
    waiting for a full background sync to complete.
    """
    # student_id=0 is a sentinel for "validate a new account before creation" —
    # no student lookup is needed in that case.  For any real student_id (>0),
    # we confirm the student exists so we can log context.
    if student_id > 0:
        student = db.query(Student).filter(Student.id == student_id).first()
        if not student:
            raise HTTPException(status_code=404, detail="Student record not found.")

    # Resolve input: explicit username takes precedence over URL
    raw_input = (payload.username or "").strip() or (payload.leetcode_url or "").strip()
    if not raw_input:
        return {
            "validation_status": "INVALID_FORMAT",
            "message": "No username or LeetCode URL provided.",
            "username": None,
            "canonical_url": None,
            "can_save": False,
        }

    # Step 1 — Format validation (no network call yet)
    parsed_username, parsed_url, u_status = extract_leetcode_username(raw_input)
    if u_status != "OK" or not parsed_username:
        return {
            "validation_status": "INVALID_FORMAT",
            "message": f"Invalid LeetCode username or URL format: {u_status}",
            "username": None,
            "canonical_url": None,
            "can_save": False,
        }

    parsed_username = parsed_username.lower()

    # Step 2 — Live LeetCode fetch (identity-checked inside fetch_leetcode_profile)
    try:
        result = await asyncio.wait_for(
            fetch_leetcode_profile(parsed_username, force_refresh=True, timeout=20.0, max_retries=2),
            timeout=25.0
        )
    except asyncio.TimeoutError:
        return {
            "validation_status": "NETWORK_ERROR",
            "message": "LeetCode API did not respond in time. Please try again.",
            "username": parsed_username,
            "canonical_url": parsed_url,
            "can_save": False,
        }
    except Exception as exc:
        logger.warning(f"[VALIDATE_LC] Unexpected error for '{parsed_username}': {exc}")
        return {
            "validation_status": "FETCH_FAILED",
            "message": f"Could not reach LeetCode API: {exc}",
            "username": parsed_username,
            "canonical_url": parsed_url,
            "can_save": False,
        }

    # Step 3 — Map fetcher status to structured validation status
    fetcher_status = result.get("status", "")

    if fetcher_status in ("success", "OK"):
        canonical = result.get("username", parsed_username)
        return {
            "validation_status": "VALID",
            "message": f"LeetCode account '{canonical}' verified successfully.",
            "username": canonical,
            "canonical_url": f"https://leetcode.com/u/{canonical}/",
            "can_save": True,
            "profile_data": {
                "total_solved": result.get("total_solved"),
                "contest_rating": result.get("contest_rating"),
                "recent_contest": result.get("recent_contest_name"),
            }
        }

    elif fetcher_status == "INVALID_USERNAME":
        return {
            "validation_status": "ACCOUNT_NOT_FOUND",
            "message": f"No LeetCode account found for username '{parsed_username}'.",
            "username": parsed_username,
            "canonical_url": None,
            "can_save": False,
        }

    elif fetcher_status == "IDENTITY_MISMATCH":
        return {
            "validation_status": "IDENTITY_MISMATCH",
            "message": result.get("error_message") or "LeetCode returned a different username than requested.",
            "username": parsed_username,
            "canonical_url": None,
            "can_save": False,
        }

    elif fetcher_status == "FETCH_FAILED":
        err = result.get("error_message") or "Profile fetch failed."
        # Distinguish rate-limit from generic failure
        if "429" in err or "rate" in err.lower():
            return {
                "validation_status": "RATE_LIMITED",
                "message": "LeetCode is temporarily rate-limiting requests. Please wait a minute and try again.",
                "username": parsed_username,
                "canonical_url": None,
                "can_save": False,
            }
        return {
            "validation_status": "FETCH_FAILED",
            "message": err,
            "username": parsed_username,
            "canonical_url": None,
            "can_save": False,
        }

    else:
        return {
            "validation_status": "FETCH_FAILED",
            "message": result.get("error_message") or f"Unexpected fetch status: {fetcher_status}",
            "username": parsed_username,
            "canonical_url": None,
            "can_save": False,
        }
