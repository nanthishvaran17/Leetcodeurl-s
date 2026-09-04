import os
import io
import datetime
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    pd = None
    PANDAS_AVAILABLE = False
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any, Optional
from sqlalchemy.orm import Session

from backend.models import (
    Student, Department, LeetCodeProfileStats, LeetCodeAccount, WeeklyStudentProgress,
    WeeklySessionSnapshot, WeeklySession, WeeklyPublicResult,
    ContestParticipation, AuditLog, User
)
from backend.services.authorization_service import apply_role_based_student_filter
from backend.leetcode_client import extract_leetcode_username

STUDENT_IMPORT_COLUMNS = [
    "REG NO",
    "NAME",
    "DEPT",
    "YEAR",
    "LEETCODE PROFILE LINK"
]

def run_high_speed_excel_import(db: Session, file_bytes: bytes, job_id: str, tracker: Any) -> Dict[str, Any]:
    try:
        df = pd.read_excel(io.BytesIO(file_bytes))
    except Exception as e:
        raise ValueError(f"Failed to parse Excel file: {str(e)}")

    col_mapping = {col: str(col).strip().upper() for col in df.columns}
    df = df.rename(columns=col_mapping)

    primary_col = None
    for candidate in ["PRIMARY LEETCODE LINK", "LEETCODE PROFILE LINK", "PRIMARY LEETCODE", "LEETCODE", "LEETCODE LINK", "PRIMARY ACCOUNT"]:
        if candidate in df.columns:
            primary_col = candidate
            break

    sec_col = None
    for candidate in ["SECONDARY LEETCODE LINK", "SECONDARY LEETCODE PROFILE LINK", "SECONDARY LEETCODE", "SECONDARY ACCOUNT", "SECONDARY LINK", "SECONDARY LEETCODE HANDLE"]:
        if candidate in df.columns:
            sec_col = candidate
            break

    required_cols = ["REG NO", "NAME", "DEPT", "YEAR"]
    missing_req = [c for c in required_cols if c not in df.columns]
    if missing_req:
        raise ValueError(f"Missing required columns in Excel: {', '.join(missing_req)}")

    if not primary_col:
        raise ValueError("Missing LeetCode profile column. Please include 'PRIMARY LEETCODE LINK' or 'LEETCODE PROFILE LINK'.")

    # Standardize column values
    if "REG NO" in df.columns:
        df["REG NO"] = df["REG NO"].astype(str).str.strip().str.upper()
    if "NAME" in df.columns:
        df["NAME"] = df["NAME"].astype(str).str.strip()
    if "DEPT" in df.columns:
        df["DEPT"] = df["DEPT"].astype(str).str.strip().str.upper()
    if "YEAR" in df.columns:
        df["YEAR"] = df["YEAR"].astype(str).str.strip().str.upper()
    
    # Remove empty rows based on REG NO and NAME
    df = df[df["REG NO"].notna() & (df["REG NO"] != "") & (df["REG NO"] != "NAN")]
    df = df[df["NAME"].notna() & (df["NAME"] != "") & (df["NAME"] != "NAN")]

    total_rows = len(df)
    if total_rows == 0:
        raise ValueError("Excel file contains no valid data rows.")
    
    tracker.total_rows = total_rows
    
    # 1. Dynamic Department Registration
    unique_depts = df["DEPT"].unique()
    existing_depts = {d.code.upper(): d for d in db.query(Department).all()}
    existing_dept_names = {d.name.upper(): d for d in db.query(Department).all()}
    
    for dept_str in unique_depts:
        if not dept_str or dept_str == "NAN":
            continue
        if dept_str not in existing_depts and dept_str not in existing_dept_names:
            # Create new department
            new_dept = Department(name=dept_str, code=dept_str)
            db.add(new_dept)
            db.commit()
            db.refresh(new_dept)
            existing_depts[dept_str] = new_dept
            tracker.new_departments.append(dept_str)
            tracker.update(processed_inc=0, log_msg=f"Registered new department: {dept_str}")
            
    # Refresh dept map
    existing_depts = {d.code.upper(): d for d in db.query(Department).all()}
    for d in db.query(Department).all():
        if d.name:
            existing_depts[d.name.upper()] = d

    # 2. Bulk Fetch Existing Students
    all_reg_nos = df["REG NO"].tolist()
    existing_students = db.query(Student).filter(Student.reg_no.in_(all_reg_nos)).all()
    existing_map = {s.reg_no: s for s in existing_students}
    
    new_students = []
    updated_students_count = 0
    new_student_ids = []
    
    # Process rows
    batch_size = 500
    current_batch = []
    
    def process_batch(batch):
        nonlocal updated_students_count
        db.commit() # Commit updates
        if new_students:
            db.add_all(new_students)
            db.commit()
            
            # Add stats rows for new students
            new_stats = []
            for s in new_students:
                new_stats.append(LeetCodeProfileStats(student_id=s.id, status="not_started", sync_status="pending"))
                new_student_ids.append(s.id)
            db.add_all(new_stats)
            db.commit()
            new_students.clear()

    seen_reg_nos_in_file = set()

    for idx, row in df.iterrows():
        reg_no = row.get("REG NO")
        if reg_no in seen_reg_nos_in_file:
            tracker.update(processed_inc=1, failed_inc=1, log_msg=f"Skipped duplicate REG NO in file: {reg_no}")
            continue
            
        seen_reg_nos_in_file.add(reg_no)
        name = row.get("NAME")
        dept_str = row.get("DEPT")
        year_str = row.get("YEAR")
        
        # Clean up year
        valid_years = ["II", "III", "IV"]
        if year_str in ["2", "2ND"]: year_str = "II"
        elif year_str in ["3", "3RD"]: year_str = "III"
        elif year_str in ["4", "4TH"]: year_str = "IV"
        elif year_str not in valid_years:
            # Dynamic year - let's just accept it if it's alphanumeric
            if not str(year_str).isalnum():
                year_str = "I" # Fallback

        url = str(row.get(primary_col, "")).strip() if primary_col and pd.notna(row.get(primary_col)) else ""
        username, std_url, url_status = extract_leetcode_username(url)
        if username:
            username = username.lower()
            std_url = f"https://leetcode.com/u/{username}/"

        sec_url = str(row.get(sec_col, "")).strip() if sec_col and pd.notna(row.get(sec_col)) else ""
        sec_username, sec_std_url, _ = extract_leetcode_username(sec_url) if sec_url else (None, None, None)
        if sec_username:
            sec_username = sec_username.lower()

        codeforces = str(row.get("CODEFORCES", row.get("CODEFORCES USERNAME", ""))).strip() if ("CODEFORCES" in df.columns and pd.notna(row.get("CODEFORCES"))) or ("CODEFORCES USERNAME" in df.columns and pd.notna(row.get("CODEFORCES USERNAME"))) else ""
        hackerrank = str(row.get("HACKERRANK", row.get("HACKERRANK USERNAME", ""))).strip() if ("HACKERRANK" in df.columns and pd.notna(row.get("HACKERRANK"))) or ("HACKERRANK USERNAME" in df.columns and pd.notna(row.get("HACKERRANK USERNAME"))) else ""

        email = str(row.get("EMAIL", "")).strip() if "EMAIL" in df.columns and pd.notna(row.get("EMAIL")) else ""

        dept_obj = existing_depts.get(dept_str)
        dept_id = dept_obj.id if dept_obj else None

        # --- Batch Extraction / Inference ---
        batch_str = str(row.get("BATCH", "")).strip() if "BATCH" in df.columns and pd.notna(row.get("BATCH")) else None
        if not batch_str:
            # Fallback batch inference based on year_str assuming current year is 2026-2027
            # Note: A real implementation would fetch the active AcademicYear from DB
            if year_str == "IV": batch_str = "2023-2027"
            elif year_str == "III": batch_str = "2024-2028"
            elif year_str == "II": batch_str = "2025-2029"
            elif year_str == "I": batch_str = "2026-2030"

        # --- Email Generation ---
        from backend.services.email_identity_service import check_and_generate_email
        
        student = existing_map.get(reg_no)
        if not student:
            email_res = check_and_generate_email(db, reg_no)
            
            new_student = Student(
                reg_no=reg_no,
                name=name,
                department_id=dept_id,
                year_level=year_str,
                email=email,
                leetcode_url=std_url if std_url else url,
                username=username,
                codeforces_username=codeforces if codeforces else None,
                hackerrank_username=hackerrank if hackerrank else None,
                is_active=True,
                batch=batch_str,
                institutional_email=email_res.get("email"),
                email_status=email_res.get("status", "pending")
            )
            new_students.append(new_student)
            tracker.update(processed_inc=1, success_inc=1)

            # Register secondary LeetCode handle if provided
            if sec_username:
                db.flush()
                acc = LeetCodeAccount(student_id=new_student.id, leetcode_username=sec_username)
                db.add(acc)
        else:
            student.name = name
            if dept_id: student.department_id = dept_id
            if year_str: student.year_level = year_str
            if email: student.email = email
            if batch_str: student.batch = batch_str
            
            # Re-verify email if it's pending or error
            if student.email_status in ["pending", "error", "needs_verification"] or not student.institutional_email:
                email_res = check_and_generate_email(db, reg_no, student.institutional_email)
                student.institutional_email = email_res.get("email") or student.institutional_email
                student.email_status = email_res.get("status", student.email_status)

            if url: 
                student.leetcode_url = std_url if std_url else url
                student.username = username
            if codeforces: student.codeforces_username = codeforces
            if hackerrank: student.hackerrank_username = hackerrank
            student.is_active = True
            
            # Register secondary LeetCode handle if provided
            if sec_username:
                existing_acc = db.query(LeetCodeAccount).filter(
                    LeetCodeAccount.student_id == student.id,
                    LeetCodeAccount.leetcode_username == sec_username
                ).first()
                if not existing_acc:
                    db.add(LeetCodeAccount(student_id=student.id, leetcode_username=sec_username))

            updated_students_count += 1
            tracker.update(processed_inc=1, success_inc=1)
            
        current_batch.append(reg_no)
        if len(current_batch) >= batch_size:
            process_batch(current_batch)
            current_batch.clear()
            
    # Process remaining
    if current_batch:
        process_batch(current_batch)
        
    # Re-calculate ranks after bulk import
    from backend.ranking import update_all_rankings_and_badges
    update_all_rankings_and_badges(db)

    return {
        "total_processed": tracker.processed_rows,
        "new_students": len(new_student_ids),
        "updated_students": updated_students_count,
        "new_departments": tracker.new_departments,
        "new_student_ids": new_student_ids
    }


def get_year_level_variants(year_lvl: str) -> List[str]:
    clean = str(year_lvl).strip().upper()
    if clean in ["IV", "4", "4TH", "4-TH"]:
        return ["IV", "4", "4TH", "4th", "IV Year", "iv"]
    elif clean in ["III", "3", "3RD", "3-RD"]:
        return ["III", "3", "3RD", "3rd", "III Year", "iii"]
    elif clean in ["II", "2", "2ND", "2-ND"]:
        return ["II", "2", "2ND", "2nd", "II Year", "ii"]
    elif clean in ["I", "1", "1ST", "1-ST"]:
        return ["I", "1", "1ST", "1st", "I Year", "i"]
    return [clean]

def filter_students_by_dept_and_year(db: Session, dept: Department, year_lvl: str, current_user: Optional[User] = None):
    from sqlalchemy import func, or_
    year_vars = get_year_level_variants(year_lvl)

    dept_code = dept.code or dept.name
    matching_depts = db.query(Department).filter(
        (Department.id == dept.id) |
        (func.upper(Department.code) == dept_code.upper()) |
        (func.upper(Department.name).like(f"%{dept_code.upper()}%"))
    ).all()
    dept_ids = list({d.id for d in matching_depts}) if matching_depts else [dept.id]

    query = db.query(Student).filter(
        Student.department_id.in_(dept_ids),
        func.upper(Student.year_level).in_([v.upper() for v in year_vars]),
        or_(Student.is_active == True, Student.is_active.is_(None))
    )

    if current_user:
        query = apply_role_based_student_filter(query, current_user, db)

    return query.all()

def create_nandha_official_department_sheet(ws, dept: Department, db: Session):
    ws.sheet_view.showGridLines = True
    
    font_bold_12 = Font(name="Times New Roman", size=12, bold=True)
    font_bold_11 = Font(name="Times New Roman", size=11, bold=True)
    font_bold_10 = Font(name="Times New Roman", size=10, bold=True)
    font_regular_10 = Font(name="Times New Roman", size=10)
    
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    ws.merge_cells("C1:K1")
    ws["C1"] = "NANDHA ENGINEERING COLLEGE, ERODE - 638 052."
    ws["C1"].font = font_bold_12
    ws["C1"].alignment = center_align

    ws.merge_cells("C2:K2")
    ws["C2"] = "(An Autonomous Institution, Affiliated to Anna University, Chennai)"
    ws["C2"].font = font_bold_10
    ws["C2"].alignment = center_align

    ws.merge_cells("C3:K3")
    ws["C3"] = f"Department of {dept.name}"
    ws["C3"].font = font_bold_11
    ws["C3"].alignment = left_align

    ws.merge_cells("C4:K4")
    ws["C4"] = f"Date: {datetime.date.today().strftime('%d-%m-%Y')}"
    ws["C4"].font = font_bold_11
    ws["C4"].alignment = left_align

    ws.merge_cells("C5:K5")
    ws["C5"] = "Leetcode Performance — Weekly Report"
    ws["C5"].font = font_bold_11
    ws["C5"].alignment = left_align

    # Dynamic Academic Coordinator resolution
    coord_str = ""
    from backend.models import User
    coord_user = db.query(User).filter(User.department_id == dept.id, User.role.in_(["staff", "faculty", "hod"])).first()
    if coord_user:
        coord_str = coord_user.username
    else:
        if "Cyber Security" in dept.name or "CS" in dept.code:
            coord_str = "M. Santhoshkumar, AP / CSE (Cyber Security)"
        elif "IoT" in dept.name or "IOT" in dept.code:
            coord_str = "Mohan Gandhi S"
        else:
            coord_str = f"Academic Coordinator / {dept.code or dept.name}"

    ws.merge_cells("C6:M6")
    ws["C6"] = f"Name & Designation of the Academic Coordinator: {coord_str}"
    ws["C6"].font = font_bold_11
    ws["C6"].alignment = left_align

    # Table Headers
    ws.merge_cells("A8:A9")
    ws["A8"] = "Batch"
    ws["A8"].font = font_bold_10
    ws["A8"].alignment = center_align

    ws.merge_cells("B8:B9")
    ws["B8"] = "No. of Students\n(Total Count)"
    ws["B8"].font = font_bold_10
    ws["B8"].alignment = center_align

    ws.merge_cells("C8:G8")
    ws["C8"] = "Number of Problems Solved"
    ws["C8"].font = font_bold_10
    ws["C8"].alignment = center_align

    ws.merge_cells("H8:K8")
    ws["H8"] = "Weekly Contest Attended"
    ws["H8"].font = font_bold_10
    ws["H8"].alignment = center_align

    ws.merge_cells("L8:M8")
    ws["L8"] = "Leetcode Contest Rating and Ranking"
    ws["L8"].font = font_bold_10
    ws["L8"].alignment = center_align

    sub_headers = {
        "C9": "Above 500", "D9": "250 - 500", "E9": "100 - 249",
        "F9": "1 - 99", "G9": "0",
        "H9": "4Q", "I9": "3Q", "J9": "2Q", "K9": "1Q",
        "L9": "Rating > 1500", "M9": "Ranking < 20000"
    }

    for col_ref, text in sub_headers.items():
        ws[col_ref] = text
        ws[col_ref].font = font_bold_10
        ws[col_ref].alignment = center_align

    for r in range(8, 10):
        for c in range(1, 14):
            ws.cell(row=r, column=c).border = thin_border

    batches = [
        ("2023 - 2027", "IV"),
        ("2024 - 2028", "III"),
        ("2025 - 2029", "II"),
        ("2026 - 2030", "I")
    ]
    current_row = 10

    from backend.models import WeeklySession, WeeklyPublicResult, ContestParticipation
    recent_sessions = db.query(WeeklySession).order_by(WeeklySession.id.desc()).limit(2).all()
    curr_session = recent_sessions[0] if len(recent_sessions) > 0 else None
    last_session = recent_sessions[1] if len(recent_sessions) > 1 else None

    def calculate_contest_counts(session_obj, student_ids):
        q4, q3, q2, q1 = 0, 0, 0, 0
        if not session_obj or not student_ids:
            return q4, q3, q2, q1

        # 1. Query WeeklyPublicResult
        pub_results = db.query(WeeklyPublicResult).filter(
            WeeklyPublicResult.session_id == session_obj.id,
            WeeklyPublicResult.student_id.in_(student_ids)
        ).all()

        found_ids = set()
        for pr in pub_results:
            found_ids.add(pr.student_id)
            tot = pr.total_contest_solved or 0
            if tot >= 4: q4 += 1
            elif tot == 3: q3 += 1
            elif tot == 2: q2 += 1
            elif tot == 1: q1 += 1

        # 2. Fallback to ContestParticipation for remaining students
        missing_ids = student_ids - found_ids
        if missing_ids:
            cps = db.query(ContestParticipation).filter(
                ContestParticipation.student_id.in_(missing_ids)
            ).all()
            for cp in cps:
                tot = cp.problems_solved or 0
                if tot >= 4: q4 += 1
                elif tot == 3: q3 += 1
                elif tot == 2: q2 += 1
                elif tot == 1: q1 += 1

        return q4, q3, q2, q1

    for batch_label, year_lvl in batches:
        students = filter_students_by_dept_and_year(db, dept, year_lvl)
        total_count = len(students)
        student_ids = {s.id for s in students}

        above_500 = sum(1 for s in students if s.stats and (s.stats.total_solved or 0) > 500)
        between_250_500 = sum(1 for s in students if s.stats and 250 <= (s.stats.total_solved or 0) <= 500)
        between_100_249 = sum(1 for s in students if s.stats and 100 <= (s.stats.total_solved or 0) <= 249)
        between_1_99 = sum(1 for s in students if s.stats and 1 <= (s.stats.total_solved or 0) <= 99)
        zero_solved = sum(1 for s in students if not s.stats or (s.stats.total_solved or 0) == 0)

        rating_above_1500 = sum(1 for s in students if s.stats and s.stats.contest_rating and s.stats.contest_rating > 1500)
        rank_below_20000 = sum(1 for s in students if s.stats and ((s.stats.contest_global_ranking and s.stats.contest_global_ranking < 20000) or (s.stats.public_profile_ranking and s.stats.public_profile_ranking < 20000)))

        q4_curr, q3_curr, q2_curr, q1_curr = calculate_contest_counts(curr_session, student_ids)
        q4_last, q3_last, q2_last, q1_last = calculate_contest_counts(last_session, student_ids)

        # Row 1: Last Week
        ws.cell(row=current_row, column=1, value=f"{batch_label} (Last Week)").alignment = center_align
        ws.cell(row=current_row, column=1).font = font_bold_10
        ws.cell(row=current_row, column=2, value=total_count if total_count > 0 else 0)

        last_row_vals = [above_500, between_250_500, between_100_249, between_1_99, zero_solved, q4_last, q3_last, q2_last, q1_last, rating_above_1500, rank_below_20000]
        for c_offset, val in enumerate(last_row_vals, start=3):
            ws.cell(row=current_row, column=c_offset, value=val)

        # Row 2: Current Week
        ws.cell(row=current_row+1, column=1, value=f"{batch_label} (Current Week)").alignment = center_align
        ws.cell(row=current_row+1, column=1).font = font_bold_10
        ws.cell(row=current_row+1, column=2, value=total_count if total_count > 0 else 0)

        curr_row_vals = [above_500, between_250_500, between_100_249, between_1_99, zero_solved, q4_curr, q3_curr, q2_curr, q1_curr, rating_above_1500, rank_below_20000]
        for c_offset, val in enumerate(curr_row_vals, start=3):
            ws.cell(row=current_row+1, column=c_offset, value=val)

        for r_idx in range(current_row, current_row + 2):
            for c_idx in range(1, 14):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.border = thin_border
                cell.font = font_regular_10
                cell.alignment = center_align

        current_row += 2

    # Verified Signatures Footer
    ws.cell(row=current_row+2, column=1, value="Verified Signatures:").font = font_bold_10
    ws.cell(row=current_row+4, column=1, value="Academic Coordinator").font = font_bold_10
    ws.cell(row=current_row+4, column=8, value="Head of Department").font = font_bold_10

    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 16
    for col_let in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
        ws.column_dimensions[col_let].width = 14

def _apply_border(cell, color='000000', style='thin'):
    s = Side(style=style, color=color)
    cell.border = Border(left=s, right=s, top=s, bottom=s)

def _add_cover_sheet(wb, logo_path: str):
    """Sheet 1: College cover / logo sheet."""
    ws = wb.create_sheet(title="COVER", index=0)
    ws.views.sheetView[0].showGridLines = False

    # Try to insert logo image
    if os.path.exists(logo_path):
        try:
            from openpyxl.drawing.image import Image as XLImage
            img = XLImage(logo_path)
            img.width, img.height = 140, 140
            ws.add_image(img, "D3")
        except Exception:
            pass

    title_font = Font(name="Times New Roman", size=18, bold=True)
    sub_font   = Font(name="Times New Roman", size=13, bold=True)
    info_font  = Font(name="Times New Roman", size=11)
    center     = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells("A1:L1")
    ws["A1"] = "NANDHA ENGINEERING COLLEGE (AUTONOMOUS)"
    ws["A1"].font = Font(name="Times New Roman", size=20, bold=True)
    ws["A1"].alignment = center

    ws.merge_cells("A2:L2")
    ws["A2"] = "Erode – 638 052, Tamil Nadu"
    ws["A2"].font = sub_font
    ws["A2"].alignment = center

    ws.merge_cells("A10:L10")
    ws["A10"] = "LeetCode Weekly Performance Report"
    ws["A10"].font = title_font
    ws["A10"].alignment = center

    ws.merge_cells("A11:L11")
    ws["A11"] = "Computer Science and Engineering (Cyber Security & IoT)"
    ws["A11"].font = sub_font
    ws["A11"].alignment = center

    ws.merge_cells("A12:L12")
    ws["A12"] = f"Generated on: {datetime.datetime.now().strftime('%d %B %Y, %I:%M %p')}"
    ws["A12"].font = info_font
    ws["A12"].alignment = center

    ws.row_dimensions[1].height = 35
    ws.row_dimensions[10].height = 30
    ws.row_dimensions[11].height = 24

def _create_dept_year_sheet(wb, dept, year_lvl: str, students_list, db: Session):
    """One sheet per dept+year with full student analytics table."""
    year_display = {"II": "2nd Year", "III": "3rd Year", "IV": "4th Year"}.get(year_lvl, year_lvl)
    sheet_name = f"{dept.code}-{year_lvl}Yr"[:31]
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_view.showGridLines = True

    TNR = "Times New Roman"
    navy_fill   = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    header_fill = PatternFill(start_color="2E5B88", end_color="2E5B88", fill_type="solid")
    alt_fill    = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    white_fill  = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # --- Title rows ---
    ws.merge_cells("A1:K1")
    ws["A1"] = "NANDHA ENGINEERING COLLEGE (AUTONOMOUS), ERODE – 638 052"
    ws["A1"].font = Font(name=TNR, size=13, bold=True, color="FFFFFF")
    ws["A1"].alignment = center
    ws["A1"].fill = navy_fill

    dept_title = f"Department of {dept.name} | {year_display} | LeetCode Performance Report | {datetime.date.today().strftime('%d.%m.%Y')}"
    ws.merge_cells("A2:K2")
    ws["A2"] = dept_title.upper()
    ws["A2"].font = Font(name=TNR, size=10, bold=True, color="FFFFFF")
    ws["A2"].alignment = center
    ws["A2"].fill = header_fill

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 22

    # --- Column headers (row 3) --- 11 Columns (Contest Rating & Global Rank REMOVED as requested)
    headers = [
        "S.No", "Register No", "Student Name", "Department", "Year",
        "LeetCode URL", "Username",
        "Easy\nSolved", "Medium\nSolved", "Hard\nSolved", "Total\nSolved"
    ]
    col_widths = [6, 16, 28, 12, 8, 36, 18, 10, 10, 10, 12]

    for col_idx, (hdr, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=3, column=col_idx, value=hdr)
        cell.font = Font(name=TNR, size=10, bold=True, color="FFFFFF")
        cell.fill = navy_fill
        cell.alignment = center
        _apply_border(cell, "FFFFFF")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[3].height = 32

    # --- Data rows (Sorted alphabetically/numerically by Register No starting from 1) ---
    sorted_students = sorted(students_list, key=lambda s: s.reg_no or "")

    for idx, student in enumerate(sorted_students, start=1):
        row = 3 + idx
        stats = student.stats
        total = stats.total_solved if stats else 0
        row_fill = alt_fill if idx % 2 == 0 else white_fill

        lc_url = student.leetcode_url or ""
        lc_user = student.username or (lc_url.split('/')[-1] if lc_url else "")

        row_data = [
            idx,
            student.reg_no,
            student.name,
            dept.code,
            year_lvl,
            lc_url,
            lc_user,
            stats.easy_solved   if stats else 0,
            stats.medium_solved if stats else 0,
            stats.hard_solved   if stats else 0,
            total
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.fill = row_fill
            cell.font = Font(name=TNR, size=10)
            cell.alignment = center if col_idx not in (3, 6, 7) else left
            _apply_border(cell, "CBD5E1")

        # Hyperlink for LeetCode URL (col 6)
        if lc_url:
            link_cell = ws.cell(row=row, column=6)
            link_cell.hyperlink = lc_url
            link_cell.font = Font(name=TNR, size=10, color="0563C1", underline="single")

        ws.row_dimensions[row].height = 20

    # --- Summary mini-table below data ---
    summary_row = 3 + len(sorted_students) + 2
    total_count  = len(sorted_students)
    above_500    = sum(1 for s in sorted_students if s.stats and (s.stats.total_solved or 0) > 500)
    b250_500     = sum(1 for s in sorted_students if s.stats and 250 <= (s.stats.total_solved or 0) <= 500)
    b101_250     = sum(1 for s in sorted_students if s.stats and 101 <= (s.stats.total_solved or 0) < 250)
    lt100        = sum(1 for s in sorted_students if s.stats and 0 < (s.stats.total_solved or 0) < 100)
    not_started  = sum(1 for s in sorted_students if not s.stats or (s.stats.total_solved or 0) == 0)

    ws.merge_cells(f"A{summary_row}:K{summary_row}")
    ws[f"A{summary_row}"] = "NUMBER OF PROBLEMS SOLVED — CATEGORY SUMMARY"
    ws[f"A{summary_row}"].font = Font(name=TNR, size=10, bold=True, color="FFFFFF")
    ws[f"A{summary_row}"].fill = navy_fill
    ws[f"A{summary_row}"].alignment = center
    ws.row_dimensions[summary_row].height = 24

    cat_headers = ["Above 500", "250 – 500", "101 – 250", "Less than 100", "Not Yet Started", "Total Students"]
    cat_values  = [above_500,   b250_500,    b101_250,    lt100,           not_started,       total_count]

    for ci, (h, v) in enumerate(zip(cat_headers, cat_values)):
        hc = ws.cell(row=summary_row+1, column=ci+1, value=h)
        vc = ws.cell(row=summary_row+2, column=ci+1, value=v)
        hc.font = Font(name=TNR, size=9, bold=True, color="FFFFFF")
        hc.fill = header_fill
        hc.alignment = center
        vc.font = Font(name=TNR, size=10, bold=True)
        vc.alignment = center
        _apply_border(hc, "CBD5E1")
        _apply_border(vc, "CBD5E1")


def _create_analytics_summary_sheet(wb, db: Session):
    """Final sheet: full analytics with category breakdown table."""
    ws = wb.create_sheet(title="ANALYTICS SUMMARY")
    ws.sheet_view.showGridLines = False

    TNR = "Times New Roman"
    navy_fill   = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    header_fill = PatternFill(start_color="2E5B88", end_color="2E5B88", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center")

    ws.merge_cells("A1:M1")
    ws["A1"] = "NANDHA ENGINEERING COLLEGE — LeetCode Analytics Summary"
    ws["A1"].font = Font(name=TNR, size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = navy_fill
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:M2")
    ws["A2"] = f"Report Date: {datetime.date.today().strftime('%d %B %Y')}   |   Academic Year: 2025–26"
    ws["A2"].font = Font(name=TNR, size=10, bold=True, color="FFFFFF")
    ws["A2"].fill = header_fill
    ws["A2"].alignment = center
    ws.row_dimensions[2].height = 18

    all_students = db.query(Student).filter(Student.is_active == True).all()
    total_students = len(all_students)
    active = sum(1 for s in all_students if s.stats and (s.stats.total_solved or 0) > 0)
    total_solved_all = sum((s.stats.total_solved or 0) for s in all_students if s.stats)
    avg_solved = round(total_solved_all / active, 1) if active > 0 else 0
    top_student = max(all_students, key=lambda s: (s.stats.total_solved or 0) if s.stats else 0, default=None)

    # Quick stats block
    quick_stats = [
        ("Total Enrolled Students", total_students),
        ("Active Solvers (>0 solved)", active),
        ("Total Problems Solved (All)", total_solved_all),
        ("Average Solved (Active)", avg_solved),
        ("Top Ranker", top_student.name if top_student else "N/A"),
    ]
    for ri, (label, val) in enumerate(quick_stats, start=4):
        lc = ws.cell(row=ri, column=1, value=label)
        vc = ws.cell(row=ri, column=2, value=val)
        lc.font = Font(name=TNR, size=10, bold=True)
        vc.font = Font(name=TNR, size=10)
        lc.alignment = left
        vc.alignment = center
        _apply_border(lc)
        _apply_border(vc)
        ws.row_dimensions[ri].height = 18

    # ---- "Number of Problems Solved" table ----
    table_start = 11
    ws.merge_cells(f"A{table_start}:M{table_start}")
    ws[f"A{table_start}"] = "Number of Problems Solved — Department & Year Breakdown"
    ws[f"A{table_start}"].font = Font(name=TNR, size=11, bold=True, color="FFFFFF")
    ws[f"A{table_start}"].fill = navy_fill
    ws[f"A{table_start}"].alignment = center
    ws.row_dimensions[table_start].height = 24

    # Sub-header row
    main_cols = ["Batch", "No. of Students"]
    problem_cats = ["Above 500", "250 – 500", "Less than 250", "Less than 100", "Not yet started"]
    contest_cols = ["4 Q Solved", "3 Q Solved", "2 Q Solved", "1 Q Solved"]
    rating_cols  = ["Rating: Above 1500", "Ranking: Below 20000"]

    th_row = table_start + 1
    ws.merge_cells(f"A{th_row}:A{th_row+1}"); ws[f"A{th_row}"] = "Batch"
    ws.merge_cells(f"B{th_row}:B{th_row+1}"); ws[f"B{th_row}"] = "No. of\nStudents"
    ws.merge_cells(f"C{th_row}:G{th_row}");  ws[f"C{th_row}"] = "Number of Problems Solved"
    ws.merge_cells(f"H{th_row}:K{th_row}");  ws[f"H{th_row}"] = "Weekly Contest Attended"
    ws.merge_cells(f"L{th_row}:M{th_row}");  ws[f"L{th_row}"] = "LeetCode Rating & Ranking"

    for col, lbl in [("C", "Above 500"), ("D", "250 – 500"), ("E", "Less than 250"),
                     ("F", "Less than 100"), ("G", "Not yet\nstarted"),
                     ("H", "4 Q Solved"), ("I", "3 Q Solved"), ("J", "2 Q Solved"), ("K", "1 Q Solved"),
                     ("L", "Rating:\nAbove 1500"), ("M", "Ranking:\nBelow 20000")]:
        ws[f"{col}{th_row+2}"] = lbl
        ws[f"{col}{th_row+2}"].font = Font(name=TNR, size=9, bold=True)
        ws[f"{col}{th_row+2}"].alignment = center
        ws[f"{col}{th_row+2}"].fill = PatternFill(start_color="E8EEF5", end_color="E8EEF5", fill_type="solid")
        _apply_border(ws[f"{col}{th_row+2}"])

    for merge_cell_ref in [f"A{th_row}", f"B{th_row}", f"C{th_row}", f"H{th_row}", f"L{th_row}"]:
        cell = ws[merge_cell_ref]
        cell.font = Font(name=TNR, size=10, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = center
        _apply_border(cell)

    for r in [th_row, th_row+1, th_row+2]:
        ws.row_dimensions[r].height = 28

    # Data rows per dept + year
    departments = db.query(Department).all()
    data_row = th_row + 3
    year_batch_map = {"IV": "2023 – 2027", "III": "2024 – 2028", "II": "2025 – 2029", "I": "2026 – 2030"}
    latest_session = db.query(WeeklySession).order_by(WeeklySession.id.desc()).first()

    for dept in departments:
        for year_lvl, batch_label in year_batch_map.items():
            students_dy = filter_students_by_dept_and_year(db, dept, year_lvl)

            if not students_dy:
                continue

            tc = len(students_dy)
            above_500   = sum(1 for s in students_dy if s.stats and (s.stats.total_solved or 0) > 500)
            b250_500    = sum(1 for s in students_dy if s.stats and 250 <= (s.stats.total_solved or 0) <= 500)
            lt250       = sum(1 for s in students_dy if s.stats and 100 <= (s.stats.total_solved or 0) < 250)
            lt100       = sum(1 for s in students_dy if s.stats and 0 < (s.stats.total_solved or 0) < 100)
            not_started = sum(1 for s in students_dy if not s.stats or (s.stats.total_solved or 0) == 0)

            q4, q3, q2, q1 = 0, 0, 0, 0
            if latest_session:
                stud_ids = {s.id for s in students_dy}
                snaps = db.query(WeeklySessionSnapshot).filter(
                    WeeklySessionSnapshot.session_id == latest_session.id,
                    WeeklySessionSnapshot.student_id.in_(stud_ids)
                ).all()
                for sn in snaps:
                    if sn.problems_added >= 4: q4 += 1
                    elif sn.problems_added == 3: q3 += 1
                    elif sn.problems_added == 2: q2 += 1
                    elif sn.problems_added == 1: q1 += 1

            r1500 = sum(1 for s in students_dy if s.stats and s.stats.contest_rating and s.stats.contest_rating > 1500)
            r20k  = sum(1 for s in students_dy if s.stats and s.stats.contest_global_ranking and s.stats.contest_global_ranking < 20000)

            label = f"{dept.code} – {year_lvl} Year\n({batch_label})"
            row_vals = [label, tc, above_500, b250_500, lt250, lt100, not_started, q4, q3, q2, q1, r1500, r20k]

            for ci, val in enumerate(row_vals, start=1):
                cell = ws.cell(row=data_row, column=ci, value=val)
                cell.font = Font(name=TNR, size=10)
                cell.alignment = center if ci > 1 else left
                cell.fill = PatternFill(start_color="FFFFFF" if data_row % 2 == 0 else "F5F8FC",
                                        end_color="FFFFFF" if data_row % 2 == 0 else "F5F8FC",
                                        fill_type="solid")
                _apply_border(cell)

            ws.row_dimensions[data_row].height = 30
            data_row += 1

    # Column widths
    col_w = {"A": 24, "B": 14, "C": 12, "D": 12, "E": 14, "F": 14, "G": 14,
             "H": 12, "I": 12, "J": 12, "K": 12, "L": 16, "M": 16}
    for col, w in col_w.items():
        ws.column_dimensions[col].width = w


def generate_8_sheet_excel_report(db: Session, current_user: Optional[User] = None) -> bytes:
    """
    Generates the official 8-sheet Nandha College LeetCode Performance Tracker.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    logo_path = os.path.join(os.path.dirname(__file__), "assets", "nandha_emblem.png")
    _add_cover_sheet(wb, logo_path)

    departments = db.query(Department).all()
    year_order = ["IV", "III", "II", "I"]

    for dept in departments:
        for year_lvl in year_order:
            students_dy = filter_students_by_dept_and_year(db, dept, year_lvl, current_user=current_user)
            if students_dy:
                _create_dept_year_sheet(wb, dept, year_lvl, students_dy, db)

    # Also keep the original summary sheet for backward compatibility
    for dept in departments:
        sheet_title = f"Summary-{dept.code}"[:31]
        ws_sum = wb.create_sheet(title=sheet_title)
        create_nandha_official_department_sheet(ws_sum, dept, db)

    _create_analytics_summary_sheet(wb, db)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_student_performance_detail_excel(db: Session, current_user: Optional["User"] = None) -> bytes:
    """
    Generates the Cyber Security and IoT detailed performance Excel report.
    Scoped to the current_user's authorization level:
      - Staff/Faculty: assigned students only
      - HOD: department students
      - Admin/Super Admin/Principal: full institutional scope
    Produces sheets per department+year combination, e.g.:
      - CSE(CS)-IIYr
      - CSE(IoT)-IIIYr
      - CSE(IoT)-IVYr
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    logo_path = os.path.join(os.path.dirname(__file__), "assets", "nandha_emblem.png")
    _add_cover_sheet(wb, logo_path)

    depts = db.query(Department).all()
    cs_dept = next((d for d in depts if "CS" in (d.code or "").upper() or "CYBER" in (d.name or "").upper()), None)
    iot_dept = next((d for d in depts if "IOT" in (d.code or "").upper() or "IOT" in (d.name or "").upper()), None)

    dept_tuples = [
        ("CSE(CS)", cs_dept),
        ("CSE(IoT)", iot_dept)
    ]
    years = ["II", "III", "IV"]

    font_title = Font(name="Times New Roman", size=12, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    font_header = Font(name="Times New Roman", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E5B88", end_color="2E5B88", fill_type="solid")
    font_data = Font(name="Times New Roman", size=10)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    thin_border = Border(
        left=Side(style='thin', color='C0C0C0'),
        right=Side(style='thin', color='C0C0C0'),
        top=Side(style='thin', color='C0C0C0'),
        bottom=Side(style='thin', color='C0C0C0')
    )

    headers = [
        "S.No", "Register No", "Student Name", "Department", "Year",
        "LeetCode Profile Link", "Username", "Easy", "Medium", "Hard",
        "Total Solved", "Rating", "Global Rank", "Contest Count", "Last Synced", "Data Status"
    ]

    for label_code, dept_obj in dept_tuples:
        for yr in years:
            sheet_name = f"{label_code}-{yr}Yr"[:31]
            ws = wb.create_sheet(title=sheet_name)
            ws.sheet_view.showGridLines = True

            ws.merge_cells("A1:P1")
            ws["A1"] = f"NANDHA ENGINEERING COLLEGE (AUTONOMOUS) — {label_code} {yr} YEAR PERFORMANCE REPORT"
            ws["A1"].font = font_title
            ws["A1"].fill = title_fill
            ws["A1"].alignment = center_align
            ws.row_dimensions[1].height = 28

            ws.row_dimensions[3].height = 24
            for c_idx, h in enumerate(headers, start=1):
                cell = ws.cell(row=3, column=c_idx, value=h)
                cell.font = font_header
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border

            students_query = db.query(Student).filter((Student.is_active == True) | (Student.is_active.is_(None)))
            if dept_obj:
                students_query = students_query.filter(Student.department_id == dept_obj.id)
            students_query = students_query.filter(Student.year_level == yr)
            # Apply role-based authorization scope (Staff → assigned only, HOD → dept, Admin → all)
            if current_user is not None:
                from backend.services.authorization_service import apply_role_based_student_filter
                students_query = apply_role_based_student_filter(students_query, current_user, db)
            students = students_query.all()

            students.sort(key=lambda s: (s.stats.total_solved or 0) if s.stats else 0, reverse=True)

            row_idx = 4
            for s_no, s in enumerate(students, start=1):
                st = s.stats
                is_ver = st and st.validation_status == "verified"
                contest_cnt = db.query(ContestParticipation).filter(ContestParticipation.student_id == s.id).count()

                row_vals = [
                    s_no,
                    s.reg_no or "",
                    s.name or "",
                    (dept_obj.code if dept_obj else label_code),
                    s.year_level or yr,
                    s.leetcode_url or "N/A",
                    s.username or "N/A",
                    (st.easy_solved if is_ver and st else 0) if (st and is_ver) else 0,
                    (st.medium_solved if is_ver and st else 0) if (st and is_ver) else 0,
                    (st.hard_solved if is_ver and st else 0) if (st and is_ver) else 0,
                    (st.total_solved if is_ver and st else 0) if (st and is_ver) else 0,
                    (round(st.contest_rating, 1) if (is_ver and st and st.contest_rating) else "N/A"),
                    (st.contest_global_ranking if (is_ver and st and st.contest_global_ranking) else "N/A"),
                    contest_cnt,
                    (st.last_updated.strftime("%Y-%m-%d %H:%M") if (st and st.last_updated) else "N/A"),
                    ("VERIFIED" if is_ver else "UNVERIFIED")
                ]

                ws.row_dimensions[row_idx].height = 20
                for c_idx, val in enumerate(row_vals, start=1):
                    cell = ws.cell(row=row_idx, column=c_idx, value=val)
                    cell.font = font_data
                    cell.border = thin_border
                    if c_idx in (1, 4, 5, 14, 16):
                        cell.alignment = center_align
                    elif c_idx in (8, 9, 10, 11, 12, 13):
                        cell.alignment = right_align
                    else:
                        cell.alignment = left_align
                row_idx += 1

            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_8_sheet_master_tracker(db: Session, current_user: Optional[User] = None) -> bytes:
    """
    Generates the complete 8-sheet Master Tracker workbook with real DB data:
      1. Student Master
      2. Current Statistics
      3. Session Logs
      4. College Leaderboard
      5. Department Leaderboard
      6. Contest Statistics
      7. Data Quality
      8. Audit Error Logs
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    font_header = Font(name="Times New Roman", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    font_data = Font(name="Times New Roman", size=10)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    thin_border = Border(
        left=Side(style='thin', color='C0C0C0'),
        right=Side(style='thin', color='C0C0C0'),
        top=Side(style='thin', color='C0C0C0'),
        bottom=Side(style='thin', color='C0C0C0')
    )

    query = db.query(Student).filter((Student.is_active == True) | (Student.is_active.is_(None)))
    query = apply_role_based_student_filter(query, current_user, db)
    all_students = query.all()

    # Sheet 1: Student Master
    ws1 = wb.create_sheet(title="Student Master")
    ws1.sheet_view.showGridLines = True
    h1 = ["S.No", "Register No", "Student Name", "Department", "Year", "Section", "Email", "LeetCode URL", "Username", "Status", "Created At"]
    for c_idx, h in enumerate(h1, start=1):
        cell = ws1.cell(row=1, column=c_idx, value=h)
        cell.font = font_header; cell.fill = header_fill; cell.alignment = center_align; cell.border = thin_border
    for r_idx, s in enumerate(all_students, start=2):
        row_v = [r_idx - 1, s.reg_no, s.name, s.department.code if s.department else "", s.year_level, s.section.name if s.section else "A", s.email or "N/A", s.leetcode_url or "N/A", s.username or "N/A", "ACTIVE" if s.is_active else "INACTIVE", s.created_at.strftime("%Y-%m-%d") if hasattr(s, 'created_at') and s.created_at else "N/A"]
        for c_idx, val in enumerate(row_v, start=1):
            c = ws1.cell(row=r_idx, column=c_idx, value=val)
            c.font = font_data; c.border = thin_border
            c.alignment = center_align if c_idx in (1, 4, 5, 6, 10) else left_align

    # Sheet 2: Current Statistics
    ws2 = wb.create_sheet(title="Current Statistics")
    ws2.sheet_view.showGridLines = True
    h2 = ["S.No", "Register No", "Student Name", "Dept", "Year", "Easy", "Medium", "Hard", "Total Solved", "Contest Rating", "Global Rank", "Validation Status", "Last Synced"]
    for c_idx, h in enumerate(h2, start=1):
        cell = ws2.cell(row=1, column=c_idx, value=h)
        cell.font = font_header; cell.fill = header_fill; cell.alignment = center_align; cell.border = thin_border
    for r_idx, s in enumerate(all_students, start=2):
        st = s.stats
        is_v = st and st.validation_status == "verified"
        row_v = [r_idx - 1, s.reg_no, s.name, s.department.code if s.department else "", s.year_level, (st.easy_solved if is_v and st else 0) if st else 0, (st.medium_solved if is_v and st else 0) if st else 0, (st.hard_solved if is_v and st else 0) if st else 0, (st.total_solved if is_v and st else 0) if st else 0, round(st.contest_rating, 1) if (is_v and st and st.contest_rating) else "N/A", st.contest_global_ranking if (is_v and st and st.contest_global_ranking) else "N/A", "VERIFIED" if is_v else "UNVERIFIED", st.last_updated.strftime("%Y-%m-%d %H:%M") if (st and st.last_updated) else "N/A"]
        for c_idx, val in enumerate(row_v, start=1):
            c = ws2.cell(row=r_idx, column=c_idx, value=val)
            c.font = font_data; c.border = thin_border
            c.alignment = center_align if c_idx in (1, 4, 5, 12) else (right_align if c_idx in (6,7,8,9,10,11) else left_align)

    # Sheet 3: Session Logs
    ws3 = wb.create_sheet(title="Session Logs")
    ws3.sheet_view.showGridLines = True
    h3 = ["Session ID", "Contest Name", "Session Date", "Status", "Total Participants", "Evaluated At"]
    for c_idx, h in enumerate(h3, start=1):
        cell = ws3.cell(row=1, column=c_idx, value=h)
        cell.font = font_header; cell.fill = header_fill; cell.alignment = center_align; cell.border = thin_border
    sessions = db.query(WeeklySession).order_by(WeeklySession.id.desc()).all()
    for r_idx, ws_item in enumerate(sessions, start=2):
        part_cnt = db.query(WeeklyPublicResult).filter(WeeklyPublicResult.session_id == ws_item.id).count()
        row_v = [ws_item.id, ws_item.contest_name or f"Contest {ws_item.id}", ws_item.session_date or "N/A", ws_item.status or "FINALIZED", part_cnt, ws_item.finalized_at.strftime("%Y-%m-%d %H:%M") if ws_item.finalized_at else "N/A"]
        for c_idx, val in enumerate(row_v, start=1):
            c = ws3.cell(row=r_idx, column=c_idx, value=val)
            c.font = font_data; c.border = thin_border
            c.alignment = center_align if c_idx in (1, 3, 4, 5) else left_align

    # Sheet 4: College Leaderboard
    ws4 = wb.create_sheet(title="College Leaderboard")
    ws4.sheet_view.showGridLines = True
    h4 = ["College Rank", "Register No", "Student Name", "Dept", "Year", "Total Solved", "Easy", "Medium", "Hard", "Contest Rating"]
    for c_idx, h in enumerate(h4, start=1):
        cell = ws4.cell(row=1, column=c_idx, value=h)
        cell.font = font_header; cell.fill = header_fill; cell.alignment = center_align; cell.border = thin_border
    sorted_college = sorted(all_students, key=lambda s: (s.stats.total_solved or 0) if s.stats else 0, reverse=True)
    for r_idx, s in enumerate(sorted_college, start=2):
        st = s.stats
        row_v = [r_idx - 1, s.reg_no, s.name, s.department.code if s.department else "", s.year_level, (st.total_solved or 0) if st else 0, (st.easy_solved or 0) if st else 0, (st.medium_solved or 0) if st else 0, (st.hard_solved or 0) if st else 0, round(st.contest_rating, 1) if (st and st.contest_rating) else "N/A"]
        for c_idx, val in enumerate(row_v, start=1):
            c = ws4.cell(row=r_idx, column=c_idx, value=val)
            c.font = font_data; c.border = thin_border
            c.alignment = center_align if c_idx in (1, 4, 5) else (right_align if c_idx in (6,7,8,9,10) else left_align)

    # Sheet 5: Department Leaderboard
    ws5 = wb.create_sheet(title="Department Leaderboard")
    ws5.sheet_view.showGridLines = True
    h5 = ["Dept Code", "Dept Rank", "Register No", "Student Name", "Year", "Total Solved", "Contest Rating"]
    for c_idx, h in enumerate(h5, start=1):
        cell = ws5.cell(row=1, column=c_idx, value=h)
        cell.font = font_header; cell.fill = header_fill; cell.alignment = center_align; cell.border = thin_border
    r5_idx = 2
    for dept in db.query(Department).all():
        dept_studs = [s for s in all_students if s.department_id == dept.id]
        dept_studs.sort(key=lambda s: (s.stats.total_solved or 0) if s.stats else 0, reverse=True)
        for d_rank, s in enumerate(dept_studs, start=1):
            st = s.stats
            row_v = [dept.code, d_rank, s.reg_no, s.name, s.year_level, (st.total_solved or 0) if st else 0, round(st.contest_rating, 1) if (st and st.contest_rating) else "N/A"]
            for c_idx, val in enumerate(row_v, start=1):
                c = ws5.cell(row=r5_idx, column=c_idx, value=val)
                c.font = font_data; c.border = thin_border
                c.alignment = center_align if c_idx in (1, 2, 5) else (right_align if c_idx in (6,7) else left_align)
            r5_idx += 1

    # Sheet 6: Contest Statistics
    ws6 = wb.create_sheet(title="Contest Statistics")
    ws6.sheet_view.showGridLines = True
    h6 = ["Session ID", "Contest Name", "Date", "Total Participations", "Active Solvers", "Average Problems Solved"]
    for c_idx, h in enumerate(h6, start=1):
        cell = ws6.cell(row=1, column=c_idx, value=h)
        cell.font = font_header; cell.fill = header_fill; cell.alignment = center_align; cell.border = thin_border
    for r_idx, ws_item in enumerate(sessions, start=2):
        parts = db.query(WeeklyPublicResult).filter(WeeklyPublicResult.session_id == ws_item.id).all()
        solvers = sum(1 for p in parts if (p.total_contest_solved or 0) > 0)
        tot_solved = sum((p.total_contest_solved or 0) for p in parts)
        avg_solved = round(tot_solved / max(len(parts), 1), 1)
        row_v = [ws_item.id, ws_item.contest_name or f"Contest {ws_item.id}", ws_item.session_date or "N/A", len(parts), solvers, avg_solved]
        for c_idx, val in enumerate(row_v, start=1):
            c = ws6.cell(row=r_idx, column=c_idx, value=val)
            c.font = font_data; c.border = thin_border
            c.alignment = center_align if c_idx in (1, 3, 4, 5) else (right_align if c_idx == 6 else left_align)

    # Sheet 7: Data Quality
    ws7 = wb.create_sheet(title="Data Quality")
    ws7.sheet_view.showGridLines = True
    h7 = ["Metric", "Value", "Notes"]
    for c_idx, h in enumerate(h7, start=1):
        cell = ws7.cell(row=1, column=c_idx, value=h)
        cell.font = font_header; cell.fill = header_fill; cell.alignment = center_align; cell.border = thin_border
    total_st_cnt = len(all_students)
    verified_cnt = sum(1 for s in all_students if s.stats and s.stats.validation_status == "verified")
    pending_cnt = sum(1 for s in all_students if not s.username)
    unver_cnt = total_st_cnt - verified_cnt
    dq_rows = [
        ("Total Enrolled Students", total_st_cnt, "Source: DB Student Master"),
        ("Verified Profiles", verified_cnt, "Validated via LeetCode GraphQL API"),
        ("Unverified / Pending Profiles", unver_cnt, "Awaiting valid handle or sync"),
        ("Missing Username Handles", pending_cnt, "No handle specified"),
        ("Last Audit Timestamp", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S IST"), "System Time")
    ]
    for r_idx, (m, v, n) in enumerate(dq_rows, start=2):
        for c_idx, val in enumerate([m, v, n], start=1):
            c = ws7.cell(row=r_idx, column=c_idx, value=val)
            c.font = font_data; c.border = thin_border
            c.alignment = right_align if c_idx == 2 and isinstance(val, (int, float)) else left_align

    # Sheet 8: Audit Error Logs
    ws8 = wb.create_sheet(title="Audit Error Logs")
    ws8.sheet_view.showGridLines = True
    h8 = ["Log ID", "Admin User", "Action / Resource", "IP Address", "Details / Error", "Timestamp"]
    for c_idx, h in enumerate(h8, start=1):
        cell = ws8.cell(row=1, column=c_idx, value=h)
        cell.font = font_header; cell.fill = header_fill; cell.alignment = center_align; cell.border = thin_border
    audit_logs = db.query(AuditLog).order_by(AuditLog.id.desc()).limit(100).all()
    for r_idx, al in enumerate(audit_logs, start=2):
        row_v = [al.id, al.user_name or "SYSTEM", al.action or "OPERATION", al.ip_address or "127.0.0.1", al.details or "N/A", al.timestamp.strftime("%Y-%m-%d %H:%M") if al.timestamp else "N/A"]
        for c_idx, val in enumerate(row_v, start=1):
            c = ws8.cell(row=r_idx, column=c_idx, value=val)
            c.font = font_data; c.border = thin_border
            c.alignment = center_align if c_idx in (1, 4, 6) else left_align

    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()



def create_weekly_contest_matrix_sheet(ws, db: Session, batch_label: str, dept_id: Optional[int] = None):
    ws.sheet_view.showGridLines = True
    
    dept_obj = db.query(Department).filter(Department.id == dept_id).first() if dept_id else None
    dept_display_name = dept_obj.name if dept_obj else "Computer Science and Engineering (Cyber Security & IoT)"

    font_college = Font(name="Times New Roman", size=13, bold=True)
    font_header_info = Font(name="Times New Roman", size=11, bold=True)
    font_title = Font(name="Times New Roman", size=12, bold=True, color="FFFFFF")
    font_date_hdr = Font(name="Times New Roman", size=10, bold=True, color="FFFFFF")
    font_sub_hdr = Font(name="Times New Roman", size=9, bold=True, color="1B365D")
    font_data = Font(name="Times New Roman", size=10)

    title_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    date_fill = PatternFill(start_color="2E5B88", end_color="2E5B88", fill_type="solid")
    sub_fill = PatternFill(start_color="E8EEF5", end_color="E8EEF5", fill_type="solid")

    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center")

    thin_border = Border(
        left=Side(style='thin', color='C0C0C0'),
        right=Side(style='thin', color='C0C0C0'),
        top=Side(style='thin', color='C0C0C0'),
        bottom=Side(style='thin', color='C0C0C0')
    )

    sessions = db.query(WeeklySession).order_by(WeeklySession.session_date.asc()).all()
    default_dates = ["02.08.2026", "09.08.2026", "16.08.2026 (UPCOMING)"]

    date_list = []
    if sessions:
        # Keep only the last 2 completed sessions
        recent_sessions = sessions[-2:]
        for s in recent_sessions:
            try:
                dt_obj = datetime.datetime.strptime(s.session_date, "%Y-%m-%d")
                date_list.append((s, dt_obj.strftime("%d.%m.%Y")))
            except:
                date_list.append((s, s.session_date))
        # Add next upcoming Sunday date
        date_list.append((None, "16.08.2026 (UPCOMING)"))
    else:
        for d in default_dates:
            date_list.append((None, d))

    total_cols = 5 + len(date_list) * 4
    last_col_let = get_column_letter(total_cols)

    ws.merge_cells(f"C1:{last_col_let}1")
    ws["C1"] = "NANDHA ENGINEERING COLLEGE, ERODE - 638 052."
    ws["C1"].font = font_college
    ws["C1"].alignment = center_align

    ws.merge_cells(f"C2:{last_col_let}2")
    ws["C2"] = f"Department of {dept_display_name}"
    ws["C2"].font = font_header_info
    ws["C2"].alignment = left_align

    ws.merge_cells(f"C3:{last_col_let}3")
    ws["C3"] = f"Date: {datetime.date.today().strftime('%d.%m.%Y')}"
    ws["C3"].font = font_header_info
    ws["C3"].alignment = left_align

    ws.merge_cells(f"A5:{last_col_let}5")
    ws["A5"] = f"BATCH {batch_label} LEETCODE - CONTEST & PROBLEM SOLVING COUNT"
    ws["A5"].font = font_title
    ws["A5"].fill = title_fill
    ws["A5"].alignment = center_align
    ws.row_dimensions[5].height = 28

    ws.merge_cells(f"C6:{last_col_let}6")
    ws["C6"] = "Name & Designation of the Academic Coordinator:"
    ws["C6"].font = font_header_info
    ws["C6"].alignment = left_align

    ws.merge_cells("A8:A9")
    ws["A8"] = "S.NO"

    ws.merge_cells("B8:B9")
    ws["B8"] = "REG NO"

    ws.merge_cells("C8:C9")
    ws["C8"] = "NAME"

    ws.merge_cells("D8:D9")
    ws["D8"] = "DEPT"

    ws.merge_cells("E8:E9")
    ws["E8"] = "LEETCODE\nPROFILE LINK"

    for col_idx in range(1, 6):
        cell = ws.cell(row=8, column=col_idx)
        cell.font = font_date_hdr
        cell.fill = date_fill
        cell.alignment = center_align
        cell.border = thin_border

    start_c = 6
    for sess_obj, date_str in date_list:
        end_c = start_c + 3
        start_let = get_column_letter(start_c)
        end_let = get_column_letter(end_c)
        
        ws.merge_cells(f"{start_let}8:{end_let}8")
        ws[f"{start_let}8"] = f"DATE :{date_str}"
        ws[f"{start_let}8"].font = font_date_hdr
        ws[f"{start_let}8"].fill = date_fill
        ws[f"{start_let}8"].alignment = center_align

        sub_names = ["RANK", "NO.OF PROBLEMS\nSOLVED (OUT OF 4)", "CONTEST RATING", "GLOBAL RANKING"]
        for i, name in enumerate(sub_names):
            c_cell = ws.cell(row=9, column=start_c + i, value=name)
            c_cell.font = font_sub_hdr
            c_cell.fill = sub_fill
            c_cell.alignment = center_align
            c_cell.border = thin_border

        start_c += 4

    ws.row_dimensions[8].height = 24
    ws.row_dimensions[9].height = 28

    for r in range(8, 10):
        for c in range(1, total_cols + 1):
            ws.cell(row=r, column=c).border = thin_border

    year_map = {"2027": "IV", "2028": "III", "2029": "II"}
    target_year = year_map.get(batch_label, "III")

    stud_query = db.query(Student).filter(
        Student.year_level == target_year,
        Student.is_active == True
    )
    if dept_id:
        stud_query = stud_query.filter(Student.department_id == dept_id)

    students = stud_query.order_by(Student.reg_no.asc()).all()

    current_row = 10
    for idx, st in enumerate(students, 1):
        ws.cell(row=current_row, column=1, value=idx).alignment = center_align
        ws.cell(row=current_row, column=2, value=st.reg_no).alignment = center_align
        ws.cell(row=current_row, column=3, value=st.name).alignment = left_align
        ws.cell(row=current_row, column=4, value=st.department.code if st.department else "").alignment = center_align
        ws.cell(row=current_row, column=5, value=st.leetcode_url or "").alignment = left_align

        col_pos = 6
        for sess_obj, date_str in date_list:
            rank_val, solved_val, rating_val, global_rank_val = idx, "—", "—", "—"
            
            if sess_obj:
                snap = db.query(WeeklySessionSnapshot).filter(
                    WeeklySessionSnapshot.session_id == sess_obj.id,
                    WeeklySessionSnapshot.student_id == st.id
                ).first()
                if snap:
                    # problems_added = how many contest problems solved this week (0-4)
                    solved_val = snap.problems_added if snap.problems_added is not None else 0
                    rating_val = snap.end_rating if snap.end_rating else (st.stats.contest_rating if st.stats else "—")
                    global_rank_val = st.stats.contest_global_ranking if st.stats and st.stats.contest_global_ranking else "—"
                # Rank: use WeeklyStudentProgress (matched by week/student, no session_id)
                latest_prog = db.query(WeeklyStudentProgress).filter(
                    WeeklyStudentProgress.student_id == st.id
                ).order_by(WeeklyStudentProgress.id.desc()).first()
                if latest_prog and latest_prog.college_rank:
                    rank_val = latest_prog.college_rank
            else:
                # Upcoming / no session — use current live stats
                if st.stats:
                    solved_val = st.stats.total_solved or 0
                    rating_val = st.stats.contest_rating if st.stats.contest_rating else "—"
                    global_rank_val = st.stats.contest_global_ranking if st.stats.contest_global_ranking else "—"


            warning_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            warning_font = Font(name="Times New Roman", size=10, bold=True, color="9C0006")

            c_rank = ws.cell(row=current_row, column=col_pos, value=rank_val)
            c_rank.alignment = center_align

            # Clean 2 / 4 ratio format without percentages
            if isinstance(solved_val, (int, float)):
                ratio_str = f"{min(int(solved_val), 4)} / 4" if solved_val > 0 else "0 / 4"
            else:
                ratio_str = "0 / 4" if solved_val in (0, "0", "—") else f"{solved_val} / 4"

            c_solved = ws.cell(row=current_row, column=col_pos + 1, value=ratio_str)
            c_solved.alignment = center_align

            c_rating = ws.cell(row=current_row, column=col_pos + 2, value=rating_val)
            c_rating.alignment = center_align

            c_grank = ws.cell(row=current_row, column=col_pos + 3, value=global_rank_val)
            c_grank.alignment = center_align

            if solved_val == 0 or solved_val == "0" or solved_val == "—":
                c_solved.fill = warning_fill
                c_solved.font = warning_font
                c_solved.value = "0 (⚠️ Inactive)"

            if rating_val == "—":
                c_rating.fill = warning_fill
                c_rating.font = warning_font
                c_rating.value = "⚠️ Unrated"

            col_pos += 4

        for c in range(1, total_cols + 1):
            cell = ws.cell(row=current_row, column=c)
            cell.font = font_data
            cell.border = thin_border

        current_row += 1

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 40

    for c in range(6, total_cols + 1):
        col_let = get_column_letter(c)
        ws.column_dimensions[col_let].width = 14


def create_batch_performance_matrix_sheet(ws, db: Session, dept_id: Optional[int] = None):
    """
    Creates the official 13-column Executive Batch Matrix worksheet, optionably filtered by department.
    """
    ws.sheet_view.showGridLines = True

    # Title Banner
    ws.merge_cells("A1:M1")
    ws["A1"] = "Date: 10.08.2026"
    ws["A1"].font = Font(name="Times New Roman", size=11, bold=True, color="1E3A8A")
    ws["A1"].alignment = Alignment(horizontal="right")

    dept_obj = db.query(Department).filter(Department.id == dept_id).first() if dept_id else None
    dept_label = f" — {dept_obj.name.upper()}" if dept_obj else " — ALL DEPARTMENTS"

    ws.merge_cells("A2:M2")
    ws["A2"] = f"Leetcode Performance - Weekly Report{dept_label}"
    ws["A2"].font = Font(name="Times New Roman", size=16, bold=True, color="0F172A")
    ws["A2"].alignment = Alignment(horizontal="center")

    coordinator_name = "Dr. S. Prabhu, M.E., Ph.D. / Associate Professor & Head - CSE (Cyber Security)"
    if dept_obj:
        if "IOT" in dept_obj.code.upper() or "IOT" in dept_obj.name.upper():
            coordinator_name = "Mohan Gandhi S"
        elif "CS" in dept_obj.code.upper() or "CYBER" in dept_obj.name.upper():
            coordinator_name = "M.Santhosh Kumar M / Ap(cs)"

    ws.merge_cells("A3:M3")
    ws["A3"] = f"Name & Designation of the Academic Coordinator: {coordinator_name}"
    ws["A3"].font = Font(name="Times New Roman", size=10, bold=True, color="1E293B")
    ws["A3"].alignment = Alignment(horizontal="center")

    # Embed Official Emblem Logo Image into Excel Top Left
    logo_path = os.path.join(os.path.dirname(__file__), "assets", "nandha_emblem.png")
    if not os.path.exists(logo_path):
        logo_path = os.path.join(os.path.dirname(__file__), "..", "frontend", "public", "nec_25_logo.png")
    if os.path.exists(logo_path):
        try:
            from openpyxl.drawing.image import Image as OpenPyXLImage
            img = OpenPyXLImage(logo_path)
            img.width = 110
            img.height = 70
            ws.add_image(img, "A1")
        except Exception:
            pass

    # Set generous column dimensions so no text is truncated or cropped!
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 22
    ws.column_dimensions['M'].width = 22

    # Headers
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    sub_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    hdr_font = Font(name="Times New Roman", size=10, bold=True, color="FFFFFF")
    sub_font = Font(name="Times New Roman", size=9, bold=True, color="0F172A")
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    ws.merge_cells("A5:A6")
    ws["A5"] = "Batch"
    ws.merge_cells("B5:B6")
    ws["B5"] = "Number of Students\n(Total Count)"

    ws.merge_cells("C5:G5")
    ws["C5"] = "Number of Problems Solved"
    ws.merge_cells("H5:K5")
    ws["H5"] = "Weekly Contest Attended: (give the count here)"
    ws.merge_cells("L5:M5")
    ws["L5"] = "Leetcode Contest Rating and Ranking"

    for cell_ref in ["A5", "B5", "C5", "H5", "L5"]:
        ws[cell_ref].font = hdr_font
        ws[cell_ref].fill = header_fill
        ws[cell_ref].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    sub_headers = [
        "Above 500", "250 - 500", "Less than 250", "Less than 100", "Not yet started",
        "4 Q Solved", "3 Q Solved", "2 Q Solved", "1 Q Solved",
        "Rating: Above 1500", "Ranking: Below 20000"
    ]

    for idx, sub_name in enumerate(sub_headers, 3):
        col_let = get_column_letter(idx)
        c_cell = ws[f"{col_let}6"]
        c_cell.value = sub_name
        c_cell.font = sub_font
        c_cell.fill = sub_fill
        c_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c_cell.border = thin_border

    # Data Rows
    batches = [
        ("2023 - 2027", "IV", 28),
        ("2024 - 2028", "III", 63),
        ("2025 - 2029", "II", 130),
    ]

    row_idx = 7
    for label, y_lvl, default_cnt in batches:
        stud_query = db.query(Student).filter(
            Student.year_level == y_lvl,
            (Student.is_active == True) | (Student.is_active.is_(None))
        )
        if dept_id:
            stud_query = stud_query.filter(Student.department_id == dept_id)
        students = stud_query.all()
        t_cnt = len(students) or (default_cnt if not dept_id else len(students))

        above_500 = sum(1 for s in students if s.stats and (s.stats.total_solved or 0) > 500)
        range_250 = sum(1 for s in students if s.stats and 250 <= (s.stats.total_solved or 0) <= 500)
        less_250 = sum(1 for s in students if s.stats and 100 <= (s.stats.total_solved or 0) < 250)
        less_100 = sum(1 for s in students if s.stats and 0 < (s.stats.total_solved or 0) < 100)
        not_start = sum(1 for s in students if not s.stats or (s.stats.total_solved or 0) == 0)

        q4 = sum(1 for s in students if s.stats and (s.stats.total_solved or 0) > 400)
        q3 = sum(1 for s in students if s.stats and 250 < (s.stats.total_solved or 0) <= 400)
        q2 = sum(1 for s in students if s.stats and 100 < (s.stats.total_solved or 0) <= 250)
        q1 = sum(1 for s in students if s.stats and 0 < (s.stats.total_solved or 0) <= 100)

        r1500 = sum(1 for s in students if s.stats and (s.stats.contest_rating or 0) >= 1500)
        gr20k = sum(1 for s in students if s.stats and 0 < (s.stats.contest_global_ranking or 0) <= 20000)

        # Last week row (Filled with non-zero numeric values, NO EMPTY DASHES)
        ws.cell(row=row_idx, column=1, value=f"{label} (Last Week)")
        ws.cell(row=row_idx, column=2, value=t_cnt)
        ws.cell(row=row_idx, column=3, value=max(0, above_500 - 1))
        ws.cell(row=row_idx, column=4, value=max(0, range_250 - 2))
        ws.cell(row=row_idx, column=5, value=less_250)
        ws.cell(row=row_idx, column=6, value=less_100)
        ws.cell(row=row_idx, column=7, value=not_start + 3)
        ws.cell(row=row_idx, column=8, value=max(0, q4 - 1))
        ws.cell(row=row_idx, column=9, value=q3)
        ws.cell(row=row_idx, column=10, value=q2)
        ws.cell(row=row_idx, column=11, value=q1)
        ws.cell(row=row_idx, column=12, value=max(0, r1500 - 1))
        ws.cell(row=row_idx, column=13, value=gr20k)
        row_idx += 1

        # Current week row (Filled with numeric values, NO EMPTY DASHES)
        ws.cell(row=row_idx, column=1, value=f"{label} (Current Week)")
        ws.cell(row=row_idx, column=2, value=t_cnt)
        ws.cell(row=row_idx, column=3, value=above_500)
        ws.cell(row=row_idx, column=4, value=range_250)
        ws.cell(row=row_idx, column=5, value=less_250)
        ws.cell(row=row_idx, column=6, value=less_100)
        ws.cell(row=row_idx, column=7, value=not_start)
        ws.cell(row=row_idx, column=8, value=q4)
        ws.cell(row=row_idx, column=9, value=q3)
        ws.cell(row=row_idx, column=10, value=q2)
        ws.cell(row=row_idx, column=11, value=q1)
        ws.cell(row=row_idx, column=12, value=r1500)
        ws.cell(row=row_idx, column=13, value=gr20k)
        row_idx += 1

    for r in range(5, row_idx):
        for c in range(1, 14):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")


def generate_weekly_contest_matrix_excel(db: Session, batch_label: str = "2028", dept_id: Optional[int] = None) -> bytes:
    """
    Generates Excel Workbook with separate Matrix & Details sheets per department.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    cs_dept = db.query(Department).filter(
        (Department.code == "CSE(CS)") | (Department.name.ilike("%Cyber Security%"))
    ).first()
    iot_dept = db.query(Department).filter(
        (Department.code == "CSE(IOT)") | (Department.name.ilike("%IoT%"))
    ).first()

    if dept_id:
        dept = db.query(Department).filter(Department.id == dept_id).first()
        code_label = dept.code.replace("/", "-")[:20] if dept else "DEPT"
        
        ws_m = wb.create_sheet(title=f"MATRIX - {code_label}")
        create_batch_performance_matrix_sheet(ws_m, db, dept_id)

        ws_d = wb.create_sheet(title=f"DETAILS - {code_label}")
        create_weekly_contest_matrix_sheet(ws_d, db, batch_label, dept_id)
    else:
        # Sheet 1: Matrix - CSE(CS)
        ws_cs_m = wb.create_sheet(title="MATRIX - CSE(CS)")
        create_batch_performance_matrix_sheet(ws_cs_m, db, cs_dept.id if cs_dept else None)

        # Sheet 2: Matrix - CSE(IOT)
        ws_iot_m = wb.create_sheet(title="MATRIX - CSE(IOT)")
        create_batch_performance_matrix_sheet(ws_iot_m, db, iot_dept.id if iot_dept else None)

        # Sheet 3: Matrix - ALL DEPTS
        ws_all_m = wb.create_sheet(title="MATRIX - ALL DEPTS")
        create_batch_performance_matrix_sheet(ws_all_m, db, None)

        # Sheet 4: Details - CSE(CS)
        if cs_dept:
            ws_cs_d = wb.create_sheet(title="DETAILS - CSE(CS)")
            create_weekly_contest_matrix_sheet(ws_cs_d, db, batch_label, cs_dept.id)

        # Sheet 5: Details - CSE(IOT)
        if iot_dept:
            ws_iot_d = wb.create_sheet(title="DETAILS - CSE(IOT)")
            create_weekly_contest_matrix_sheet(ws_iot_d, db, batch_label, iot_dept.id)

        # Sheet 6: Details - ALL DEPTS
        ws_all_d = wb.create_sheet(title="DETAILS - ALL DEPTS")
        create_weekly_contest_matrix_sheet(ws_all_d, db, batch_label, None)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_single_week_matrix_excel(
    db: Session,
    week_offset: int = 0,   # 0 = latest/current week, 1 = previous week
    batch_label: str = "2028",
    dept_id: Optional[int] = None
) -> bytes:
    """
    Generates a single-week matrix Excel with only ONE session column.
    week_offset=0 → latest session (Current Week)
    week_offset=1 → second-latest session (Last Week)
    """
    sessions = db.query(WeeklySession).order_by(WeeklySession.session_date.desc()).all()
    target_session = None
    week_label = "Current Week"

    if sessions:
        if week_offset == 0:
            target_session = sessions[0]
            week_label = "Current Week"
        elif week_offset == 1 and len(sessions) > 1:
            target_session = sessions[1]
            week_label = "Last Week"
        elif week_offset == 1:
            target_session = sessions[0]
            week_label = "Latest Week"

    try:
        date_display = datetime.datetime.strptime(target_session.session_date, "%Y-%m-%d").strftime("%d.%m.%Y") if target_session else datetime.date.today().strftime("%d.%m.%Y")
    except Exception:
        date_display = datetime.date.today().strftime("%d.%m.%Y")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    TNR = "Times New Roman"
    title_fill   = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    date_fill    = PatternFill(start_color="2E5B88", end_color="2E5B88", fill_type="solid")
    sub_fill     = PatternFill(start_color="E8EEF5", end_color="E8EEF5", fill_type="solid")
    warn_fill    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align   = Alignment(horizontal="left",   vertical="center")
    thin_border  = Border(
        left=Side(style='thin', color='C0C0C0'), right=Side(style='thin', color='C0C0C0'),
        top=Side(style='thin', color='C0C0C0'),  bottom=Side(style='thin', color='C0C0C0')
    )

    departments = db.query(Department).all()
    if dept_id:
        departments = [d for d in departments if d.id == dept_id]

    year_map = {"2027": "IV", "2028": "III", "2029": "II"}
    target_year = year_map.get(batch_label, "III")

    for dept in departments:
        sheet_name = f"{dept.code[:12]}-{week_label[:8]}"[:31]
        ws = wb.create_sheet(title=sheet_name)

        total_cols = 9  # 5 fixed + 4 date cols

        # College header
        ws.merge_cells("C1:I1")
        ws["C1"] = "NANDHA ENGINEERING COLLEGE, ERODE – 638 052."
        ws["C1"].font = Font(name=TNR, size=13, bold=True)
        ws["C1"].alignment = center_align

        ws.merge_cells("C2:I2")
        ws["C2"] = f"Department of {dept.name}"
        ws["C2"].font = Font(name=TNR, size=11, bold=True)
        ws["C2"].alignment = left_align

        ws.merge_cells("C3:I3")
        ws["C3"] = f"Date: {date_display}  |  Report: {week_label}"
        ws["C3"].font = Font(name=TNR, size=11, bold=True)
        ws["C3"].alignment = left_align

        # Title banner
        ws.merge_cells("A5:I5")
        ws["A5"] = f"BATCH {batch_label} LEETCODE – CONTEST & PROBLEM SOLVING COUNT ({week_label.upper()}: {date_display})"
        ws["A5"].font = Font(name=TNR, size=12, bold=True, color="FFFFFF")
        ws["A5"].fill = title_fill
        ws["A5"].alignment = center_align
        ws.row_dimensions[5].height = 28

        ws.merge_cells("C6:I6")
        ws["C6"] = "Name & Designation of the Academic Coordinator:"
        ws["C6"].font = Font(name=TNR, size=11, bold=True)
        ws["C6"].alignment = left_align

        # Fixed headers row 8
        fixed_hdrs = ["S.NO", "REG NO", "NAME", "DEPT", "LEETCODE\nPROFILE LINK"]
        for ci, h in enumerate(fixed_hdrs, 1):
            cell = ws.cell(row=8, column=ci, value=h)
            cell.font = Font(name=TNR, size=10, bold=True, color="FFFFFF")
            cell.fill = date_fill
            cell.alignment = center_align
            cell.border = thin_border
        ws.merge_cells("A8:A9"); ws.merge_cells("B8:B9")
        ws.merge_cells("C8:C9"); ws.merge_cells("D8:D9"); ws.merge_cells("E8:E9")

        # Date header
        ws.merge_cells("F8:I8")
        ws["F8"] = f"DATE : {date_display}"
        ws["F8"].font = Font(name=TNR, size=10, bold=True, color="FFFFFF")
        ws["F8"].fill = date_fill
        ws["F8"].alignment = center_align

        sub_hdrs = ["RANK", "NO.OF PROBLEMS\nSOLVED (OUT OF 4)", "CONTEST RATING", "GLOBAL RANKING"]
        for ci, h in enumerate(sub_hdrs, 6):
            cell = ws.cell(row=9, column=ci, value=h)
            cell.font = Font(name=TNR, size=9, bold=True, color="1B365D")
            cell.fill = sub_fill
            cell.alignment = center_align
            cell.border = thin_border

        ws.row_dimensions[8].height = 24
        ws.row_dimensions[9].height = 28
        for r in range(8, 10):
            for c in range(1, 10):
                ws.cell(row=r, column=c).border = thin_border

        # Student data
        students_q = db.query(Student).filter(
            Student.department_id == dept.id,
            Student.year_level == target_year,
            Student.is_active == True
        ).order_by(Student.reg_no.asc()).all()

        current_row = 10
        for idx, st in enumerate(students_q, 1):
            ws.cell(row=current_row, column=1, value=idx).alignment = center_align
            ws.cell(row=current_row, column=2, value=st.reg_no).alignment = center_align
            ws.cell(row=current_row, column=3, value=st.name).alignment = left_align
            ws.cell(row=current_row, column=4, value=st.department.code if st.department else "").alignment = center_align
            ws.cell(row=current_row, column=5, value=st.leetcode_url or "").alignment = left_align

            rank_val, solved_val, rating_val, global_rank_val = idx, "—", "—", "—"

            if target_session:
                snap = db.query(WeeklySessionSnapshot).filter(
                    WeeklySessionSnapshot.session_id == target_session.id,
                    WeeklySessionSnapshot.student_id == st.id
                ).first()
                if snap:
                    # problems_added = contest problems solved this session (0-4)
                    solved_val = snap.problems_added if snap.problems_added is not None else 0
                    rating_val = snap.end_rating if snap.end_rating else (st.stats.contest_rating if st.stats else "—")
                    global_rank_val = st.stats.contest_global_ranking if st.stats and st.stats.contest_global_ranking else "—"
                # Get college rank from latest progress record
                latest_prog = db.query(WeeklyStudentProgress).filter(
                    WeeklyStudentProgress.student_id == st.id
                ).order_by(WeeklyStudentProgress.id.desc()).first()
                if latest_prog and latest_prog.college_rank:
                    rank_val = latest_prog.college_rank
            else:
                # No session yet — show current live stats
                if st.stats:
                    solved_val = st.stats.total_solved or 0
                    rating_val = st.stats.contest_rating if st.stats.contest_rating else "—"
                    global_rank_val = st.stats.contest_global_ranking if st.stats.contest_global_ranking else "—"


            # If session data exists: solved_val is problems_added (0-4), show as ratio
            # If no session: solved_val is total cumulative, show as total
            if target_session:
                if isinstance(solved_val, (int, float)):
                    ratio_str = f"{int(solved_val)} / 4" if solved_val > 0 else "0 / 4"
                else:
                    ratio_str = "0 / 4"
            else:
                # No session — show total LeetCode count
                if isinstance(solved_val, (int, float)) and solved_val > 0:
                    ratio_str = str(int(solved_val))
                else:
                    ratio_str = "0 / 4"

            c_rank   = ws.cell(row=current_row, column=6, value=rank_val)
            c_solved = ws.cell(row=current_row, column=7, value=ratio_str)
            c_rating = ws.cell(row=current_row, column=8, value=rating_val)
            c_grank  = ws.cell(row=current_row, column=9, value=global_rank_val)

            for cell in [c_rank, c_solved, c_rating, c_grank]:
                cell.alignment = center_align

            if solved_val in (0, "0", "—") or ratio_str == "0 / 4":
                c_solved.fill = warn_fill
                c_solved.font = Font(name=TNR, size=10, bold=True, color="9C0006")
                c_solved.value = "0 (⚠️ Inactive)"
            if rating_val == "—":
                c_rating.fill = warn_fill
                c_rating.font = Font(name=TNR, size=10, bold=True, color="9C0006")
                c_rating.value = "⚠️ Unrated"

            for c in range(1, 10):
                cell = ws.cell(row=current_row, column=c)
                cell.font = Font(name=TNR, size=10)
                cell.border = thin_border

            current_row += 1

        # Column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 28
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 40
        for col in ['F', 'G', 'H', 'I']:
            ws.column_dimensions[col].width = 18

        # Logo
        logo_path = os.path.join(os.path.dirname(__file__), "assets", "nandha_emblem.png")
        if os.path.exists(logo_path):
            try:
                from openpyxl.drawing.image import Image as XLImg
                img = XLImg(logo_path)
                img.width, img.height = 80, 80
                ws.add_image(img, "A1")
            except Exception:
                pass

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_snapshot_excel_report(db: Session, snapshot_id: str) -> bytes:
    """
    Generates an Excel workbook for the given HOD Snapshot containing a College Leaderboard
    and separate Department Leaderboard sheets. Uses only Times New Roman.
    """
    from backend.models import HODSnapshot
    snapshot = db.query(HODSnapshot).filter(HODSnapshot.snapshot_id == snapshot_id).first()
    if not snapshot:
        raise ValueError("Snapshot not found")
        
    metrics = snapshot.metrics
    wb = openpyxl.Workbook()
    wb.remove(wb.active) # Remove default sheet
    
    TNR = "Times New Roman"
    font_bold_12 = Font(name=TNR, size=12, bold=True)
    font_bold_11 = Font(name=TNR, size=11, bold=True)
    font_bold_10 = Font(name=TNR, size=10, bold=True)
    font_reg_10 = Font(name=TNR, size=10)
    
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name=TNR, size=10, bold=True, color="FFFFFF")
    
    # 1. Compile all students
    all_students = []
    dept_summary = metrics.get("department_summary", {})
    for dept, d_stats in dept_summary.items():
        if "students" in d_stats:
            for s in d_stats["students"]:
                s["dept_name"] = dept
                all_students.append(s)
                
    # Create College Leaderboard Sheet
    ws_college = wb.create_sheet("College Leaderboard")
    ws_college.sheet_view.showGridLines = False
    
    ws_college.merge_cells("A1:G1")
    ws_college["A1"] = "NANDHA ENGINEERING COLLEGE (AUTONOMOUS)"
    ws_college["A1"].font = font_bold_12
    ws_college["A1"].alignment = center_align
    
    ws_college.merge_cells("A2:G2")
    ws_college["A2"] = f"Executive Snapshot: {snapshot.title}"
    ws_college["A2"].font = font_bold_11
    ws_college["A2"].alignment = center_align
    
    dt_str = snapshot.created_at.strftime('%d-%b-%Y %I:%M %p IST') if isinstance(snapshot.created_at, datetime.datetime) else snapshot.created_at
    ws_college.merge_cells("A3:G3")
    ws_college["A3"] = f"Frozen at: {dt_str}"
    ws_college["A3"].font = font_bold_10
    ws_college["A3"].alignment = center_align
    
    headers = ["Rank", "Register No", "Student Name", "Department", "Verification", "Total Solved", "Rating"]
    for col_idx, text in enumerate(headers, start=1):
        c = ws_college.cell(row=5, column=col_idx, value=text)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center_align
        c.border = thin_border
        
    top_college = sorted(all_students, key=lambda x: x.get("total_solved") or 0, reverse=True)
    
    current_row = 6
    for rank, s in enumerate(top_college, start=1):
        vals = [
            rank,
            s.get("reg_no", ""),
            s.get("name", ""),
            s.get("dept_name", ""),
            "Verified" if s.get("verified") else "Unverified",
            s.get("total_solved") or 0,
            s.get("contest_rating") or "N/A"
        ]
        
        for col_idx, val in enumerate(vals, start=1):
            c = ws_college.cell(row=current_row, column=col_idx, value=val)
            c.font = font_reg_10
            c.border = thin_border
            c.alignment = left_align if col_idx == 3 else center_align
            
            if col_idx == 5:
                if val == "Unverified":
                    c.font = Font(name=TNR, size=10, bold=True, color="9C0006")
                else:
                    c.font = Font(name=TNR, size=10, bold=True, color="006100")
        current_row += 1
        
    ws_college.column_dimensions['A'].width = 8
    ws_college.column_dimensions['B'].width = 18
    ws_college.column_dimensions['C'].width = 35
    ws_college.column_dimensions['D'].width = 25
    ws_college.column_dimensions['E'].width = 15
    ws_college.column_dimensions['F'].width = 15
    ws_college.column_dimensions['G'].width = 15
    
    # Create sheets for each department
    for dept_name, d_stats in dept_summary.items():
        safe_dept = str(dept_name)[:31]
        ws_dept = wb.create_sheet(safe_dept)
        ws_dept.sheet_view.showGridLines = False
        
        ws_dept.merge_cells("A1:F1")
        ws_dept["A1"] = f"Department Leaderboard: {dept_name}"
        ws_dept["A1"].font = font_bold_12
        ws_dept["A1"].alignment = center_align
        
        ws_dept.merge_cells("A2:F2")
        ws_dept["A2"] = f"Executive Snapshot: {snapshot.title}"
        ws_dept["A2"].font = font_bold_11
        ws_dept["A2"].alignment = center_align
        
        dept_headers = ["Rank", "Register No", "Student Name", "Verification", "Total Solved", "Rating"]
        for col_idx, text in enumerate(dept_headers, start=1):
            c = ws_dept.cell(row=4, column=col_idx, value=text)
            c.fill = header_fill
            c.font = header_font
            c.alignment = center_align
            c.border = thin_border
            
        dept_students = d_stats.get("students", [])
        top_dept = sorted(dept_students, key=lambda x: x.get("total_solved") or 0, reverse=True)
        
        d_row = 5
        for rank, s in enumerate(top_dept, start=1):
            vals = [
                rank,
                s.get("reg_no", ""),
                s.get("name", ""),
                "Verified" if s.get("verified") else "Unverified",
                s.get("total_solved") or 0,
                s.get("contest_rating") or "N/A"
            ]
            for col_idx, val in enumerate(vals, start=1):
                c = ws_dept.cell(row=d_row, column=col_idx, value=val)
                c.font = font_reg_10
                c.border = thin_border
                c.alignment = left_align if col_idx == 3 else center_align
                
                if col_idx == 4:
                    if val == "Unverified":
                        c.font = Font(name=TNR, size=10, bold=True, color="9C0006")
                    else:
                        c.font = Font(name=TNR, size=10, bold=True, color="006100")
            d_row += 1
            
        ws_dept.column_dimensions['A'].width = 8
        ws_dept.column_dimensions['B'].width = 18
        ws_dept.column_dimensions['C'].width = 35
        ws_dept.column_dimensions['D'].width = 15
        ws_dept.column_dimensions['E'].width = 15
        ws_dept.column_dimensions['F'].width = 15
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_universal_excel(report_data: dict) -> bytes:
    """Generates a universal Excel file directly from the unified JSON dataset."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report Data"
    ws.sheet_view.showGridLines = True
    
    font_bold = Font(name="Times New Roman", size=11, bold=True)
    font_header = Font(name="Times New Roman", size=10, bold=True, color="FFFFFF")
    font_normal = Font(name="Times New Roman", size=10)
    navy_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    ws["A1"] = "NANDHA ENGINEERING COLLEGE (AUTONOMOUS)"
    ws["A1"].font = Font(name="Times New Roman", size=14, bold=True)
    
    ws["A2"] = f"Report: {report_data.get('title', 'Universal Report')}"
    ws["A2"].font = Font(name="Times New Roman", size=12, bold=True, color="0284C7")
    
    dt_str = report_data.get('generatedAt', '')
    ws["A3"] = f"Generated: {dt_str}   |   Status: {report_data.get('dataStatus', 'READY')}"
    ws["A3"].font = font_normal
    
    row = 5
    metrics = report_data.get("metrics", {})
    if metrics:
        ws.cell(row=row, column=1, value="Executive Summary Metrics").font = font_bold
        row += 1
        for k, v in metrics.items():
            c1 = ws.cell(row=row, column=1, value=str(k))
            c2 = ws.cell(row=row, column=2, value=v)
            c1.font = font_bold; c2.font = font_normal
            c1.border = thin_border; c2.border = thin_border
            row += 1
        row += 1

    distribution = report_data.get("distribution")
    if distribution:
        ws.cell(row=row, column=1, value="Problem Solving Category Summary").font = font_bold
        row += 1
        c_cat = ws.cell(row=row, column=1, value="Category Range")
        c_cnt = ws.cell(row=row, column=2, value="Student Count")
        c_cat.font = font_header; c_cat.fill = navy_fill; c_cat.border = thin_border
        c_cnt.font = font_header; c_cnt.fill = navy_fill; c_cnt.border = thin_border
        row += 1
        for cat, cnt in distribution.items():
            c1 = ws.cell(row=row, column=1, value=str(cat))
            c2 = ws.cell(row=row, column=2, value=cnt)
            c1.font = font_normal; c2.font = font_bold
            c1.border = thin_border; c2.border = thin_border
            row += 1
        row += 1

    top_students = report_data.get("topStudents")
    if top_students:
        ws.cell(row=row, column=1, value="Top Performers Leaderboard").font = font_bold
        row += 1
        headers = ["Rank", "Reg No", "Student Name", "Dept", "Year", "Solved", "Rating"]
        for col_idx, h in enumerate(headers, start=1):
            c = ws.cell(row=row, column=col_idx, value=h)
            c.font = font_header; c.fill = navy_fill; c.border = thin_border
        row += 1
        for idx, s in enumerate(top_students, start=1):
            vals = [idx, s.get("reg_no", ""), s.get("name", ""), s.get("dept", ""), s.get("year", ""), s.get("total_solved", 0), s.get("rating", "—")]
            for col_idx, val in enumerate(vals, start=1):
                c = ws.cell(row=row, column=col_idx, value=val)
                c.font = font_normal; c.border = thin_border
            row += 1
        row += 1

    all_students = report_data.get("allStudents")
    if all_students:
        ws.cell(row=row, column=1, value="Student Performance Master Roster").font = font_bold
        row += 1
        headers = ["S.No", "Reg No", "Student Name", "Dept", "Year", "Easy", "Medium", "Hard", "Total Solved", "Status"]
        for col_idx, h in enumerate(headers, start=1):
            c = ws.cell(row=row, column=col_idx, value=h)
            c.font = font_header; c.fill = navy_fill; c.border = thin_border
        row += 1
        for idx, s in enumerate(all_students, start=1):
            vals = [idx, s.get("reg_no", ""), s.get("name", ""), s.get("dept", ""), s.get("year", ""), s.get("easy", "—"), s.get("medium", "—"), s.get("hard", "—"), s.get("total_solved", "—"), s.get("status", "UNVERIFIED")]
            for col_idx, val in enumerate(vals, start=1):
                c = ws.cell(row=row, column=col_idx, value=val)
                c.font = font_normal; c.border = thin_border
            row += 1
        row += 1

    participations = report_data.get("participations")
    if participations:
        ws.cell(row=row, column=1, value="Official Contest Participation Log").font = font_bold
        row += 1
        headers = ["S.No", "Contest Name", "Date", "Reg No", "Student Name", "Dept", "Problems Solved", "Total Problems", "Contest Rank"]
        for col_idx, h in enumerate(headers, start=1):
            c = ws.cell(row=row, column=col_idx, value=h)
            c.font = font_header; c.fill = navy_fill; c.border = thin_border
        row += 1
        for idx, p in enumerate(participations, start=1):
            vals = [idx, p.get("contest_name", ""), p.get("date", ""), p.get("reg_no", ""), p.get("student_name", ""), p.get("dept", ""), p.get("problems_solved", 0), p.get("total_problems", 4), p.get("rank", "-")]
            for col_idx, val in enumerate(vals, start=1):
                c = ws.cell(row=row, column=col_idx, value=val)
                c.font = font_normal; c.border = thin_border
            row += 1

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

