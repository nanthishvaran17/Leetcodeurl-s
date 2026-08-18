"""
Master Weekly Report Engine Unit Tests
Verifies canonical dataset generation, full roster classification, category equations,
and Excel/PDF/DOCX consumption of identical canonical metadata.
"""
import os
import unittest
import datetime
import openpyxl
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from backend.database import Base
from backend.models import Student, Department, LeetCodeProfileStats, WeeklySession, WeeklyPublicResult, WeeklyVirtualResult
from backend.services.report_data_service import get_problem_category
from backend.services.weekly_report_service import generate_weekly_performance_data
from backend.exporters.weekly_excel_generator import build_weekly_performance_excel
from backend.pdf_generator import build_weekly_performance_pdf
from backend.word_generator import build_weekly_performance_docx


class TestWeeklyReportEngine(unittest.TestCase):

    def setUp(self):
        self.engine = create_engine("sqlite:///:memory:")
        Base.metadata.create_all(bind=self.engine)
        Session = sessionmaker(bind=self.engine)
        self.db = Session()

        # Seed test departments
        self.dept_cs = Department(id=1, name="Computer Science and Engineering (Cyber Security)", code="CSE(CS)")
        self.dept_iot = Department(id=2, name="Computer Science and Engineering (IoT)", code="CSE(IoT)")
        self.db.add_all([self.dept_cs, self.dept_iot])
        self.db.commit()

        # Seed test sessions with required week_number
        self.s513 = WeeklySession(id=1, week_number=1, contest_name="Weekly Contest 513", session_date="02.08.2026", status="FINALIZED")
        self.s514 = WeeklySession(id=2, week_number=2, contest_name="Weekly Contest 514", session_date="09.08.2026", status="FINALIZED")
        self.db.add_all([self.s513, self.s514])
        self.db.commit()

        # Seed test students across all 4 batches (I, II, III, IV)
        self.students_data = [
            ("732224CC001", "Student Above 500", "CSE(CS)", "III", "https://leetcode.com/u/user1/", "user1", 600, 300, 200, 100),
            ("732224CC002", "Student 500 Bound", "CSE(CS)", "III", "https://leetcode.com/u/user2/", "user2", 500, 250, 150, 100),
            ("732224CC003", "Student 250 Bound", "CSE(IoT)", "II",  "https://leetcode.com/u/user3/", "user3", 250, 150, 80, 20),
            ("732224CC004", "Student 101 Bound", "CSE(IoT)", "II",  "https://leetcode.com/u/user4/", "user4", 101, 60, 40, 1),
            ("732224CC005", "Student 100 Bound", "CSE(CS)", "IV",   "https://leetcode.com/u/user5/", "user5", 100, 50, 40, 10),
            ("732224CC006", "Student 1 Bound",   "CSE(CS)", "IV",   "https://leetcode.com/u/user6/", "user6", 1, 1, 0, 0),
            ("732224CC007", "Student 0 Solved",  "CSE(IoT)", "I",   "https://leetcode.com/u/user7/", "user7", 0, 0, 0, 0),
            ("732224CC008", "Student Missing",   "CSE(IoT)", "I",   None, None, None, None, None, None),
        ]

        for reg, name, d_code, yr, url, uname, tot, ez, med, hd in self.students_data:
            dept_id = 1 if d_code == "CSE(CS)" else 2
            st = Student(reg_no=reg, name=name, department_id=dept_id, year_level=yr, leetcode_url=url, username=uname, is_active=True)
            self.db.add(st)
            self.db.commit()
            self.db.refresh(st)

            if tot is not None:
                stats = LeetCodeProfileStats(
                    student_id=st.id,
                    total_solved=tot,
                    easy_solved=ez,
                    medium_solved=med,
                    hard_solved=hd,
                    status="verified",
                    sync_status="success"
                )
            else:
                stats = LeetCodeProfileStats(
                    student_id=st.id,
                    total_solved=None,
                    status="pending",
                    sync_status="failed",
                    error_code="MISSING_LINK"
                )
            self.db.add(stats)
            self.db.commit()

            # Seed contest results
            solved_c = 4 if reg == "732224CC001" else (3 if reg == "732224CC002" else 0)
            part_st = "PUBLIC_ATTENDED" if solved_c > 0 else "PUBLIC_NOT_ATTENDED"
            pub_res = WeeklyPublicResult(
                session_id=self.s514.id,
                student_id=st.id,
                reg_no=reg,
                name=name,
                dept=d_code,
                year=yr,
                participation_status=part_st,
                total_contest_solved=solved_c,
                data_fetch_status="SUCCESS" if solved_c > 0 else "DATA_UNAVAILABLE"
            )
            self.db.add(pub_res)
        self.db.commit()

    def tearDown(self):
        self.db.close()

    def test_category_boundary_assignment(self):
        """Verifies exact category boundaries: 0, 1, 100, 101, 250, 500, 501."""
        self.assertEqual(get_problem_category(501, True), "Above 500")
        self.assertEqual(get_problem_category(500, True), "250-500")
        self.assertEqual(get_problem_category(250, True), "250-500")
        self.assertEqual(get_problem_category(249, True), "101-250")
        self.assertEqual(get_problem_category(101, True), "101-250")
        self.assertEqual(get_problem_category(100, True), "Less than 100")
        self.assertEqual(get_problem_category(1, True), "Less than 100")
        self.assertEqual(get_problem_category(0, True), "Not Yet Started")
        self.assertEqual(get_problem_category(None, False), "Data Unavailable")

    def test_full_roster_represented(self):
        """All 8 seeded students must be represented in canonical dataset."""
        data = generate_weekly_performance_data(self.db, report_date="17-08-2026")
        self.assertEqual(data["total_students"], 8)
        self.assertEqual(len(data["all_students_current"]), 8)

        # Batch 2026-2030 (Year I) is present
        batch_labels = [b["batch"] for b in data["batch_summaries"]]
        self.assertIn("2026 - 2030", batch_labels)
        self.assertIn("2025 - 2029", batch_labels)
        self.assertIn("2024 - 2028", batch_labels)
        self.assertIn("2023 - 2027", batch_labels)

    def test_exporters_consume_identical_canonical_dataset(self):
        """Excel, PDF, and DOCX all consume identical dataset without raising errors."""
        data = generate_weekly_performance_data(self.db, report_date="17-08-2026")

        # Excel
        test_xlsx = "test_weekly_report.xlsx"
        try:
            build_weekly_performance_excel(data, test_xlsx)
            self.assertTrue(os.path.exists(test_xlsx))
            wb = openpyxl.load_workbook(test_xlsx)
            self.assertIn("01_CSE_CS_Matrix", wb.sheetnames)
            self.assertIn("02_CSE_IOT_Matrix", wb.sheetnames)
            self.assertIn("03_II_Year", wb.sheetnames)
            self.assertIn("04_III_Year", wb.sheetnames)
            self.assertIn("05_IV_Year", wb.sheetnames)
            self.assertIn("06_Overall_Summary", wb.sheetnames)
            self.assertIn("07_Not_Attended", wb.sheetnames)
        finally:
            if os.path.exists(test_xlsx):
                os.remove(test_xlsx)

        # PDF
        pdf_bytes = build_weekly_performance_pdf(data)
        self.assertIsInstance(pdf_bytes, bytes)
        self.assertGreater(len(pdf_bytes), 1000)

        # DOCX
        docx_bytes = build_weekly_performance_docx(data)
        self.assertIsInstance(docx_bytes, bytes)
        self.assertGreater(len(docx_bytes), 1000)


if __name__ == "__main__":
    unittest.main()
