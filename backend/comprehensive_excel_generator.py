import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime

def generate_comprehensive_excel(db) -> bytes:
    """
    Generates a master institutional Excel workbook (.xlsx) containing:
    1. Executive Summary (Nandha Engineering College 13-column matrix)
    2. Cyber Security (Full Student Details)
    3. IoT (Full Student Details)
    4. II Year (Batch 2029)
    5. III Year (Batch 2028)
    6. IV Year (Batch 2027)
    """
    from backend.word_generator import _compute_dept_matrix, BATCH_CONFIG

    wb = openpyxl.Workbook()
    ws_summary = wb.active
    ws_summary.title = "Executive Summary"

    # Styling definitions
    navy_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    brand_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    sub_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
    header_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    font_title = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    font_sub = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    font_header_bold = Font(name="Calibri", size=10, bold=True, color="0F172A")
    font_bold = Font(name="Calibri", size=10, bold=True)
    font_regular = Font(name="Calibri", size=10)

    thin_border_side = Side(style='thin', color='CBD5E1')
    grid_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')

    # 1. Summary Sheet Banner
    ws_summary.merge_cells('A1:M1')
    ws_summary['A1'] = "NANDHA ENGINEERING COLLEGE, ERODE - 638 052 (AUTONOMOUS)"
    ws_summary['A1'].font = font_title
    ws_summary['A1'].fill = navy_fill
    ws_summary['A1'].alignment = align_center

    ws_summary.merge_cells('A2:M2')
    ws_summary['A2'] = f"LEETCODE PERFORMANCE WEEKLY REPORT — Date: {datetime.datetime.now().strftime('%d.%m.%Y')}"
    ws_summary['A2'].font = font_sub
    ws_summary['A2'].fill = brand_fill
    ws_summary['A2'].alignment = align_center

    ws_summary.row_dimensions[1].height = 30
    ws_summary.row_dimensions[2].height = 22

    cur_row = 4

    for did, dept_title in [(1, "Department of Computer Science and Engineering (Cyber Security)"), (2, "Department of Computer Science and Engineering (IoT)")]:
        ws_summary.merge_cells(f'A{cur_row}:M{cur_row}')
        ws_summary[f'A{cur_row}'] = dept_title
        ws_summary[f'A{cur_row}'].font = font_sub
        ws_summary[f'A{cur_row}'].fill = sub_fill
        ws_summary[f'A{cur_row}'].alignment = align_center
        ws_summary.row_dimensions[cur_row].height = 22
        cur_row += 1

        # Multi-level headers
        ws_summary.merge_cells(f'A{cur_row}:A{cur_row+1}')
        ws_summary[f'A{cur_row}'] = "Batch"
        ws_summary.merge_cells(f'B{cur_row}:B{cur_row+1}')
        ws_summary[f'B{cur_row}'] = "Number of Students\n(Total Count)"

        ws_summary.merge_cells(f'C{cur_row}:G{cur_row}')
        ws_summary[f'C{cur_row}'] = "Number of Problems Solved"

        ws_summary.merge_cells(f'H{cur_row}:K{cur_row}')
        ws_summary[f'H{cur_row}'] = "Weekly Contest Attended"

        ws_summary.merge_cells(f'L{cur_row}:M{cur_row}')
        ws_summary[f'L{cur_row}'] = "Leetcode Contest Rating and Ranking"

        sub_headers = ["Above 500", "250 - 500", "100 - 249", "1 - 99", "0", "4Q", "3Q", "2Q", "1Q", "Rating > 1500", "Ranking < 20000"]
        for idx, sh in enumerate(sub_headers, start=3):
            col_letter = get_column_letter(idx)
            ws_summary[f'{col_letter}{cur_row+1}'] = sh

        for r in range(cur_row, cur_row+2):
            ws_summary.row_dimensions[r].height = 22
            for c in range(1, 14):
                cell = ws_summary.cell(row=r, column=c)
                cell.font = font_header_bold
                cell.fill = header_fill
                cell.alignment = align_center
                cell.border = grid_border

        cur_row += 2

        data_m = _compute_dept_matrix(db, did)
        for b_key in ["2023_2027", "2024_2028", "2025_2029"]:
            b_info = BATCH_CONFIG[b_key]
            b_label = b_info["label"]
            lw = data_m[b_key]["last_week"]
            cw = data_m[b_key]["current_week"]

            # Last week row
            ws_summary.append([
                f"{b_label} (Last Week)", lw["total_students"],
                lw["prob_above_500"], lw["prob_250_500"], lw["prob_100_249"], lw["prob_1_99"], lw["prob_0"],
                lw["q4"], lw["q3"], lw["q2"], lw["q1"],
                lw["rating_above_1500"], lw["rank_below_20k"]
            ])
            ws_summary.row_dimensions[cur_row].height = 20
            for col in range(1, 14):
                c_cell = ws_summary.cell(row=cur_row, column=col)
                c_cell.font = font_regular
                c_cell.border = grid_border
                c_cell.alignment = align_left if col == 1 else align_center
                if cur_row % 2 == 0:
                    c_cell.fill = zebra_fill
            cur_row += 1

            # Current week row
            ws_summary.append([
                f"{b_label} (Current Week)", cw["total_students"],
                cw["prob_above_500"], cw["prob_250_500"], cw["prob_100_249"], cw["prob_1_99"], cw["prob_0"],
                cw["q4"], cw["q3"], cw["q2"], cw["q1"],
                cw["rating_above_1500"], cw["rank_below_20k"]
            ])
            ws_summary.row_dimensions[cur_row].height = 20
            for col in range(1, 14):
                c_cell = ws_summary.cell(row=cur_row, column=col)
                c_cell.font = font_bold if col in (1, 2) else font_regular
                c_cell.border = grid_border
                c_cell.alignment = align_left if col == 1 else align_center
                if cur_row % 2 == 0:
                    c_cell.fill = zebra_fill
            cur_row += 1

        cur_row += 2

    # Auto fit summary columns
    for col in ws_summary.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row in (1, 2):
                continue
            val = str(cell.value or '')
            max_len = max(max_len, len(val))
        ws_summary.column_dimensions[col_letter].width = max(max_len + 4, 14)

    # 2. Add Student Roster Sheets
    from backend.models import Student, Department
    students = db.query(Student).filter(Student.is_active == 1).order_by(Student.department_id, Student.year_level, Student.reg_no).all()

    def add_roster_sheet(title, filtered_students):
        ws = wb.create_sheet(title=title)
        ws.merge_cells('A1:J1')
        ws['A1'] = f"NANDHA ENGINEERING COLLEGE — {title.upper()}"
        ws['A1'].font = font_title
        ws['A1'].fill = navy_fill
        ws['A1'].alignment = align_center

        ws.merge_cells('A2:J2')
        ws['A2'] = f"Total Students: {len(filtered_students)} | Generated: {datetime.datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws['A2'].font = font_sub
        ws['A2'].fill = brand_fill
        ws['A2'].alignment = align_center

        headers = ["S.No", "Register No", "Student Name", "Department", "Year", "LeetCode Username", "Total Solved", "Contest Rating", "Global Rank", "Status"]
        ws.row_dimensions[4].height = 24
        for idx, h in enumerate(headers, start=1):
            c = ws.cell(row=4, column=idx, value=h)
            c.font = font_header_bold
            c.fill = header_fill
            c.alignment = align_center
            c.border = grid_border

        for r_idx, s in enumerate(filtered_students, start=1):
            row_num = r_idx + 4
            d_name = "Cyber Security" if s.department_id == 1 else "IoT"
            t_solved = s.stats.total_solved if s.stats and s.stats.total_solved is not None else 0
            c_rating = s.stats.contest_rating if s.stats and s.stats.contest_rating is not None else "—"
            c_rank = f"#{s.stats.contest_global_ranking:,}" if s.stats and s.stats.contest_global_ranking else "—"
            ws.append([
                r_idx,
                s.reg_no,
                s.name,
                d_name,
                f"Year {s.year_level}",
                s.username or "—",
                t_solved,
                c_rating,
                c_rank,
                "ACTIVE"
            ])
            ws.row_dimensions[row_num].height = 20
            for col in range(1, 11):
                cell = ws.cell(row=row_num, column=col)
                cell.font = font_regular
                cell.border = grid_border
                cell.alignment = align_center if col in (1, 2, 4, 5, 8, 9, 10) else (align_left if col in (3, 6) else align_right)
                if row_num % 2 == 0:
                    cell.fill = zebra_fill

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row in (1, 2):
                    continue
                val = str(cell.value or '')
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 13)

    add_roster_sheet("Cyber Security", [s for s in students if s.department_id == 1])
    add_roster_sheet("IoT", [s for s in students if s.department_id == 2])
    add_roster_sheet("II Year (2029)", [s for s in students if s.year_level == 2])
    add_roster_sheet("III Year (2028)", [s for s in students if s.year_level == 3])
    add_roster_sheet("IV Year (2027)", [s for s in students if s.year_level == 4])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
