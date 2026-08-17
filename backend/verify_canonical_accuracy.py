import sys
import openpyxl
from backend.database import SessionLocal
from backend.services.canonical_contest_engine import build_canonical_contest_dataset
from backend.exporters.excel_exporter import export_excel_from_dataset
from backend.routes.weekly_contests import get_session_matrix
from backend.routes.reports import _get_dataset_for_id

def run_full_accuracy_audit():
    print("================================================================================")
    print("NANDHA ENGINEERING COLLEGE — LEETCODE DATA ACCURACY & RECONCILIATION AUDIT")
    print("================================================================================")

    db = SessionLocal()
    session_id = 5

    # 1. Test Canonical Dataset Generation
    dataset = build_canonical_contest_dataset(session_id, db)
    total_master = dataset["metrics"]["totalStudents"]
    all_rows = dataset["all_rows"]

    print(f"\n1. CANONICAL DATASET INTEGRITY:")
    print(f"   - Contest: {dataset['contestName']} ({dataset['sessionDate']})")
    print(f"   - Total Master Roster Count: {total_master} Students")
    print(f"   - Public Attended: {dataset['statusCounts']['PUBLIC']}")
    print(f"   - Virtual Attended: {dataset['statusCounts']['VIRTUAL']}")
    print(f"   - Confirmed Not Attended: {dataset['statusCounts']['NOT_ATTENDED']}")
    print(f"   - Explicit Quality Issues / Unlinked: {dataset['statusCounts']['USERNAME_NOT_FOUND']}")
    print(f"   - Participation Rate: {dataset['metrics']['participationPercentage']}%")
    assert total_master == 302, f"Expected 302 students, got {total_master}"
    print("   [PASS] Master Student Count verified (302 students).")

    # 2. Mathematical Reconciliation Verification
    reconciliation = dataset["reconciliation"]
    print(f"\n2. MATHEMATICAL RECONCILIATION ENGINE:")
    print(f"   - Master Count: {reconciliation['masterCount']}")
    print(f"   - Sum of Department Totals: {reconciliation['deptSum']}")
    print(f"   - Sum of Year Totals: {reconciliation['yearSum']}")
    print(f"   - Sum of Status Category Totals: {reconciliation['statusSum']}")
    print(f"   - Gatekeeper Status: {'PASSED' if reconciliation['passed'] else 'FAILED'}")
    assert reconciliation["passed"] is True, "Reconciliation failed!"
    print("   [PASS] 100% Department, Year, and Status Reconciliation verified.")

    # 3. Question-Wise Sum and Authenticity Rule
    print(f"\n3. QUESTION-WISE FORMULA & NULL INTEGRITY CHECK:")
    participants_checked = 0
    non_participants_checked = 0
    for r in all_rows:
        status = r["status"]
        if status in ("PUBLIC", "VIRTUAL"):
            participants_checked += 1
            q1 = r["q1"] or 0
            q2 = r["q2"] or 0
            q3 = r["q3"] or 0
            q4 = r["q4"] or 0
            solved = r["total_solved"]
            assert solved == (q1 + q2 + q3 + q4), f"Mismatch for {r['reg_no']}: solved={solved} != sum({q1},{q2},{q3},{q4})"
            assert q1 in (0, 1) and q2 in (0, 1) and q3 in (0, 1) and q4 in (0, 1)
        else:
            non_participants_checked += 1
            # Must not have fake question counts
            assert r["q1"] is None or r["q1"] == "—"
            assert r["q2"] is None or r["q2"] == "—"
            assert r["q3"] is None or r["q3"] == "—"
            assert r["q4"] is None or r["q4"] == "—"

    print(f"   - Verified Participants Checked: {participants_checked} (Formula `Solved === Q1+Q2+Q3+Q4` satisfied for 100%)")
    print(f"   - Non-Participants / Unresolved Checked: {non_participants_checked} (All preserved as NULL/—, 0 fabricated questions)")
    print("   [PASS] Absolute Question Authenticity Rule verified.")

    # 4. Multi-Sheet Excel Export Integrity
    print(f"\n4. MULTI-SHEET EXCEL EXPORT VERIFICATION:")
    excel_bytes = export_excel_from_dataset(dataset)
    import io
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
    sheet_names = wb.sheetnames
    print(f"   - Generated Workbook Sheets ({len(sheet_names)}): {', '.join(sheet_names)}")
    
    assert "Weekly Contest Summary" in sheet_names
    assert "Student Performance" in sheet_names
    assert "Public Attended" in sheet_names
    assert "Public Not Attended" in sheet_names
    assert "CSE(CS)-II" in sheet_names or "CSE(CS)-II" in "".join(sheet_names)
    
    ws_perf = wb["Student Performance"]
    # Excel header is at row 3, data begins row 4
    excel_perf_rows = ws_perf.max_row - 3
    print(f"   - Excel 'Student Performance' Sheet Row Count: {excel_perf_rows}")
    assert excel_perf_rows == total_master, f"Excel rows ({excel_perf_rows}) does not match Master Count ({total_master})"
    print("   [PASS] UI <-> Excel 1:1 Row Count Parity verified.")

    # 5. UI API Endpoints Parity
    print(f"\n5. UI / PREVIEW / MATRIX API ENDPOINTS PARITY:")
    ui_matrix = get_session_matrix(session_id=session_id, db=db)
    report_data, fn = _get_dataset_for_id(str(session_id), db)
    
    assert len(ui_matrix["rows"]) == total_master
    assert len(report_data["rows"]) == total_master
    assert ui_matrix["metrics"]["officialAttended"] == report_data["metrics"]["officialAttended"]
    assert ui_matrix["metrics"]["participationRate"] == report_data["metrics"]["participationRate"]
    print("   - UI Table Rows: " + str(len(ui_matrix["rows"])))
    print("   - Report Preview Rows: " + str(len(report_data["rows"])))
    print("   - UI Official Attended: " + str(ui_matrix["metrics"]["officialAttended"]))
    print("   - Report Official Attended: " + str(report_data["metrics"]["officialAttended"]))
    print("   [PASS] Single Source of Truth verified across all layers.")

    print("\n================================================================================")
    print(">>> AUDIT RESULT: 100% RECONCILIATION & DATA ACCURACY PASSED SUCCESSFULLY! <<<")
    print("================================================================================\n")

if __name__ == "__main__":
    run_full_accuracy_audit()
