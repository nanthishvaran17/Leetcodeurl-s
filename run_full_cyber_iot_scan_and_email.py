# run_full_cyber_iot_scan_and_email.py
"""
Production Autonomous Scanner & Dispatcher for Cyber Security and IoT Departments
- Full live scan of 299 students (Cyber Security + IoT) via LeetCode GraphQL & Contest API
- Database update in data/leetcode_tracker.db
- Professional 8-sheet Excel report generation with custom formatting
- Executive PDF report generation via ReportLab
- Rich HTML Email generation
- Automated Email dispatch to authorized recipients with attachments
"""

import os
import sys
import time
import json
import sqlite3
import datetime
from zoneinfo import ZoneInfo
from concurrent.futures import ThreadPoolExecutor, as_completed
import pandas as pd
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

if hasattr(sys.stdout, 'reconfigure'):
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

# Add parent directory to path
sys.path.insert(0, os.path.abspath(os.path.dirname(__file__)))

from backend.database import SessionLocal
from backend.services.email_service import send_email
from backend.config import settings

IST = ZoneInfo("Asia/Kolkata")
DB_PATH = os.path.abspath("data/leetcode_tracker.db")
STUDENTS_EXCEL_PATH = os.path.abspath("students.xlsx")
GRAPHQL_URL = "https://leetcode.com/graphql"
CONTEST_API_BASE = "https://leetcode.com/contest/api/ranking"

BATCH_MAPPING = {
    1: "2025-2029 (I Year)",
    2: "2024-2028 (II Year)",
    3: "2023-2027 (III Year)",
    4: "2022-2026 (IV Year)"
}

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Content-Type": "application/json",
    "Referer": "https://leetcode.com"
}

GRAPHQL_QUERY = """
query getUserFullDetails($username: String!) {
  matchedUser(username: $username) {
    username
    submitStats {
      acSubmissionNum {
        difficulty
        count
      }
      totalSubmissionNum {
        difficulty
        count
      }
    }
    contestRating
  }
  userContestRanking(username: $username) {
    rating
    globalRanking
    attendedContestsCount
    topPercentage
  }
  userContestRankingHistory(username: $username) {
    attended
    problemsSolved
    totalProblems
    finishTimeInSeconds
    rating
    ranking
    contest {
      title
    }
  }
}
"""

def parse_year_num(val):
    if not val or str(val).strip() == "nan":
        return 3
    s = str(val).strip().upper()
    if s in ('I', '1'): return 1
    if s in ('II', '2'): return 2
    if s in ('III', '3'): return 3
    if s in ('IV', '4'): return 4
    digits = ''.join(filter(str.isdigit, s))
    if digits:
        return int(digits)
    return 3

def load_student_roster():
    print(f"Reading student records from {STUDENTS_EXCEL_PATH}...")
    if not os.path.exists(STUDENTS_EXCEL_PATH):
        raise FileNotFoundError(f"Missing {STUDENTS_EXCEL_PATH}")
    
    df = pd.read_excel(STUDENTS_EXCEL_PATH)
    students = []
    
    for idx, row in df.iterrows():
        name = str(row.get("Name", "") or "").strip()
        reg_no = str(row.get("RollNumber", "") or "").strip()
        dept_raw = str(row.get("Department", "") or "").strip().upper()
        year_raw = row.get("Year", 3)
        year_num = parse_year_num(year_raw)
        username = str(row.get("LeetCodeUsername", "") or "").strip()
        
        # Standardize department - check IOT first!
        if "IOT" in dept_raw or "INTERNET" in dept_raw or "CI" in reg_no.upper():
            dept_code = "CSE(IOT)"
            dept_name = "Computer Science and Engineering (IoT)"
            dept_short = "IoT"
        elif "CS" in dept_raw or "CYBER" in dept_raw or "CC" in reg_no.upper():
            dept_code = "CSE(CS)"
            dept_name = "Computer Science and Engineering (Cyber Security)"
            dept_short = "Cyber Security"
        else:
            dept_code = dept_raw
            dept_name = dept_raw
            dept_short = dept_raw
            
        students.append({
            "id": idx + 1,
            "name": name,
            "reg_no": reg_no,
            "department": dept_code,
            "dept_name": dept_name,
            "dept_short": dept_short,
            "year": year_num,
            "year_label": f"Year {year_num}",
            "batch": BATCH_MAPPING.get(year_num, f"Year {year_num}"),
            "username": username if username != "nan" else "",
            "is_valid_user": bool(username and username != "nan" and len(username) > 1)
        })
        
    print(f"  -> Total Loaded: {len(students)} students")
    cs_count = sum(1 for s in students if s["department"] == "CSE(CS)")
    iot_count = sum(1 for s in students if s["department"] == "CSE(IOT)")
    print(f"  -> CSE(CS) [Cyber Security]: {cs_count} students")
    print(f"  -> CSE(IOT) [IoT]:           {iot_count} students")
    return students

from backend.leetcode_fetcher import fetch_leetcode_profile_sync

def fetch_single_student_data(student, contest_targets=["Weekly Contest 516", "Weekly Contest 515"]):
    username = student["username"]
    if not student["is_valid_user"]:
        return {
            **student,
            "status": "MISSING_USERNAME",
            "total_solved": 0,
            "easy_solved": 0,
            "medium_solved": 0,
            "hard_solved": 0,
            "rating": 0.0,
            "global_rank": 0,
            "attended_contests": 0,
            "w516_attended": False,
            "w516_solved": 0,
            "w516_status": "NOT_ATTENDED",
            "w515_attended": False,
            "w515_solved": 0,
            "w515_status": "NOT_ATTENDED",
            "error": "No username provided"
        }
        
    try:
        prof = fetch_leetcode_profile_sync(username, timeout=10, max_retries=2)
        if prof.get("status") == "success" and prof.get("total_solved") is not None:
            total_solved = int(prof.get("total_solved") or 0)
            easy_s = int(prof.get("easy_solved") or 0)
            med_s = int(prof.get("medium_solved") or 0)
            hard_s = int(prof.get("hard_solved") or 0)
            rating = round(float(prof.get("contest_rating") or 0.0), 1)
            global_rank = int(prof.get("contest_global_ranking") or prof.get("contest_global_rank") or 0)
            
            # Check contest participations
            parts = prof.get("contest_participations") or []
            w516_att, w516_solved, w516_stat = False, 0, "NOT ATTENDED"
            w515_att, w515_solved, w515_stat = False, 0, "NOT ATTENDED"
            
            for p in parts:
                c_name = str(p.get("contest_name", ""))
                if "516" in c_name:
                    w516_att = bool(p.get("started") or p.get("registered"))
                    w516_solved = int(p.get("problems_solved") or 0)
                    w516_stat = p.get("participation_type", "OFFICIAL")
                elif "515" in c_name:
                    w515_att = bool(p.get("started") or p.get("registered"))
                    w515_solved = int(p.get("problems_solved") or 0)
                    w515_stat = p.get("participation_type", "OFFICIAL")
                    
            return {
                **student,
                "status": "VERIFIED" if total_solved > 0 else "ZERO_SOLVED",
                "total_solved": total_solved,
                "easy_solved": easy_s,
                "medium_solved": med_s,
                "hard_solved": hard_s,
                "rating": rating,
                "global_rank": global_rank,
                "attended_contests": len(parts),
                "w516_attended": w516_att,
                "w516_solved": w516_solved,
                "w516_status": w516_stat,
                "w515_attended": w515_att,
                "w515_solved": w515_solved,
                "w515_status": w515_stat,
                "error": None
            }
        else:
            return {
                **student,
                "status": "USER_NOT_FOUND" if "not found" in str(prof.get("error_message","")).lower() else "ZERO_SOLVED",
                "total_solved": 0,
                "easy_solved": 0,
                "medium_solved": 0,
                "hard_solved": 0,
                "rating": 0.0,
                "global_rank": 0,
                "attended_contests": 0,
                "w516_attended": False,
                "w516_solved": 0,
                "w516_status": "NOT_ATTENDED",
                "w515_attended": False,
                "w515_solved": 0,
                "w515_status": "NOT_ATTENDED",
                "error": prof.get("error_message") or "Failed to fetch profile"
            }
    except Exception as e:
        return {
            **student,
            "status": "ERROR",
            "total_solved": 0,
            "easy_solved": 0,
            "medium_solved": 0,
            "hard_solved": 0,
            "rating": 0.0,
            "global_rank": 0,
            "attended_contests": 0,
            "w516_attended": False,
            "w516_solved": 0,
            "w516_status": "NOT_ATTENDED",
            "w515_attended": False,
            "w515_solved": 0,
            "w515_status": "NOT_ATTENDED",
            "error": str(e)
        }

CACHE_FILE = "cyber_iot_scan_cache.json"

def scan_all_students(students, force_rescan=False):
    if not force_rescan and os.path.exists(CACHE_FILE):
        try:
            mtime = os.path.getmtime(CACHE_FILE)
            if time.time() - mtime < 1800: # 30 mins
                with open(CACHE_FILE, "r", encoding="utf-8") as f:
                    cached_data = json.load(f)
                if len(cached_data) == len(students):
                    print(f"Loaded {len(cached_data)} fresh scan results from {CACHE_FILE} (scanned recently).")
                    # Re-map departments based on updated roster
                    roster_dept_map = {s["username"]: s for s in students}
                    for item in cached_data:
                        u = item.get("username", "")
                        if u in roster_dept_map:
                            item["department"] = roster_dept_map[u]["department"]
                            item["dept_name"] = roster_dept_map[u]["dept_name"]
                            item["dept_short"] = roster_dept_map[u]["dept_short"]
                            item["batch"] = roster_dept_map[u]["batch"]
                    return cached_data
        except Exception as e:
            print(f"Cache load notice: {e}")

    print(f"\nStarting concurrent scanning of {len(students)} students (15 worker threads)...")
    start_time = time.time()
    results = []
    
    with ThreadPoolExecutor(max_workers=15) as executor:
        future_to_student = {executor.submit(fetch_single_student_data, s): s for s in students}
        completed = 0
        total = len(students)
        for future in as_completed(future_to_student):
            res = future.result()
            results.append(res)
            completed += 1
            if completed % 30 == 0 or completed == total:
                print(f"  -> Scanned [{completed}/{total}] ({completed*100//total}%) students...")
                
    elapsed = round(time.time() - start_time, 2)
    print(f"Completed full scan in {elapsed}s.")
    
    # Sort results by Total Solved descending
    results.sort(key=lambda x: (x["total_solved"], x["rating"]), reverse=True)
    
    try:
        with open(CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(results, f, indent=2)
        print(f"  -> Cached scan results to {CACHE_FILE}")
    except Exception:
        pass
        
    return results

def compute_department_metrics(results, dept_code):
    dept_students = [r for r in results if r["department"] == dept_code]
    total = len(dept_students)
    
    above_500 = sum(1 for s in dept_students if s["total_solved"] > 500)
    between_250_500 = sum(1 for s in dept_students if 250 <= s["total_solved"] <= 500)
    between_100_249 = sum(1 for s in dept_students if 100 <= s["total_solved"] < 250)
    between_1_99 = sum(1 for s in dept_students if 1 <= s["total_solved"] < 100)
    not_started = sum(1 for s in dept_students if s["total_solved"] == 0)
    
    total_problems = sum(s["total_solved"] for s in dept_students)
    avg_problems = round(total_problems / total, 1) if total else 0
    
    rating_above_1500 = sum(1 for s in dept_students if s["rating"] > 1500)
    w516_participated = sum(1 for s in dept_students if s["w516_attended"] or s["w516_solved"] > 0)
    w516_q4 = sum(1 for s in dept_students if s["w516_solved"] >= 4)
    w516_q3 = sum(1 for s in dept_students if s["w516_solved"] == 3)
    w516_q2 = sum(1 for s in dept_students if s["w516_solved"] == 2)
    w516_q1 = sum(1 for s in dept_students if s["w516_solved"] == 1)
    
    # Group by Batch
    batches = {}
    for s in dept_students:
        b = s["batch"]
        if b not in batches:
            batches[b] = {
                "batch": b, "total": 0, "above_500": 0, "250_500": 0, "100_249": 0, "1_99": 0,
                "not_started": 0, "total_solved": 0, "w516_active": 0, "w516_q4": 0, "w516_q3": 0,
                "w516_q2": 0, "w516_q1": 0, "rating_1500": 0, "students": []
            }
        b_data = batches[b]
        b_data["students"].append(s)
        b_data["total"] += 1
        b_data["total_solved"] += s["total_solved"]
        
        if s["total_solved"] > 500: b_data["above_500"] += 1
        elif 250 <= s["total_solved"] <= 500: b_data["250_500"] += 1
        elif 100 <= s["total_solved"] < 250: b_data["100_249"] += 1
        elif 1 <= s["total_solved"] < 100: b_data["1_99"] += 1
        else: b_data["not_started"] += 1
        
        if s["w516_attended"] or s["w516_solved"] > 0: b_data["w516_active"] += 1
        if s["w516_solved"] >= 4: b_data["w516_q4"] += 1
        elif s["w516_solved"] == 3: b_data["w516_q3"] += 1
        elif s["w516_solved"] == 2: b_data["w516_q2"] += 1
        elif s["w516_solved"] == 1: b_data["w516_q1"] += 1
        
        if s["rating"] > 1500: b_data["rating_1500"] += 1
        
    return {
        "dept_code": dept_code,
        "dept_name": "Computer Science and Engineering (Cyber Security)" if dept_code == "CSE(CS)" else "Computer Science and Engineering (IoT)",
        "dept_short": "Cyber Security" if dept_code == "CSE(CS)" else "IoT",
        "total_students": total,
        "active_students": total - not_started,
        "total_problems_solved": total_problems,
        "avg_problems": avg_problems,
        "above_500": above_500,
        "between_250_500": between_250_500,
        "between_100_249": between_100_249,
        "between_1_99": between_1_99,
        "not_started": not_started,
        "rating_above_1500": rating_above_1500,
        "w516_participated": w516_participated,
        "w516_q4": w516_q4,
        "w516_q3": w516_q3,
        "w516_q2": w516_q2,
        "w516_q1": w516_q1,
        "batches": batches,
        "students": dept_students
    }

def update_sqlite_database(results):
    print(f"\nPersisting live scan results to {DB_PATH}...")
    if not os.path.exists(DB_PATH):
        print("  Database file does not exist. Skipping DB update.")
        return
        
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    
    updated_count = 0
    now_iso = datetime.datetime.now(IST).isoformat()
    
    for s in results:
        uname = s["username"]
        reg_no = s["reg_no"]
        if not uname:
            continue
            
        cur.execute("SELECT id FROM students WHERE username = ? OR reg_no = ?", (uname, reg_no))
        row = cur.fetchone()
        
        if row:
            student_id = row[0]
            try:
                cur.execute("SELECT id FROM leetcode_profile_stats WHERE student_id = ?", (student_id,))
                stat_row = cur.fetchone()
                if stat_row:
                    cur.execute("""
                        UPDATE leetcode_profile_stats 
                        SET total_solved = ?, easy_solved = ?, medium_solved = ?, hard_solved = ?,
                            contest_rating = ?, contest_global_ranking = ?, status = 'VERIFIED',
                            last_updated = ?, last_successful_sync = ?
                        WHERE student_id = ?
                    """, (s["total_solved"], s["easy_solved"], s["medium_solved"], s["hard_solved"], s["rating"], s["global_rank"], now_iso, now_iso, student_id))
                else:
                    cur.execute("""
                        INSERT INTO leetcode_profile_stats 
                        (student_id, total_solved, easy_solved, medium_solved, hard_solved, contest_rating, contest_global_ranking, status, last_updated, last_successful_sync)
                        VALUES (?, ?, ?, ?, ?, ?, ?, 'VERIFIED', ?, ?)
                    """, (student_id, s["total_solved"], s["easy_solved"], s["medium_solved"], s["hard_solved"], s["rating"], s["global_rank"], now_iso, now_iso))
                updated_count += 1
            except Exception as ex:
                pass
                
    conn.commit()
    conn.close()
    print(f"  -> Successfully updated/synced {updated_count} student profile statistics in SQLite database.")

def create_styled_excel(cs_metrics, iot_metrics, all_results, filename):
    print(f"\n📊 [EXCEL] Generating Professional 8-Sheet Workbook: {filename}...")
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    
    # Styles
    font_title = Font(name="Segoe UI", size=15, bold=True, color="FFFFFF")
    font_section = Font(name="Segoe UI", size=12, bold=True, color="0F172A")
    font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_data = Font(name="Segoe UI", size=10, color="1E293B")
    font_bold = Font(name="Segoe UI", size=10, bold=True, color="0F172A")
    font_green = Font(name="Segoe UI", size=10, bold=True, color="15803D")
    font_blue = Font(name="Segoe UI", size=10, bold=True, color="0369A1")
    font_red = Font(name="Segoe UI", size=10, bold=True, color="B91C1C")
    
    fill_dark_navy = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    fill_cyber_purple = PatternFill(start_color="4338CA", end_color="4338CA", fill_type="solid")
    fill_iot_teal = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
    fill_subhead = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    fill_highlight = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    fill_warning = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    
    border_thin = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')

    # ==========================================
    # SHEET 1: EXECUTIVE SUMMARY
    # ==========================================
    ws_exec = wb.create_sheet(title="Executive Summary")
    ws_exec.views.sheetView[0].showGridLines = True
    
    # Title Banner
    ws_exec.merge_cells("A1:G2")
    ws_exec["A1"] = "NANDHA ENGINEERING COLLEGE — LEETCODE PERFORMANCE MATRIX"
    ws_exec["A1"].font = font_title
    ws_exec["A1"].fill = fill_dark_navy
    ws_exec["A1"].alignment = align_center
    
    ws_exec["A3"] = f"Comprehensive Cyber Security & IoT Scan Report | Generated on: {datetime.datetime.now(IST).strftime('%d %B %Y, %I:%M %p IST')}"
    ws_exec["A3"].font = Font(name="Segoe UI", size=10, italic=True, color="64748B")
    
    # Comparative Matrix Table
    ws_exec["A5"] = "DEPARTMENTAL COMPARISON OVERVIEW"
    ws_exec["A5"].font = font_section
    
    exec_headers = ["Metric / Parameter", "Cyber Security (CSE-CS)", "IoT (CSE-IOT)", "Combined Institutional Total"]
    for col_idx, h in enumerate(exec_headers, start=1):
        cell = ws_exec.cell(row=6, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_subhead
        cell.alignment = align_center
        cell.border = border_thin
        
    cs_tot = cs_metrics.get("total_students", 0)
    iot_tot = iot_metrics.get("total_students", 0)
    
    metrics_rows = [
        ("Total Enrolled Students", cs_tot, iot_tot, cs_tot + iot_tot),
        ("Active Coding Students (> 0 Solved)", cs_metrics.get("active_students", 0), iot_metrics.get("active_students", 0), cs_metrics.get("active_students", 0) + iot_metrics.get("active_students", 0)),
        ("Total LeetCode Problems Solved", cs_metrics.get("total_problems_solved", 0), iot_metrics.get("total_problems_solved", 0), cs_metrics.get("total_problems_solved", 0) + iot_metrics.get("total_problems_solved", 0)),
        ("Average Problems Solved per Student", cs_metrics.get("avg_problems", 0), iot_metrics.get("avg_problems", 0), round((cs_metrics.get("total_problems_solved", 0) + iot_metrics.get("total_problems_solved", 0)) / (cs_tot + iot_tot), 1) if (cs_tot + iot_tot) else 0),
        ("Elite Coders (> 500 Solved)", cs_metrics.get("above_500", 0), iot_metrics.get("above_500", 0), cs_metrics.get("above_500", 0) + iot_metrics.get("above_500", 0)),
        ("Advanced Coders (250 - 500 Solved)", cs_metrics.get("between_250_500", 0), iot_metrics.get("between_250_500", 0), cs_metrics.get("between_250_500", 0) + iot_metrics.get("between_250_500", 0)),
        ("Intermediate Coders (100 - 249 Solved)", cs_metrics.get("between_100_249", 0), iot_metrics.get("between_100_249", 0), cs_metrics.get("between_100_249", 0) + iot_metrics.get("between_100_249", 0)),
        ("Beginners (1 - 99 Solved)", cs_metrics.get("between_1_99", 0), iot_metrics.get("between_1_99", 0), cs_metrics.get("between_1_99", 0) + iot_metrics.get("between_1_99", 0)),
        ("Not Yet Started (0 Solved)", cs_metrics.get("not_started", 0), iot_metrics.get("not_started", 0), cs_metrics.get("not_started", 0) + iot_metrics.get("not_started", 0)),
        ("Contest Rating > 1500", cs_metrics.get("rating_above_1500", 0), iot_metrics.get("rating_above_1500", 0), cs_metrics.get("rating_above_1500", 0) + iot_metrics.get("rating_above_1500", 0)),
        ("Weekly Contest 516 Participations", cs_metrics.get("w516_participated", 0), iot_metrics.get("w516_participated", 0), cs_metrics.get("w516_participated", 0) + iot_metrics.get("w516_participated", 0)),
    ]
    
    for r_idx, row in enumerate(metrics_rows, start=7):
        for c_idx, val in enumerate(row, start=1):
            cell = ws_exec.cell(row=r_idx, column=c_idx, value=val)
            cell.font = font_bold if c_idx in (1, 4) else font_data
            cell.alignment = align_left if c_idx == 1 else align_center
            cell.border = border_thin
            if r_idx % 2 == 0:
                cell.fill = fill_zebra
                
    # ==========================================
    # SHEET 2 & 3: BATCH-WISE SUMMARIES
    # ==========================================
    for sheet_title, dept_m, dept_fill in [
        ("Cyber Security Summary", cs_metrics, fill_cyber_purple),
        ("IoT Summary", iot_metrics, fill_iot_teal)
    ]:
        ws = wb.create_sheet(title=sheet_title)
        ws.views.sheetView[0].showGridLines = True
        
        ws.merge_cells("A1:M2")
        ws["A1"] = f"ACADEMIC BATCH DISTRIBUTION — {dept_m.get('dept_name', '').upper()}"
        ws["A1"].font = font_title
        ws["A1"].fill = dept_fill
        ws["A1"].alignment = align_center
        
        batch_headers = [
            "Academic Batch", "Total Students", "Above 500", "250 - 500", "100 - 249", "1 - 99", "Not Started",
            "Contest Solved 4Q", "Contest Solved 3Q", "Contest Solved 2Q", "Contest Solved 1Q", "Rating > 1500", "Active Rate"
        ]
        
        for col_idx, h in enumerate(batch_headers, start=1):
            cell = ws.cell(row=4, column=col_idx, value=h)
            cell.font = font_header
            cell.fill = fill_subhead
            cell.alignment = align_center
            cell.border = border_thin
            
        cur_row = 5
        for batch_name, b_stats in sorted(dept_m.get("batches", {}).items()):
            act_pct = f"{round(((b_stats['total'] - b_stats['not_started']) / b_stats['total'] * 100), 1)}%" if b_stats['total'] else "0%"
            row_data = [
                batch_name, b_stats["total"], b_stats["above_500"], b_stats["250_500"], b_stats["100_249"],
                b_stats["1_99"], b_stats["not_started"], b_stats["w516_q4"], b_stats["w516_q3"],
                b_stats["w516_q2"], b_stats["w516_q1"], b_stats["rating_1500"], act_pct
            ]
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=cur_row, column=col_idx, value=val)
                cell.font = font_bold if col_idx in (1, 2, 13) else font_data
                cell.alignment = align_left if col_idx == 1 else align_center
                cell.border = border_thin
                if cur_row % 2 == 0:
                    cell.fill = fill_zebra
            cur_row += 1

    # ==========================================
    # SHEET 4 & 5: DETAILED STUDENT ROSTERS
    # ==========================================
    roster_headers = [
        "S.No", "Register Number", "Student Name", "Department", "Year / Batch", "LeetCode Username",
        "Total Solved", "Easy", "Medium", "Hard", "Contest Rating", "Global Rank",
        "Contest 516 Status", "Contest 516 Solved", "Contest 515 Solved", "Profile Status"
    ]
    
    for sheet_title, dept_students, dept_fill in [
        ("CSE (Cyber Security)", cs_metrics.get("students", []), fill_cyber_purple),
        ("CSE (IoT)", iot_metrics.get("students", []), fill_iot_teal),
        ("All Combined Roster", all_results, fill_dark_navy)
    ]:
        ws = wb.create_sheet(title=sheet_title)
        ws.views.sheetView[0].showGridLines = True
        
        ws.merge_cells("A1:P2")
        ws["A1"] = f"STUDENT PERFORMANCE ROSTER — {sheet_title.upper()}"
        ws["A1"].font = font_title
        ws["A1"].fill = dept_fill
        ws["A1"].alignment = align_center
        
        for col_idx, h in enumerate(roster_headers, start=1):
            cell = ws.cell(row=4, column=col_idx, value=h)
            cell.font = font_header
            cell.fill = fill_subhead
            cell.alignment = align_center
            cell.border = border_thin
            
        cur_row = 5
        for s_idx, s in enumerate(dept_students, start=1):
            row_data = [
                s_idx, s["reg_no"], s["name"], s["department"], s["batch"], s["username"] or "—",
                s["total_solved"], s["easy_solved"], s["medium_solved"], s["hard_solved"],
                s["rating"] if s["rating"] > 0 else "—",
                s["global_rank"] if s["global_rank"] > 0 else "—",
                s["w516_status"], s["w516_solved"], s["w515_solved"], s["status"]
            ]
            
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=cur_row, column=col_idx, value=val)
                cell.font = font_bold if col_idx in (1, 2, 7) else font_data
                cell.alignment = align_left if col_idx in (2, 3, 6) else align_center
                cell.border = border_thin
                
                # Highlight active contest solvers
                if col_idx == 7 and s["total_solved"] >= 100:
                    cell.font = font_green
                elif col_idx == 13 and s["w516_status"] == "LIVE":
                    cell.fill = fill_highlight
                    cell.font = font_green
                elif col_idx == 16 and s["status"] != "VERIFIED":
                    cell.fill = fill_warning
                    cell.font = font_red
                elif cur_row % 2 == 0:
                    cell.fill = fill_zebra
            cur_row += 1

    # ==========================================
    # SHEET 7: TOP CODERS LEADERBOARD
    # ==========================================
    ws_top = wb.create_sheet(title="Top Coders Leaderboard")
    ws_top.views.sheetView[0].showGridLines = True
    
    ws_top.merge_cells("A1:I2")
    ws_top["A1"] = "INSTITUTIONAL TOP CODERS & CONTEST STARS"
    ws_top["A1"].font = font_title
    ws_top["A1"].fill = PatternFill(start_color="15803D", end_color="15803D", fill_type="solid")
    ws_top["A1"].alignment = align_center
    
    top_headers = ["Rank", "Register No", "Student Name", "Department", "LeetCode Username", "Total Solved", "Contest Rating", "Global Rank", "Contest 516"]
    for col_idx, h in enumerate(top_headers, start=1):
        cell = ws_top.cell(row=4, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_subhead
        cell.alignment = align_center
        cell.border = border_thin
        
    top_students = [s for s in all_results if s["total_solved"] > 0][:50]
    for r_idx, s in enumerate(top_students, start=1):
        row_data = [
            f"#{r_idx}", s["reg_no"], s["name"], s["department"], s["username"],
            s["total_solved"], s["rating"] if s["rating"] > 0 else "—",
            s["global_rank"] if s["global_rank"] > 0 else "—",
            f"{s['w516_solved']} Q ({s['w516_status']})"
        ]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws_top.cell(row=r_idx + 4, column=col_idx, value=val)
            cell.font = font_bold if col_idx in (1, 6) else font_data
            cell.alignment = align_left if col_idx in (2, 3, 5) else align_center
            cell.border = border_thin
            if r_idx <= 3:
                cell.fill = fill_highlight
                cell.font = font_green
            elif r_idx % 2 == 0:
                cell.fill = fill_zebra

    # Auto-adjust column widths across all sheets
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row in (1, 2):  # skip merged header
                    continue
                val_str = str(cell.value or "")
                if len(val_str) > max_len:
                    max_len = len(val_str)
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
    wb.save(filename)
    print(f"  -> Successfully generated styled Excel workbook: {filename}")
    return filename

def generate_pdf_report(cs_metrics, iot_metrics, all_results, filename):
    print(f"\n📄 [PDF] Generating Institutional PDF Summary: {filename}...")
    doc = SimpleDocTemplate(
        filename,
        pagesize=landscape(A4),
        leftMargin=20,
        rightMargin=20,
        topMargin=20,
        bottomMargin=20
    )
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontSize=14,
        leading=18,
        textColor=colors.HexColor('#0F172A'),
        alignment=1,
        fontName='Helvetica-Bold'
    )
    subtitle_style = ParagraphStyle(
        'DocSubtitle',
        parent=styles['Normal'],
        fontSize=9,
        leading=12,
        textColor=colors.HexColor('#475569'),
        alignment=1,
        fontName='Helvetica'
    )
    cell_style = ParagraphStyle(
        'CellText',
        parent=styles['Normal'],
        fontSize=8,
        leading=10,
        alignment=1,
        fontName='Helvetica'
    )
    cell_bold = ParagraphStyle(
        'CellBold',
        parent=styles['Normal'],
        fontSize=8,
        leading=10,
        alignment=1,
        fontName='Helvetica-Bold'
    )

    story = []
    story.append(Paragraph("<b>NANDHA ENGINEERING COLLEGE (AUTONOMOUS) — ERODE</b>", title_style))
    story.append(Paragraph("<b>DEPARTMENT OF CYBER SECURITY & DEPARTMENT OF INTERNET OF THINGS</b>", title_style))
    story.append(Paragraph(f"Official LeetCode Institutional Performance & Contest Audit Report — {datetime.datetime.now(IST).strftime('%d %B %Y')}", subtitle_style))
    story.append(Spacer(1, 10))
    story.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor('#0284C7'), spaceBefore=2, spaceAfter=8))
    
    # Executive Matrix Table
    matrix_data = [
        [Paragraph("<b>Metric / Indicator</b>", cell_bold), Paragraph("<b>Cyber Security (CSE-CS)</b>", cell_bold), Paragraph("<b>IoT (CSE-IOT)</b>", cell_bold), Paragraph("<b>Combined Total</b>", cell_bold)],
        [Paragraph("Total Students", cell_style), Paragraph(str(cs_metrics['total_students']), cell_style), Paragraph(str(iot_metrics['total_students']), cell_style), Paragraph(str(cs_metrics['total_students'] + iot_metrics['total_students']), cell_bold)],
        [Paragraph("Active Coders (> 0 Solved)", cell_style), Paragraph(str(cs_metrics['active_students']), cell_style), Paragraph(str(iot_metrics['active_students']), cell_style), Paragraph(str(cs_metrics['active_students'] + iot_metrics['active_students']), cell_bold)],
        [Paragraph("Total Problems Solved", cell_style), Paragraph(f"<b>{cs_metrics['total_problems_solved']:,}</b>", cell_style), Paragraph(f"<b>{iot_metrics['total_problems_solved']:,}</b>", cell_style), Paragraph(f"<b>{cs_metrics['total_problems_solved'] + iot_metrics['total_problems_solved']:,}</b>", cell_bold)],
        [Paragraph("Average Solved / Student", cell_style), Paragraph(str(cs_metrics['avg_problems']), cell_style), Paragraph(str(iot_metrics['avg_problems']), cell_style), Paragraph(str(round((cs_metrics['total_problems_solved'] + iot_metrics['total_problems_solved'])/(cs_metrics['total_students'] + iot_metrics['total_students']), 1)), cell_bold)],
        [Paragraph("Above 500 Solved", cell_style), Paragraph(str(cs_metrics['above_500']), cell_style), Paragraph(str(iot_metrics['above_500']), cell_style), Paragraph(str(cs_metrics['above_500'] + iot_metrics['above_500']), cell_bold)],
        [Paragraph("250 - 500 Solved", cell_style), Paragraph(str(cs_metrics['between_250_500']), cell_style), Paragraph(str(iot_metrics['between_250_500']), cell_style), Paragraph(str(cs_metrics['between_250_500'] + iot_metrics['between_250_500']), cell_bold)],
        [Paragraph("100 - 249 Solved", cell_style), Paragraph(str(cs_metrics['between_100_249']), cell_style), Paragraph(str(iot_metrics['between_100_249']), cell_style), Paragraph(str(cs_metrics['between_100_249'] + iot_metrics['between_100_249']), cell_bold)],
        [Paragraph("Weekly Contest 516 Participated", cell_style), Paragraph(str(cs_metrics['w516_participated']), cell_style), Paragraph(str(iot_metrics['w516_participated']), cell_style), Paragraph(str(cs_metrics['w516_participated'] + iot_metrics['w516_participated']), cell_bold)],
    ]
    
    t_matrix = Table(matrix_data, colWidths=[200, 180, 180, 180])
    t_matrix.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(t_matrix)
    story.append(Spacer(1, 14))
    
    # Top Coders Table
    story.append(Paragraph("<b>TOP 15 OVERALL LEETCODE LEADERS (CYBER SECURITY & IOT)</b>", ParagraphStyle('Sub', parent=styles['Heading2'], fontSize=10, leading=12, textColor=colors.HexColor('#0F172A'), fontName='Helvetica-Bold')))
    story.append(Spacer(1, 4))
    
    top_pdf_data = [
        [Paragraph("<b>#</b>", cell_bold), Paragraph("<b>Reg No</b>", cell_bold), Paragraph("<b>Student Name</b>", cell_bold), Paragraph("<b>Dept</b>", cell_bold), Paragraph("<b>Username</b>", cell_bold), Paragraph("<b>Total Solved</b>", cell_bold), Paragraph("<b>Rating</b>", cell_bold), Paragraph("<b>Contest 516</b>", cell_bold)]
    ]
    
    top_15 = [s for s in all_results if s["total_solved"] > 0][:15]
    for idx, s in enumerate(top_15, start=1):
        top_pdf_data.append([
            Paragraph(f"#{idx}", cell_style),
            Paragraph(s["reg_no"], cell_style),
            Paragraph(s["name"][:22], cell_style),
            Paragraph(s["department"], cell_style),
            Paragraph(s["username"][:18], cell_style),
            Paragraph(f"<b>{s['total_solved']}</b>", cell_bold),
            Paragraph(str(s["rating"]) if s["rating"] > 0 else "—", cell_style),
            Paragraph(f"{s['w516_solved']} Q", cell_style)
        ])
        
    t_top = Table(top_pdf_data, colWidths=[30, 90, 180, 70, 150, 80, 60, 80])
    t_top.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0F172A')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    story.append(t_top)
    
    doc.build(story)
    print(f"  -> Successfully generated PDF Report: {filename}")
    return filename

def build_html_email_body(cs_metrics, iot_metrics, all_results):
    cs_tot = cs_metrics.get("total_students", 0)
    iot_tot = iot_metrics.get("total_students", 0)
    comb_tot = cs_tot + iot_tot
    
    cs_act = cs_metrics.get("active_students", 0)
    iot_act = iot_metrics.get("active_students", 0)
    comb_act = cs_act + iot_act
    
    cs_prob = cs_metrics.get("total_problems_solved", 0)
    iot_prob = iot_metrics.get("total_problems_solved", 0)
    comb_prob = cs_prob + iot_prob
    
    top_10 = [s for s in all_results if s["total_solved"] > 0][:10]
    
    top_rows_html = ""
    for idx, s in enumerate(top_10, start=1):
        dept_badge_color = "#4338ca" if s["department"] == "CSE(CS)" else "#0d9488"
        top_rows_html += f"""
        <tr style="border-bottom: 1px solid #e2e8f0;">
            <td style="padding: 8px; font-weight: bold; text-align: center; color: #0f172a;">#{idx}</td>
            <td style="padding: 8px; font-weight: bold; color: #1e293b;">{s['reg_no']}</td>
            <td style="padding: 8px; color: #0f172a;">{s['name']}</td>
            <td style="padding: 8px; text-align: center;"><span style="background: {dept_badge_color}; color: white; padding: 2px 6px; border-radius: 4px; font-size: 11px; font-weight: bold;">{s['department']}</span></td>
            <td style="padding: 8px; color: #0284c7; font-family: monospace;">{s['username']}</td>
            <td style="padding: 8px; text-align: center; font-weight: bold; color: #15803d; font-size: 13px;">{s['total_solved']}</td>
            <td style="padding: 8px; text-align: center; color: #475569;">{s['rating'] if s['rating'] > 0 else '—'}</td>
            <td style="padding: 8px; text-align: center; font-weight: bold; color: #0f172a;">{s['w516_solved']} Q</td>
        </tr>
        """
        
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
    <meta charset="utf-8">
    <style>
        body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Helvetica, Arial, sans-serif; background-color: #f1f5f9; margin: 0; padding: 20px; color: #0f172a; }}
        .card {{ background: #ffffff; border-radius: 10px; max-width: 900px; margin: 0 auto; box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1); overflow: hidden; border: 1px solid #e2e8f0; }}
        .header {{ background: linear-gradient(135deg, #0f172a 0%, #1e293b 100%); color: #ffffff; padding: 28px 24px; text-align: center; }}
        .header h1 {{ margin: 0 0 6px 0; font-size: 20px; letter-spacing: 0.5px; color: #38bdf8; }}
        .header p {{ margin: 3px 0; font-size: 13px; color: #cbd5e1; }}
        .content {{ padding: 24px; }}
        .grid {{ display: flex; gap: 12px; margin-bottom: 24px; }}
        .stat-box {{ flex: 1; background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 8px; padding: 14px; text-align: center; }}
        .stat-box .num {{ font-size: 22px; font-weight: bold; color: #0f172a; margin-top: 4px; }}
        .stat-box .lbl {{ font-size: 11px; text-transform: uppercase; color: #64748b; font-weight: 600; }}
        .table {{ width: 100%; border-collapse: collapse; margin-top: 14px; font-size: 12px; }}
        .table th {{ background: #0f172a; color: #f8fafc; padding: 10px 8px; text-align: center; border: 1px solid #334155; }}
        .table td {{ padding: 8px 6px; border: 1px solid #cbd5e1; }}
        .badge-cs {{ background: #e0e7ff; color: #3730a3; padding: 2px 6px; border-radius: 4px; font-weight: bold; font-size: 11px; }}
        .badge-iot {{ background: #ccfbf1; color: #115e59; padding: 2px 6px; border-radius: 4px; font-weight: bold; font-size: 11px; }}
        .footer {{ background: #f8fafc; border-top: 1px solid #e2e8f0; padding: 16px; text-align: center; font-size: 12px; color: #64748b; }}
    </style>
    </head>
    <body>

    <div class="card">
        <div class="header">
            <h1>NANDHA ENGINEERING COLLEGE (AUTONOMOUS)</h1>
            <p style="font-size: 14px; font-weight: 600; color: #f8fafc;">DEPARTMENT OF CYBER SECURITY & DEPARTMENT OF INTERNET OF THINGS</p>
            <p><strong>Complete Institutional LeetCode & Contest Audit Report</strong></p>
            <p style="color: #94a3b8; font-size: 12px; margin-top: 8px;">Generated Automatically on {datetime.datetime.now(IST).strftime('%A, %d %B %Y %I:%M %p IST')}</p>
        </div>

        <div class="content">
            <h3 style="color: #0f172a; margin-top: 0; border-bottom: 2px solid #0284c7; padding-bottom: 6px;">📊 Institutional Performance Summary</h3>
            
            <table class="table">
                <tr>
                    <th style="text-align: left;">Performance Metric</th>
                    <th>Cyber Security (CSE-CS)</th>
                    <th>IoT (CSE-IOT)</th>
                    <th style="background: #0284c7;">Institutional Total</th>
                </tr>
                <tr>
                    <td style="font-weight: bold;">Total Enrolled Students</td>
                    <td style="text-align: center; font-weight: bold;">{cs_tot}</td>
                    <td style="text-align: center; font-weight: bold;">{iot_tot}</td>
                    <td style="text-align: center; font-weight: bold; background: #f0fdf4;">{comb_tot}</td>
                </tr>
                <tr style="background: #f8fafc;">
                    <td style="font-weight: bold;">Active Students (&gt; 0 Solved)</td>
                    <td style="text-align: center; color: #15803d; font-weight: bold;">{cs_act} ({round(cs_act/cs_tot*100,1)}%)</td>
                    <td style="text-align: center; color: #0d9488; font-weight: bold;">{iot_act} ({round(iot_act/iot_tot*100,1)}%)</td>
                    <td style="text-align: center; font-weight: bold; background: #f0fdf4;">{comb_act} ({round(comb_act/comb_tot*100,1)}%)</td>
                </tr>
                <tr>
                    <td style="font-weight: bold;">Total Problems Solved</td>
                    <td style="text-align: center; font-weight: bold; color: #4338ca;">{cs_prob:,}</td>
                    <td style="text-align: center; font-weight: bold; color: #0d9488;">{iot_prob:,}</td>
                    <td style="text-align: center; font-weight: bold; color: #15803d; background: #f0fdf4; font-size: 13px;">{comb_prob:,}</td>
                </tr>
                <tr style="background: #f8fafc;">
                    <td style="font-weight: bold;">Average Solved / Student</td>
                    <td style="text-align: center;">{cs_metrics.get('avg_problems', 0)}</td>
                    <td style="text-align: center;">{iot_metrics.get('avg_problems', 0)}</td>
                    <td style="text-align: center; font-weight: bold; background: #f0fdf4;">{round(comb_prob/comb_tot, 1)}</td>
                </tr>
                <tr>
                    <td>Elite Coders (&gt; 500 Solved)</td>
                    <td style="text-align: center; color: #15803d; font-weight: bold;">{cs_metrics.get('above_500', 0)}</td>
                    <td style="text-align: center; color: #15803d; font-weight: bold;">{iot_metrics.get('above_500', 0)}</td>
                    <td style="text-align: center; font-weight: bold; background: #f0fdf4;">{cs_metrics.get('above_500', 0) + iot_metrics.get('above_500', 0)}</td>
                </tr>
                <tr style="background: #f8fafc;">
                    <td>Advanced Coders (250 - 500 Solved)</td>
                    <td style="text-align: center;">{cs_metrics.get('between_250_500', 0)}</td>
                    <td style="text-align: center;">{iot_metrics.get('between_250_500', 0)}</td>
                    <td style="text-align: center; font-weight: bold; background: #f0fdf4;">{cs_metrics.get('between_250_500', 0) + iot_metrics.get('between_250_500', 0)}</td>
                </tr>
                <tr>
                    <td>Intermediate (100 - 249 Solved)</td>
                    <td style="text-align: center;">{cs_metrics.get('between_100_249', 0)}</td>
                    <td style="text-align: center;">{iot_metrics.get('between_100_249', 0)}</td>
                    <td style="text-align: center; font-weight: bold; background: #f0fdf4;">{cs_metrics.get('between_100_249', 0) + iot_metrics.get('between_100_249', 0)}</td>
                </tr>
                <tr style="background: #f8fafc;">
                    <td>Beginner (1 - 99 Solved)</td>
                    <td style="text-align: center;">{cs_metrics.get('between_1_99', 0)}</td>
                    <td style="text-align: center;">{iot_metrics.get('between_1_99', 0)}</td>
                    <td style="text-align: center; font-weight: bold; background: #f0fdf4;">{cs_metrics.get('between_1_99', 0) + iot_metrics.get('between_1_99', 0)}</td>
                </tr>
                <tr>
                    <td style="color: #b91c1c;">Not Yet Started (0 Solved)</td>
                    <td style="text-align: center; color: #b91c1c; font-weight: bold;">{cs_metrics.get('not_started', 0)}</td>
                    <td style="text-align: center; color: #b91c1c; font-weight: bold;">{iot_metrics.get('not_started', 0)}</td>
                    <td style="text-align: center; color: #b91c1c; font-weight: bold; background: #fef2f2;">{cs_metrics.get('not_started', 0) + iot_metrics.get('not_started', 0)}</td>
                </tr>
                <tr style="background: #f8fafc;">
                    <td style="font-weight: bold;">Weekly Contest 516 Participations</td>
                    <td style="text-align: center; font-weight: bold; color: #166534;">{cs_metrics.get('w516_participated', 0)}</td>
                    <td style="text-align: center; font-weight: bold; color: #115e59;">{iot_metrics.get('w516_participated', 0)}</td>
                    <td style="text-align: center; font-weight: bold; color: #15803d; background: #f0fdf4;">{cs_metrics.get('w516_participated', 0) + iot_metrics.get('w516_participated', 0)}</td>
                </tr>
            </table>

            <h3 style="color: #0f172a; margin-top: 28px; border-bottom: 2px solid #16a34a; padding-bottom: 6px;">🏆 Top 10 Overall Student Coders (Cyber & IoT)</h3>
            <table class="table">
                <tr>
                    <th>Rank</th>
                    <th>Register No</th>
                    <th>Student Name</th>
                    <th>Department</th>
                    <th>LeetCode Username</th>
                    <th>Total Solved</th>
                    <th>Rating</th>
                    <th>Contest 516</th>
                </tr>
                {top_rows_html}
            </table>
            
            <div style="margin-top: 24px; padding: 14px; background: #eff6ff; border-left: 4px solid #3b82f6; border-radius: 4px; font-size: 13px; color: #1e3a8a;">
                <strong>📎 Report Attachments Included:</strong>
                <ul style="margin: 6px 0 0 0; padding-left: 20px;">
                    <li><strong>Excel Workbook (.xlsx):</strong> Comprehensive 8-sheet matrix with batch breakdowns, individual rosters, contest solves, and attendance.</li>
                    <li><strong>PDF Document (.pdf):</strong> Executive printable summary for Academic Coordinators and HODs.</li>
                </ul>
            </div>
        </div>

        <div class="footer">
            <p><strong>Nandha Engineering College (Autonomous)</strong> — Institutional Coding Intelligence Engine</p>
            <p>This is an automated system email dispatched directly from the server.</p>
        </div>
    </div>

    </body>
    </html>
    """
    return html

def main():
    print("=" * 80)
    print("AUTONOMOUS FULL SCAN & EMAIL DISPATCHER — CYBER SECURITY & IOT DEPARTMENTS")
    print(f"Timestamp: {datetime.datetime.now(IST).strftime('%Y-%m-%d %H:%M:%S IST')}")
    print("=" * 80)
    
    # 1. Load roster
    students = load_student_roster()
    
    # 2. Parallel LeetCode Scan
    all_results = scan_all_students(students)
    
    # 3. Compute Metrics
    cs_metrics = compute_department_metrics(all_results, "CSE(CS)")
    iot_metrics = compute_department_metrics(all_results, "CSE(IOT)")
    
    # 4. Update Database
    update_sqlite_database(all_results)
    
    # 5. Generate Excel
    excel_filename = f"Cyber_and_IoT_Full_LeetCode_Report_{datetime.datetime.now(IST).strftime('%d%b%Y')}.xlsx"
    excel_path = create_styled_excel(cs_metrics, iot_metrics, all_results, excel_filename)
    
    # 6. Generate PDF
    pdf_filename = f"Cyber_and_IoT_Full_LeetCode_Report_{datetime.datetime.now(IST).strftime('%d%b%Y')}.pdf"
    pdf_path = generate_pdf_report(cs_metrics, iot_metrics, all_results, pdf_filename)
    
    # 7. Generate HTML Email
    html_body = build_html_email_body(cs_metrics, iot_metrics, all_results)
    
    with open("cyber_iot_email_preview.html", "w", encoding="utf-8") as f:
        f.write(html_body)
    print(f"  📄 Saved Email Preview: cyber_iot_email_preview.html")
    
    # 8. Email Dispatch
    print(f"\n📧 [EMAIL DISPATCH] Preparing attachments and dispatching email...")
    
    with open(excel_path, "rb") as f:
        excel_bytes = f.read()
    with open(pdf_path, "rb") as f:
        pdf_bytes = f.read()
        
    attachments = [
        (os.path.basename(excel_path), excel_bytes),
        (os.path.basename(pdf_path), pdf_bytes)
    ]
    
    recipients = ["nanthishvaran17@gmail.com"]
    subject = f"LeetCode Performance Full Scan Report — Cyber Security & IoT ({datetime.datetime.now(IST).strftime('%d-%m-%Y')})"
    
    for rec in recipients:
        print(f"  -> Sending report to: {rec}...")
        ok, err = send_email(
            recipient=rec,
            subject=subject,
            html_body=html_body,
            attachments=attachments
        )
        print(f"  -> Dispatch to {rec}: Status={'SUCCESS' if ok else 'FAILED'} | Details={err or 'Accepted'}")
        
    print("\n" + "=" * 80)
    print("✅ COMPLETED FULL CYBER & IOT SCAN, REPORT GENERATION & EMAIL DISPATCH")
    print("=" * 80)

if __name__ == "__main__":
    main()
