"""
report_excel.py — Generates multi-sheet Excel workbook.
Sheet 1: Historical Raw Data
Sheet 2: This Week's Summary
Sheet 3: Rating Chart (embedded line chart)
"""
import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)

# ─── Theme colours ─────────────────────────────────────────────────────────────
NAVY   = "1B2A4A"
GOLD   = "F5B800"
GREEN  = "1DB954"
RED    = "E53935"
LIGHT  = "F0F4FA"
WHITE  = "FFFFFF"
GRAY   = "B0BEC5"

def _hdr(text: str, bold: bool = True, size: int = 11, color: str = WHITE) -> dict:
    return {"value": text, "font": Font(bold=bold, size=size, color=color),
            "fill": PatternFill("solid", fgColor=NAVY),
            "alignment": Alignment(horizontal="center", vertical="center")}

def _cell(ws, row: int, col: int, value, bold=False, color=None, fill=None, align="left"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, color=color or "000000")
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    return c

def _border(ws, min_row, max_row, min_col, max_col):
    thin = Side(style="thin", color=GRAY)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                             min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border


# ─── Sheet 1: Historical Raw Data ──────────────────────────────────────────────

def _sheet_raw(wb: openpyxl.Workbook, history: list[dict], settings: dict) -> None:
    ws = wb.active
    ws.title = "📊 Contest History"
    ws.sheet_properties.tabColor = NAVY

    # Title row
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = f"{settings.get('student', {}).get('name', 'Student')} — LeetCode Contest History"
    title_cell.font = Font(bold=True, size=14, color=WHITE)
    title_cell.fill = PatternFill("solid", fgColor=NAVY)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Headers
    headers = ["#", "Contest", "Date", "Problems Solved", "Rating", "Δ Rating", "Global Rank", "Finish Time"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = Font(bold=True, size=10, color=WHITE)
        c.fill = PatternFill("solid", fgColor="2C3E6B")
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # Data rows
    prev_rating = None
    for idx, row in enumerate(reversed(history), 1):  # oldest first for chart
        r = idx + 2
        rating = row.get("rating") or 0
        delta = round(rating - prev_rating, 1) if prev_rating is not None else 0
        prev_rating = rating

        date_str = datetime.fromtimestamp(
            row.get("contest_start") or 0, tz=timezone.utc
        ).strftime("%d %b %Y")

        bg = LIGHT if idx % 2 == 0 else WHITE
        vals = [
            idx,
            row.get("contest_title", ""),
            date_str,
            f"{row.get('problems_solved') or 0} / {row.get('total_problems') or 4}",
            round(rating, 1),
            delta,
            row.get("ranking") or "",
            _fmt_time(row.get("finish_time_s")),
        ]
        aligns = ["center", "left", "center", "center", "center", "center", "center", "center"]
        for col, (v, a) in enumerate(zip(vals, aligns), 1):
            c = ws.cell(row=r, column=col, value=v)
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal=a, vertical="center")
            if col == 6 and isinstance(v, (int, float)):  # delta
                c.font = Font(bold=True, color=GREEN if v >= 0 else RED)

    # Column widths
    widths = [5, 36, 14, 16, 10, 10, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _border(ws, 2, 2 + len(history), 1, 8)


# ─── Sheet 2: This Week's Summary ──────────────────────────────────────────────

def _sheet_summary(wb: openpyxl.Workbook, analysis: dict, settings: dict) -> None:
    ws = wb.create_sheet("📋 This Week")
    ws.sheet_properties.tabColor = GOLD

    def section_header(row: int, text: str):
        ws.merge_cells(f"A{row}:E{row}")
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(bold=True, size=11, color=WHITE)
        c.fill = PatternFill("solid", fgColor="2C3E6B")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 22

    def kv(row: int, key: str, value, green_if_positive: bool = False):
        ws.cell(row=row, column=1, value=key).font = Font(bold=True, size=10)
        c = ws.cell(row=row, column=2, value=value)
        c.font = Font(size=10)
        if green_if_positive and isinstance(value, (int, float)):
            c.font = Font(size=10, bold=True, color=GREEN if value >= 0 else RED)
        ws.row_dimensions[row].height = 18

    row = 1
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22

    # Title
    ws.merge_cells("A1:E1")
    t = ws.cell(row=1, column=1, value=f"Weekly Report — {analysis.get('latest_title', '')}")
    t.font = Font(bold=True, size=13, color=WHITE)
    t.fill = PatternFill("solid", fgColor=NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    row = 2

    section_header(row, "🏆  Headline Stats"); row += 1
    kv(row, "Current Rating",    round(analysis.get("current_rating") or 0, 1)); row += 1
    kv(row, "Rating Change",     analysis.get("rating_delta"), green_if_positive=True); row += 1
    kv(row, "5-Week Avg Rating", analysis.get("rolling_avg_5")); row += 1
    kv(row, "Global Rank",       f"#{(analysis.get('current_ranking') or 0):,}"); row += 1
    kv(row, "Rank Change",       analysis.get("rank_delta"), green_if_positive=False); row += 1
    kv(row, "Problems Solved",   f"{analysis.get('problems_solved') or 0} / {analysis.get('total_problems') or 4}"); row += 1
    kv(row, "Finish Time",       _fmt_time(analysis.get("finish_time_s"))); row += 1
    kv(row, "Participation Streak", f"{analysis.get('streak', 0)} weeks"); row += 1
    row += 1

    section_header(row, "📉  Tag Weakness Analysis (Weakest First)"); row += 1
    ws.cell(row=row, column=1, value="Topic Tag").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Accuracy %").font = Font(bold=True)
    ws.cell(row=row, column=3, value="Attempted").font = Font(bold=True)
    ws.cell(row=row, column=4, value="Accepted").font = Font(bold=True)
    row += 1
    for tag_row in (analysis.get("weak_tags") or []):
        ws.cell(row=row, column=1, value=tag_row["tag"])
        acc_cell = ws.cell(row=row, column=2, value=f"{tag_row['accuracy']}%")
        acc_cell.font = Font(color=RED if tag_row["accuracy"] < 50 else "000000")
        ws.cell(row=row, column=3, value=tag_row["total"])
        ws.cell(row=row, column=4, value=tag_row["accepted"])
        row += 1
    row += 1

    section_header(row, "✍️  Auto-Generated Performance Narrative"); row += 1
    ws.merge_cells(f"A{row}:E{row + 3}")
    c = ws.cell(row=row, column=1, value=analysis.get("narrative", ""))
    c.alignment = Alignment(wrap_text=True, vertical="top")
    c.font = Font(size=10, italic=True)
    ws.row_dimensions[row].height = 72

    if analysis.get("milestones_crossed"):
        row += 5
        section_header(row, "🎉  Milestones This Week"); row += 1
        for m in analysis["milestones_crossed"]:
            ws.cell(row=row, column=1, value=f"🏅 Rating {m} crossed!").font = Font(bold=True, color=GOLD)
            row += 1


# ─── Sheet 3: Rating Chart ─────────────────────────────────────────────────────

def _sheet_chart(wb: openpyxl.Workbook, history: list[dict]) -> None:
    ws = wb.create_sheet("📈 Rating Chart")
    ws.sheet_properties.tabColor = GREEN

    # Write data for chart (oldest first)
    ws.cell(row=1, column=1, value="Contest").font = Font(bold=True)
    ws.cell(row=1, column=2, value="Rating").font = Font(bold=True)
    for idx, row in enumerate(reversed(history), 2):
        ws.cell(row=idx, column=1, value=row.get("contest_title", f"W{idx-1}")[:20])
        ws.cell(row=idx, column=2, value=round(row.get("rating") or 0, 1))

    n = len(history)
    if n < 2:
        return

    chart = LineChart()
    chart.title = "Rating Over Time"
    chart.style = 10
    chart.y_axis.title = "Rating"
    chart.x_axis.title = "Contest"
    chart.width = 22
    chart.height = 14

    data = Reference(ws, min_col=2, min_row=1, max_row=n + 1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=n + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.line.solidFill = "1B2A4A"
    chart.series[0].graphicalProperties.line.width = 18000  # 2pt

    ws.add_chart(chart, "D2")
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 12


# ─── Public API ────────────────────────────────────────────────────────────────

def _fmt_time(seconds: Optional[int]) -> str:
    if not seconds:
        return "—"
    m, s = divmod(seconds, 60)
    return f"{m}m {s}s"


def generate_excel(
    history: list[dict],
    analysis: dict,
    settings: dict,
    out_dir: Path,
) -> Path:
    wb = openpyxl.Workbook()

    _sheet_raw(wb, history, settings)
    _sheet_summary(wb, analysis, settings)
    _sheet_chart(wb, history)

    # Remove default empty sheet if present (shouldn't be after _sheet_raw sets active)
    # openpyxl creates a default 'Sheet' — we already renamed it in _sheet_raw

    out_dir.mkdir(parents=True, exist_ok=True)
    today = datetime.now().strftime("%Y-%m-%d")
    path = out_dir / f"LeetCode_Report_{today}.xlsx"
    wb.save(path)
    log.info(f"[EXCEL] Saved: {path}")
    return path
